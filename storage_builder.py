"""
storage_builder.py — Builds the Storage & Small Parts Invoice
Mirrors the layout of Sample_Promethean_Storage_Small_Parts_Invoice.xlsx
"""

import json
import shutil
import re as _re
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

TEMPLATE_FILE = Path(__file__).parent / "template" / "Sample_Promethean_Storage_Small_Parts_Invoice.xlsx"
DEFAULT_WHITELIST = Path(__file__).parent / "whitelist_default.json"
STORAGE_PRICES_FILE = Path(__file__).parent / "storage_prices.json"

ACCT_FMT = '_(\"$\"* #,##0.00_);_(\"$\"* \\(#,##0.00\\);_(\"$\"* \"-\"??_);_(@_)'
INT_FMT  = '#,##0'
DATE_FMT = "mm-dd-yy"

TAX_RATE = 0.07

# ── Default pricing (overridable via storage_prices.json) ─────────────────────
DEFAULT_PART_TYPE_PRICES = {
    "PSU":                              7,
    "Mainboard Configure for Dispatch": 52,
    "Mainboard":                        37,
    "AC-PCA":                           4,
    "Keypad":                           4,
    "Maintouch":                        37,
    "EXT-INPUT":                        7,
    "OPS-PCA":                          4,
    "USB":                              4,
    "SPEAKER":                          4,
    "CONSOLE":                          0,
}

DEFAULT_LINE_PRICES = {
    "unit_storage":      8.00,
    "pallet_storage":   23.50,
    "unit_receipt":     15.00,
    "small_part_checkin": 0.77,
    "unit_pick":         8.00,
    "small_part_pick":   8.00,
}

# Part code → type classification keywords
PART_TYPE_MAP = {
    "PSU":                              ["PSU"],
    "Mainboard Configure for Dispatch": ["MAINBOARD", "MAINBRD"],
    "AC-PCA":                           ["AC-PCA"],
    "Keypad":                           ["KEYPAD"],
    "Maintouch":                        ["MAINTOUCHPCA", "MAINTOUCH"],
    "EXT-INPUT":                        ["EXT-INPUT"],
    "OPS-PCA":                          ["OPS-PCA"],
    "USB":                              ["OPS-", "OPS4-", "AP-WIFI", "WIFI", "BT", "-NP", "-7P", "-5P", "-CP", "-C1-"],
    "SPEAKER":                          ["SPEAKER"],
    "CONSOLE":                          ["CONSOLE"],
}

RMA_PATTERN = _re.compile(r'^M\d{8}$', _re.IGNORECASE)   # M + exactly 8 digits, nothing else


# ── Pricing helpers ────────────────────────────────────────────────────────────

def load_prices():
    """Return (part_type_prices, line_prices) — merged from file over defaults."""
    import sys
    part_prices = dict(DEFAULT_PART_TYPE_PRICES)
    line_prices = dict(DEFAULT_LINE_PRICES)
    if STORAGE_PRICES_FILE.exists():
        try:
            data = json.loads(STORAGE_PRICES_FILE.read_text())
            part_prices.update(data.get("part_type_prices", {}))
            line_prices.update(data.get("line_prices", {}))
        except Exception as e:
            # Falling back to defaults here means any price customized via
            # the Config page silently reverts — surface it instead of
            # letting a corrupt storage_prices.json go unnoticed.
            print(f"[storage_builder] WARNING: could not parse {STORAGE_PRICES_FILE} — "
                  f"using default pricing only. Error: {e}", file=sys.stderr)
    return part_prices, line_prices


def save_prices(part_type_prices: dict, line_prices: dict):
    STORAGE_PRICES_FILE.write_text(json.dumps({
        "part_type_prices": part_type_prices,
        "line_prices":      line_prices,
    }, indent=2))


# ── Part classification ────────────────────────────────────────────────────────

def classify_part_type(part_code: str) -> str:
    """Map a part code string to its testing type label. Returns None if unmatched (→ blank in Excel)."""
    if not isinstance(part_code, str) or not part_code.strip():
        return None
    pc = part_code.upper()
    for ptype, keywords in PART_TYPE_MAP.items():
        for kw in keywords:
            if kw in pc:
                return ptype
    return None   # unmatched → leave Type blank, price = 0


# ── CSV cleaning ───────────────────────────────────────────────────────────────

def clean_csv_val(s):
    if isinstance(s, str):
        return s.replace('="', '').replace('"', '').strip()
    return s


def clean_csv_df(df: pd.DataFrame) -> pd.DataFrame:
    # Clean both column names AND cell values — Excel CSV export wraps everything in ="..."
    df = df.copy()
    df.columns = [clean_csv_val(c) for c in df.columns]
    return df.map(clean_csv_val)


# ── Whitelist ──────────────────────────────────────────────────────────────────

def _load_whitelist(whitelist_path=None):
    if whitelist_path and Path(whitelist_path).exists():
        wl = pd.read_excel(whitelist_path, sheet_name=None)
        parts_list = set(wl["Parts"]["Parts List"].dropna().astype(str).str.strip().str.upper())
        units_df = wl["Units"]
        unit_models = set()
        for col in units_df.columns:
            unit_models.update(units_df[col].dropna().astype(str).str.strip().str.upper())
        return parts_list, unit_models
    else:
        with open(DEFAULT_WHITELIST) as f:
            data = json.load(f)
        return (set(p.upper() for p in data["parts"]),
                set(u.upper() for u in data["unit_models"]))


# ══════════════════════════════════════════════════════════════════════════════
# FedEx Master Sheet parsing — shared by the Storage invoice AND the Workshop
# invoice's "Parts Testing & Configuration" section (moved there in v20; see
# PROJECT_BRIEF.md). Keeping this in one place means both invoices always
# render identical Part Testing data from the same source file.
# ══════════════════════════════════════════════════════════════════════════════

def analyze_fedex(fedex_path, period_start, period_end,
                   part_prices, small_part_pick_price, log=print):
    """
    Parse a FedEx Master Sheet and return Part Testing & Configuration data
    for the given billing period.

    Returns a dict with:
      - programming_df:   per-row detail (MSO, dates, part #, type, qty, fees)
      - part_type_totals: {part_type: total_quantity} for Breakdown line items
      - small_part_picks: sum of normalized Quantity values in the period
    """
    fedex = pd.read_excel(fedex_path)
    fedex["Request Date"] = pd.to_datetime(fedex["Request Date"], errors="coerce")
    fedex_period = fedex[
        (fedex["Request Date"] >= pd.Timestamp(period_start)) &
        (fedex["Request Date"] <= pd.Timestamp(period_end))
    ].copy()
    quantity = pd.to_numeric(fedex_period["Quantity"], errors="coerce").fillna(1)
    quantity = quantity.mask(quantity <= 0, 1)
    fractional = quantity[quantity.mod(1) != 0]
    if len(fractional):
        source_rows = ", ".join(str(int(i) + 2) for i in fractional.index[:10])
        raise ValueError(
            "FedEx Quantity must contain whole numbers; fractional value(s) "
            f"found on source row(s): {source_rows}")
    fedex_period["Quantity"] = quantity.astype(int)

    small_part_picks = int(fedex_period["Quantity"].sum())
    unique_msos = fedex_period["MSO"].dropna().nunique()
    log(f"Small Part Picks: {small_part_picks} total part(s) across {unique_msos} MSO order(s)")

    # ALL FedEx rows in period go into Part Testing — unknown types get price 0
    fedex_period["_part_type"] = fedex_period["Part/Component Reported Product Code"].apply(classify_part_type)
    programming_df = fedex_period[[
        "MSO", "Request Date", "Outbound Tracking",
        "Part/Component Reported Product Code", "_part_type", "Serial Number", "Quantity"
    ]].rename(columns={
        "Part/Component Reported Product Code": "Part #",
        "_part_type": "Type",
        "Serial Number": "Serial",
    }).copy()
    programming_df["Individual Part Fee"] = programming_df["Type"].map(
        lambda t: part_prices.get(t, 0) if t else 0
    )
    programming_df["Total Programming Fee"] = (
        programming_df["Individual Part Fee"] * programming_df["Quantity"]
    )
    programming_df["Part Pick Fee"] = small_part_pick_price

    # Aggregate totals by type (for Breakdown) — skip None/unclassified types
    part_type_totals = {}
    for _, row in programming_df.iterrows():
        ptype = row["Type"]
        if not ptype:   # None → unclassified, price=0, skip from Breakdown totals
            continue
        qty = int(row["Quantity"]) if pd.notna(row["Quantity"]) else 1
        part_type_totals[ptype] = part_type_totals.get(ptype, 0) + qty
    log(f"Part Testing rows: {len(programming_df)}, classified types: {list(part_type_totals.keys())}")

    return {
        "programming_df":   programming_df,
        "part_type_totals": part_type_totals,
        "small_part_picks": small_part_picks,
    }


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — analyze_storage(): load files, compute everything, return review lists
# ══════════════════════════════════════════════════════════════════════════════

def analyze_storage(
    inventory_path, shipping_path, receipt_path, fedex_path,
    pallet_count, date_from, date_to,
    whitelist_path=None, log=print,
):
    """
    Load and process all source files. Return a dict with:
      - all computed dataframes (ready to pass to build_storage_invoice)
      - review_spc_no_rma:  small parts with whitelist match but no valid M+8 RMA
      - review_non_whitelist: receipt items not on whitelist (parts OR units)
      - review_unit_recv_non_wl: units_received rows whose model isn't on unit whitelist
    The caller shows these review lists to the user for manual confirmation,
    then calls build_storage_invoice() with the confirmed inclusions.
    """
    part_prices, line_prices = load_prices()

    parts_list, unit_models = _load_whitelist(whitelist_path)
    source = "uploaded" if whitelist_path and Path(whitelist_path).exists() else "built-in default"
    log(f"Whitelist ({source}): {len(parts_list)} parts, {len(unit_models)} unit models")

    # ── Inventory ─────────────────────────────────────────────────────────────
    period_start = date_from.replace(day=1)
    period_end   = date_to
    inv = clean_csv_df(pd.read_csv(inventory_path))
    units_inv = inv[inv["Item Type"].str.strip().str.lower() == "unit"].copy()
    units_inv = units_inv.drop_duplicates(subset=["Serial Number"])
    log(f"Inventory: {len(inv)} rows, {len(units_inv)} unique units")

    # ── Shipping (period) ─────────────────────────────────────────────────────
    ship = clean_csv_df(pd.read_csv(shipping_path))
    # A unit is considered shipped only when "Pickup Date" is filled in
    ship["Pickup Date"]  = pd.to_datetime(ship["Pickup Date"],  format="%m-%d-%Y", errors="coerce")
    ship["Shipped Date"] = pd.to_datetime(ship.get("Shipped Date", pd.Series(dtype="object")),
                                          format="%m-%d-%Y", errors="coerce")
    ship_month = ship[
        ship["Pickup Date"].notna() &
        (ship["Pickup Date"] >= pd.Timestamp(period_start)) &
        (ship["Pickup Date"] <= pd.Timestamp(period_end))
    ].copy().drop_duplicates(subset=["Serial Number"])
    log(f"Shipped this period (Pickup Date filled): {len(ship_month)} units")

    # ── Unit Storage ──────────────────────────────────────────────────────────
    inv_serials  = set(units_inv["Serial Number"].astype(str).str.strip())
    ship_serials = set(ship_month["Serial Number"].astype(str).str.strip())
    storage_rows = []
    for _, r in units_inv.iterrows():
        storage_rows.append({"ActualModel": r["Model"], "Actual Serial": r["Serial Number"],
                              "Storage": line_prices["unit_storage"]})
    for _, r in ship_month.iterrows():
        sn = str(r["Serial Number"]).strip()
        if sn not in inv_serials:
            storage_rows.append({"ActualModel": r["Model"], "Actual Serial": sn,
                                  "Storage": line_prices["unit_storage"]})
    unit_storage_df = pd.DataFrame(storage_rows)
    log(f"Unit Storage: {len(unit_storage_df)} units")

    # ── Receipt log ───────────────────────────────────────────────────────────
    receipt = clean_csv_df(pd.read_csv(receipt_path))
    receipt["Received Date"] = pd.to_datetime(receipt["Received Date"], format="%m-%d-%Y", errors="coerce")
    receipt_period = receipt[
        (receipt["Received Date"] >= pd.Timestamp(period_start)) &
        (receipt["Received Date"] <= pd.Timestamp(period_end))
    ].copy()

    units_received = receipt_period[
        receipt_period["Item Type"].str.strip().str.lower() == "unit"
    ].copy()

    parts_received = receipt_period[
        receipt_period["Item Type"].isna() | (receipt_period["Item Type"].str.strip() == "")
    ].copy()
    log(f"Receipts in period: {len(units_received)} units, {len(parts_received)} parts")

    # ── Small Parts: RMA logic + whitelist review ─────────────────────────────
    # Column that holds RMA: "Serial Number" in receipt log
    # Valid RMA = starts with M, total length 9 chars (M + 8 digits)
    def has_valid_rma(row):
        sn = str(row.get("Serial Number", "") or "").strip()
        return bool(RMA_PATTERN.match(sn))

    parts_received = parts_received.copy()
    parts_received["_has_rma"] = parts_received.apply(has_valid_rma, axis=1)
    parts_received["_on_whitelist"] = parts_received["Model"].apply(
        lambda m: str(m).strip().upper() in parts_list
    )

    # Auto-include ONLY whitelist parts with a strict M+8digit RMA in the Serial column
    spc_auto = parts_received[
        parts_received["_on_whitelist"] & parts_received["_has_rma"]
    ].copy()

    # Whitelist parts with no valid M+8 RMA (blank, non-M, or M with suffix) → Unmatched sheet
    spc_wl_no_rma = parts_received[
        parts_received["_on_whitelist"] & ~parts_received["_has_rma"]
    ].copy()

    # Not on whitelist at all → Unmatched sheet
    spc_non_wl = parts_received[~parts_received["_on_whitelist"]].copy()

    # Units whose model isn't on the unit whitelist → also Unmatched sheet
    units_recv_non_wl = units_received[
        ~units_received["Model"].apply(lambda m: str(m).strip().upper() in unit_models)
    ].copy()

    def _spc_row(r):
        sn = str(r.get("Serial Number", "") or "").strip()
        # spc_auto only contains rows that passed RMA_PATTERN so sn IS the M+8digit RMA
        return {
            "Checkin Date": r["Received Date"],
            "ID": sn,
            "Part #": r.get("Model", ""),
            "Price": line_prices["small_part_checkin"],
            "_src_serial": sn,
        }

    auto_spc_rows = [_spc_row(r) for _, r in spc_auto.iterrows()]

    total_unmatched = len(spc_wl_no_rma) + len(spc_non_wl) + len(units_recv_non_wl)
    log(f"Small Parts Check In: {len(auto_spc_rows)} (whitelist + valid M+8 RMA)")
    log(f"Unmatched sheet: {total_unmatched} items "
        f"({len(spc_wl_no_rma)} WL/no-RMA, {len(spc_non_wl)} non-WL parts, "
        f"{len(units_recv_non_wl)} non-WL units)")

    # Build unmatched DataFrame — ALL original receipt log columns preserved
    # Tag each row with a Source column so user knows where it came from
    def _tag_df(df, source_label):
        out = df.copy()
        out.insert(0, "_Source", source_label)
        # Drop internal helper columns
        for col in ["_has_rma", "_on_whitelist"]:
            if col in out.columns:
                out = out.drop(columns=[col])
        return out

    unmatched_parts = []
    if len(spc_wl_no_rma):
        unmatched_parts.append(_tag_df(spc_wl_no_rma, "WL Part – No Valid RMA"))
    if len(spc_non_wl):
        unmatched_parts.append(_tag_df(spc_non_wl, "Non-Whitelist Part"))
    if len(units_recv_non_wl):
        unmatched_parts.append(_tag_df(units_recv_non_wl, "Non-Whitelist Unit"))

    if unmatched_parts:
        unmatched_df = pd.concat(unmatched_parts, ignore_index=True, sort=False)
    else:
        unmatched_df = pd.DataFrame(columns=["_Source"])

    # ── Unit Picks ────────────────────────────────────────────────────────────
    unit_picks_count = len(ship_month)

    # ── FedEx / Part Testing ─────────────────────────────────────────────────
    fedex_result = analyze_fedex(
        fedex_path, period_start, period_end,
        part_prices, line_prices["small_part_pick"], log=log)
    programming_df   = fedex_result["programming_df"]
    part_type_totals = fedex_result["part_type_totals"]
    small_part_picks = fedex_result["small_part_picks"]

    return {
        "unit_storage_df":   unit_storage_df,
        "units_received":    units_received,
        "programming_df":    programming_df,
        "part_type_totals":  part_type_totals,
        "ship_month":        ship_month,
        "period_start":      period_start,
        "period_end":        period_end,
        "unit_picks_count":  unit_picks_count,
        "small_part_picks":  small_part_picks,
        "pallet_count":      pallet_count,
        "auto_spc_rows":     auto_spc_rows,
        "unmatched_df":      unmatched_df,
    }


def _df_to_review(df, cols, date_cols=None):
    """Convert a DataFrame subset to a list-of-dicts with dates as strings."""
    if df is None or len(df) == 0:
        return []
    available = [c for c in cols if c in df.columns]
    out = []
    for _, r in df[available].iterrows():
        row = {}
        for c in available:
            v = r[c]
            if date_cols and c in date_cols:
                row[c] = v.strftime("%m/%d/%Y") if pd.notna(v) else ""
            elif pd.isna(v) if not isinstance(v, str) else False:
                row[c] = ""
            else:
                row[c] = str(v) if not isinstance(v, (int, float, bool)) else v
        out.append(row)
    return out


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — build_storage_invoice(): write the Excel file
# ══════════════════════════════════════════════════════════════════════════════

def build_storage_invoice(
    # Pass the analysis result dict directly
    analysis: dict,
    invoice_date=None, completed_date=None, call_id="", customer="",
    output_path: Path = None,
    log=print,
):
    _, line_prices = load_prices()

    unit_storage_df  = analysis["unit_storage_df"]
    units_received   = analysis["units_received"].copy()
    programming_df   = analysis["programming_df"]
    ship_month       = analysis["ship_month"]
    unit_picks_count = analysis["unit_picks_count"]
    small_part_picks = analysis["small_part_picks"]
    pallet_count     = analysis["pallet_count"]
    unmatched_df     = analysis.get("unmatched_df")

    # ── Build small parts DataFrame ───────────────────────────────────────────
    all_spc_rows = list(analysis["auto_spc_rows"])

    small_parts_df = pd.DataFrame(all_spc_rows) if all_spc_rows else pd.DataFrame(
        columns=["Checkin Date", "ID", "Part #", "Price"])
    if "_src_serial" in small_parts_df.columns:
        small_parts_df = small_parts_df.drop(columns=["_src_serial"])

    small_parts_count = len(small_parts_df)
    small_parts_total = round(small_parts_df["Price"].sum() if small_parts_count else 0, 2)

    log(f"Building invoice: {len(unit_storage_df)} stored units, "
        f"{len(units_received)} received, {small_parts_count} small parts")
    if unmatched_df is not None and len(unmatched_df):
        log(f"Unmatched sheet: {len(unmatched_df)} items")

    # ── Write workbook ────────────────────────────────────────────────────────
    shutil.copy(TEMPLATE_FILE, output_path)
    wb = load_workbook(output_path)
    wb._external_links = []
    for _ws in wb.worksheets:
        _ws._tables.clear()

    _build_breakdown(wb, invoice_date, completed_date, call_id, customer,
                     unit_storage_df, pallet_count, units_received,
                     small_parts_df, unit_picks_count, small_part_picks,
                     line_prices)
    _build_unit_storage(wb, unit_storage_df, line_prices)
    _build_unit_receiving(wb, units_received, line_prices)
    _build_units_shipped(wb, ship_month, line_prices)
    _build_part_testing(wb, programming_df)
    _build_small_parts(wb, small_parts_df)
    if unmatched_df is not None:
        _build_unmatched(wb, unmatched_df)

    for sheet in wb.worksheets:
        sheet.conditional_formatting._cf_rules.clear()

    wb.save(output_path)
    log(f"Invoice saved → {output_path.name}")

    # Totals
    subtotal = (
        line_prices["unit_storage"]    * len(unit_storage_df) +
        line_prices["pallet_storage"]  * pallet_count +
        line_prices["unit_receipt"]    * len(units_received) +
        small_parts_total +
        line_prices["unit_pick"]       * unit_picks_count +
        line_prices["small_part_pick"] * small_part_picks
    )
    tax   = round(subtotal * TAX_RATE, 2)
    total = round(subtotal + tax, 2)

    return {
        "unit_storage_count":   len(unit_storage_df),
        "pallet_count":         pallet_count,
        "units_received_count": len(units_received),
        "small_parts_count":    small_parts_count,
        "unit_picks_count":     unit_picks_count,
        "small_part_picks":     small_part_picks,
        "subtotal":             subtotal,
        "tax":                  tax,
        "total":                total,
    }


# ══════════════════════════════════════════════════════════════════════════════
# Sheet builders
# ══════════════════════════════════════════════════════════════════════════════

# ── Helpers ───────────────────────────────────────────────────────────────────

def _tf(bold=False, size=10, theme=1):
    c = Color(theme=theme, type="theme")
    return Font(name="Calibri", bold=bold, size=size, color=c)

def _hdr_fill():
    f = PatternFill(fill_type="solid")
    f.fgColor = Color(theme=0, tint=-0.249977111117893, type="theme")
    return f

def _total_fill():
    f = PatternFill(fill_type="solid")
    f.fgColor = Color(theme=4, tint=0.7999816888943144, type="theme")
    return f

def _section_fill():
    f = PatternFill(fill_type="solid")
    f.fgColor = Color(theme=4, tint=0.5999938962981048, type="theme")
    return f

def _add_table(ws, ref, name, style="TableStyleMedium2"):
    """Add an Excel Table to a sheet covering the given ref.

    Some templates (e.g. the Workshop invoice's "Part Testing & Programming"
    sheet) ship with a table of this name already baked in for the Breakdown
    sheet's SUMIFS formulas to reference. openpyxl raises on a duplicate
    table name, so drop any existing same-named table first — we're about
    to fully rebuild that sheet's rows anyway.
    """
    if name in ws.tables:
        del ws.tables[name]
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name=style, showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True)
    ws.add_table(t)

def _write_header_row(ws, headers, row=1, col_widths=None):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, size=11)
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w


# ── Breakdown ─────────────────────────────────────────────────────────────────

def _line(ws, row, label, uom, price, qty_val, qty_formula=None, total_override=None):
    """Write one line item row on the Breakdown sheet.
    - qty_val:      hard-coded quantity (shown until Excel recalculates)
    - qty_formula:  if given, written into col D instead of qty_val (e.g. =COUNTA(...)-1)
    - total_override: if given, written into col E instead of =D*C (e.g. SUMIF for tiered)
    """
    ws.cell(row=row, column=1, value=label).font = _tf(size=10)
    if uom:
        ws.cell(row=row, column=2, value=uom).font = _tf(size=10)
    if price is not None:
        c = ws.cell(row=row, column=3, value=price)
        c.font = _tf(size=10); c.number_format = ACCT_FMT
    # Column D: quantity
    cd = ws.cell(row=row, column=4, value=qty_formula if qty_formula else qty_val)
    cd.font = _tf(size=10); cd.number_format = INT_FMT
    # Column E: total
    if total_override:
        ce = ws.cell(row=row, column=5, value=total_override)
    elif price is not None:
        ce = ws.cell(row=row, column=5, value=f"=D{row}*C{row}")
    else:
        ce = ws.cell(row=row, column=5, value=0)
    ce.font = _tf(size=10); ce.number_format = ACCT_FMT


def _build_breakdown(wb, invoice_date, completed_date, call_id, customer,
                     unit_storage_df, pallet_count, units_received,
                     small_parts_df, unit_picks_count, small_part_picks,
                     line_prices):
    ws = wb["Breakdown"]
    ws.merged_cells.ranges.clear()
    ws.delete_rows(1, ws.max_row)
    for col, w in zip("ABCDE", [44.55, 14.55, 17.44, 16.44, 16.0]):
        ws.column_dimensions[col].width = w

    def addr(row, val, col="A"):
        c = ws.cell(row=row, column=ord(col)-64, value=val)
        c.font = Font(name="Calibri", size=8)
        return c
    def lbl(row, col, val):
        c = ws.cell(row=row, column=ord(col)-64, value=val)
        c.font = Font(name="Calibri", bold=True, size=9)
        c.alignment = Alignment(horizontal="right")
        return c
    def val(row, col, v, fmt=None, as_date=False):
        c = ws.cell(row=row, column=ord(col)-64, value=v)
        c.font = Font(name="Calibri", size=9)
        if as_date: c.number_format = DATE_FMT
        elif fmt:   c.number_format = fmt
        return c

    addr(1, "9145 Ellis Road"); addr(2, "Melbourne, FL 32904")
    a3 = addr(3, "www.ussiglobal.com")
    a3.font = Font(name="Calibri", size=8, underline="single",
                   color=Color(theme=10, type="theme"))

    lbl(1,"D","Call ID:");       val(1,"E", call_id)
    lbl(2,"D","Currency:");      val(2,"E","USD")
    lbl(3,"D","Invoice Date:");  val(3,"E", invoice_date, as_date=True)
    lbl(4,"D","Completed Date:"); val(4,"E", completed_date, as_date=True)
    lbl(5,"D","Customer:");      val(5,"E", customer)
    lbl(6,"D","Terms:");         val(6,"E","Net Amt Due in 45")

    ws.row_dimensions[8].height = 21.0
    ws.merge_cells("A8:E8")
    a8 = ws["A8"]
    a8.value = '=CONCATENATE("PROFORMA INVOICE: ",TEXT(E4,"MMMM YYYY"))'
    a8.font = Font(name="Calibri", bold=True, size=16)
    a8.alignment = Alignment(horizontal="center", vertical="center")

    def section_hdr(row, label):
        ws.row_dimensions[row].height = 15
        for ci, h in enumerate(["","Units of Measure","Unit Price","Quantity","Total Amount"],1):
            c = ws.cell(row=row, column=ci, value=h if ci>1 else label)
            c.fill = _hdr_fill()
            c.font = _tf(bold=True, size=10)
            c.alignment = Alignment(horizontal="center" if ci>1 else None)

    # ── Warehouse Storage ─────────────────────────────────────────────────────
    section_hdr(9, "Warehouse Storage")
    # Unit Storage: qty from live COUNTA of Unit Storage sheet, total = D10*C10
    _line(ws, 10, "Unit Storage",  "Each",   line_prices["unit_storage"],
          len(unit_storage_df),   qty_formula="=COUNTA('Unit Storage'!B:B)-1")
    # Pallet Storage: qty is hard-entered number, total = D11*C11
    _line(ws, 11, "Pallet Storage","Pallet", line_prices["pallet_storage"],
          pallet_count)

    # ── Warehouse Ins & Outs ──────────────────────────────────────────────────
    section_hdr(13, "Warehouse Ins & Outs")
    _line(ws, 14, "Unit Receipt & Processing", "Each",   line_prices["unit_receipt"],
          len(units_received),    qty_formula="=COUNTA('Unit Receiving'!B:B)-1")
    # Small Parts: tiered — qty from COUNTA, total from SUMIF on price col
    _line(ws, 15, "Small Parts Check In", "Tiered", None, len(small_parts_df),
          qty_formula="=COUNTA('Small Parts Check In'!B:B)-1",
          total_override="=SUMIF('Small Parts Check In'!D:D,\">0\",'Small Parts Check In'!D:D)")
    _line(ws, 16, "Unit Picks",      "Each",  line_prices["unit_pick"],
          unit_picks_count,       qty_formula="=COUNTA('Units Shipped'!B:B)-1")
    # Small Part Picks bill by summed Quantity on the retained FedEx detail sheet.
    _line(ws, 17, "Small Part Picks", "Each", line_prices["small_part_pick"],
          small_part_picks,
          qty_formula="=SUM('Part Testing & Programming'!G:G)")

    # ── Totals ────────────────────────────────────────────────────────────────
    ws["D31"].value = "Subtotal"; ws["D31"].font = _tf(bold=True)
    ws["E31"].value = "=SUM(E8:E30)"; ws["E31"].number_format = ACCT_FMT
    ws["D32"].value = "Tax";      ws["D32"].font = _tf(bold=True)
    ws["E32"].value = "=E31*7%";  ws["E32"].number_format = ACCT_FMT
    ws["D33"].value = "Total";    ws["D33"].font = _tf(bold=True)
    ws["E33"].value = "=SUM(E31:E32)"
    ws["E33"].font = _tf(bold=True); ws["E33"].fill = _total_fill()
    ws["E33"].number_format = ACCT_FMT


# ── Unit Storage ──────────────────────────────────────────────────────────────

def _build_unit_storage(wb, df, line_prices):
    ws = wb["Unit Storage"]
    ws.delete_rows(1, ws.max_row)
    headers = ["ActualModel", "Actual Serial", "Storage"]
    widths  = [26, 30, 12]
    _write_header_row(ws, headers, col_widths=widths)
    for i, (_, r) in enumerate(df.iterrows(), 2):
        ws.cell(row=i,column=1,value=r["ActualModel"]).font = _tf(size=11)
        ws.cell(row=i,column=2,value=r["Actual Serial"]).font = _tf(size=11)
        c = ws.cell(row=i,column=3,value=line_prices["unit_storage"])
        c.font = _tf(size=11); c.number_format = ACCT_FMT
    last = max(2, len(df)+1)
    _add_table(ws, f"A1:C{last}", "TblUnitStorage")


# ── Unit Receiving ────────────────────────────────────────────────────────────

def _build_unit_receiving(wb, units_received, line_prices):
    ws = wb["Unit Receiving"]
    ws.delete_rows(1, ws.max_row)
    headers = ["Actual Model", "Actual Serial", "Receipt Fee"]
    widths  = [26, 30, 14]
    _write_header_row(ws, headers, col_widths=widths)
    for i, (_, r) in enumerate(units_received.iterrows(), 2):
        ws.cell(row=i,column=1,value=r.get("Model","")).font = _tf(size=11)
        ws.cell(row=i,column=2,value=r.get("Serial Number","")).font = _tf(size=11)
        c = ws.cell(row=i,column=3,value=line_prices["unit_receipt"])
        c.font = _tf(size=11); c.number_format = ACCT_FMT
    last = max(2, len(units_received)+1)
    _add_table(ws, f"A1:C{last}", "TblUnitReceiving")


# ── Units Shipped ─────────────────────────────────────────────────────────────

def _build_units_shipped(wb, ship_df, line_prices):
    ws = wb["Units Shipped"]
    ws.delete_rows(1, ws.max_row)
    headers = ["MSO","Pickup Date","Model","Serial","Tracking","Sales Order Number","Out Fee"]
    widths  = [14, 12, 26, 30, 14, 18, 12]
    _write_header_row(ws, headers, col_widths=widths)
    for i, (_, r) in enumerate(ship_df.iterrows(), 2):
        vals = [r.get("Ticket Number",""), r.get("Pickup Date",""),
                r.get("Model",""), r.get("Serial Number",""),
                r.get("Tracking Number",""), r.get("Sales Order Number",""),
                line_prices["unit_pick"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=i,column=ci,value=v)
            c.font = _tf(size=11)
            if ci==2 and isinstance(v, pd.Timestamp): c.number_format = DATE_FMT
            if ci==7: c.number_format = ACCT_FMT
    last = max(2, len(ship_df)+1)
    _add_table(ws, f"A1:G{last}", "TblUnitsShipped")


# ── Part Testing & Programming ────────────────────────────────────────────────

def _build_part_testing(wb, prog_df):
    ws = wb["Part Testing & Programming"]
    ws.delete_rows(1, ws.max_row)
    headers = ["MSO","Request Date","Outbound Tracking","Part #",
               "Type","Serial","Quantity","Individual Part Fee",
               "Total Programming Fee","Part Pick Fee"]
    widths  = [14,13,18,30,28,16,10,20,22,14]
    _write_header_row(ws, headers, col_widths=widths)
    for i, (_, r) in enumerate(prog_df.iterrows(), 2):
        vals = [r.get("MSO",""), r.get("Request Date",""),
                r.get("Outbound Tracking",""), r.get("Part #",""),
                r.get("Type",""), r.get("Serial",""),
                r.get("Quantity",1), r.get("Individual Part Fee",0),
                r.get("Total Programming Fee",0), r.get("Part Pick Fee",8)]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=i,column=ci,value=v)
            c.font = _tf(size=11)
            if ci==2 and isinstance(v, pd.Timestamp): c.number_format = DATE_FMT
            if ci in (8,9,10): c.number_format = ACCT_FMT
    last = max(2, len(prog_df)+1)
    _add_table(ws, f"A1:J{last}", "TblPartTesting")


# ── Small Parts Check In ──────────────────────────────────────────────────────

def _build_small_parts(wb, small_parts_df):
    ws = wb["Small Parts Check In"]
    ws.delete_rows(1, ws.max_row)
    headers = ["Checkin Date","ID","Part #","Price"]
    widths  = [14, 20, 30, 12]
    _write_header_row(ws, headers, col_widths=widths)
    for i, (_, r) in enumerate(small_parts_df.iterrows(), 2):
        vals = [r.get("Checkin Date",""), r.get("ID",""),
                r.get("Part #",""), r.get("Price", 0.77)]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=i,column=ci,value=v)
            c.font = _tf(size=11)
            if ci==1 and isinstance(v, pd.Timestamp): c.number_format = DATE_FMT
            if ci==4: c.number_format = ACCT_FMT
    last = max(2, len(small_parts_df)+1)
    _add_table(ws, f"A1:D{last}", "TblSmallParts")

# ── Unmatched Parts & Units (extra sheet) ─────────────────────────────────────

def _build_unmatched(wb, unmatched_df):
    """Write all receipt-log items that didn't match the whitelist to an extra
    sheet — preserving ALL original columns so the user has full context."""
    SHEET_NAME = "Unmatched Parts & Units"

    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(SHEET_NAME)

    if unmatched_df is None or len(unmatched_df) == 0:
        ws.cell(row=1, column=1, value="No unmatched items this period.")
        return

    headers = list(unmatched_df.columns)
    # Auto-width: guess from header length (min 12, max 40)
    for ci, h in enumerate(headers, 1):
        col_letter = get_column_letter(ci)
        ws.column_dimensions[col_letter].width = max(12, min(40, len(str(h)) + 4))
    _write_header_row(ws, headers)

    for i, (_, r) in enumerate(unmatched_df.iterrows(), 2):
        for ci, col in enumerate(headers, 1):
            v = r.get(col, "")
            if v is None or (not isinstance(v, str) and pd.isna(v)):
                v = ""
            c = ws.cell(row=i, column=ci, value=v)
            c.font = _tf(size=11)
            if isinstance(v, pd.Timestamp):
                c.number_format = DATE_FMT

    last = max(2, len(unmatched_df) + 1)
    last_col = get_column_letter(len(headers))
    _add_table(ws, f"A1:{last_col}{last}", "TblUnmatched", style="TableStyleMedium9")
