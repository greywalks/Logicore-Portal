"""
SMS NonConforming — units received at the warehouse that don't conform
(wrong item, damaged, etc.), tracked from intake through resolution.

Backs the "SMS NonConforming" portal tab. Own SQLite database
(nonconforming.db), separate from portal_auth.db and training_planner.db,
following the same per-request-connection-on-flask.g pattern as
portal_auth.py.

Number generation: each filer gets a sequential counter per calendar year,
keyed off their initials (portal_auth.initials_for) — e.g. the 3rd item
Matt Shaw (MS) files in 2026 is "MS26-3". Counters live in their own table
(`number_counters`) rather than being derived by scanning `items` for the
highest existing suffix, so two people filing at the same moment can't both
land on the same number — the increment happens inside one transaction.
"""
import io
import sqlite3
from pathlib import Path
from datetime import datetime, date

from flask import g
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "nonconforming.db"

# Columns a caller may set directly on an item. `date_added`, `number`,
# `filed_by_user_id`, `filed_by_username`, `created_at` are always derived
# server-side and never taken from client input.
EDITABLE_FIELDS = [
    "ticket_no", "model", "serial", "ra_no", "tracking", "carrier",
    "address", "status", "ussi_resolution", "addtl_info",
    "origin_company", "store_no",
]
REQUIRED_FIELDS = ["model", "serial", "carrier"]


# ─────────────────────────────────────────────────────────────────────────
# DB
# ─────────────────────────────────────────────────────────────────────────

def get_db():
    if "_nc_db" not in g:
        db = sqlite3.connect(DB_PATH)
        db.row_factory = sqlite3.Row
        db.execute("PRAGMA foreign_keys = ON")
        g._nc_db = db
    return g._nc_db


def close_db(exception=None):
    db = g.pop("_nc_db", None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date_added TEXT NOT NULL,
            ticket_no TEXT,
            model TEXT NOT NULL,
            serial TEXT NOT NULL,
            number TEXT UNIQUE NOT NULL,
            ra_no TEXT,
            tracking TEXT,
            carrier TEXT NOT NULL,
            address TEXT,
            status TEXT,
            ussi_resolution TEXT,
            addtl_info TEXT,
            origin_company TEXT,
            store_no TEXT,
            filed_by_user_id INTEGER,
            filed_by_username TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS number_counters (
            initials TEXT NOT NULL,
            year INTEGER NOT NULL,
            count INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY (initials, year)
        )
        """
    )
    # Helpful for the search box hitting several columns at once.
    db.execute("CREATE INDEX IF NOT EXISTS idx_items_number ON items(number)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_items_serial ON items(serial)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_items_model ON items(model)")
    db.commit()
    db.close()


# ─────────────────────────────────────────────────────────────────────────
# Number generation
# ─────────────────────────────────────────────────────────────────────────

def generate_number(initials, year=None):
    """Atomically bump and return this initials/year's next sequence number,
    formatted like 'MS26-3'. BEGIN IMMEDIATE takes the write lock up front
    so two concurrent filers under the same initials can't both read the
    same count before either commits."""
    initials = (initials or "XX").strip().upper()
    year = year or date.today().year
    yy = year % 100
    db = get_db()
    db.execute("BEGIN IMMEDIATE")
    try:
        row = db.execute(
            "SELECT count FROM number_counters WHERE initials = ? AND year = ?",
            (initials, year),
        ).fetchone()
        next_count = (row["count"] if row else 0) + 1
        db.execute(
            "INSERT INTO number_counters (initials, year, count) VALUES (?, ?, ?) "
            "ON CONFLICT(initials, year) DO UPDATE SET count = excluded.count",
            (initials, year, next_count),
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return f"{initials}{yy}-{next_count}"


def preview_next_number(initials, year=None):
    """Read-only look at what the *next* number would be, for the form's
    live preview — does not consume a counter slot."""
    initials = (initials or "XX").strip().upper()
    year = year or date.today().year
    db = get_db()
    row = db.execute(
        "SELECT count FROM number_counters WHERE initials = ? AND year = ?",
        (initials, year),
    ).fetchone()
    next_count = (row["count"] if row else 0) + 1
    return f"{initials}{year % 100}-{next_count}"


# ─────────────────────────────────────────────────────────────────────────
# CRUD
# ─────────────────────────────────────────────────────────────────────────

def add_item(fields, user):
    """fields: dict from the form (only EDITABLE_FIELDS are read). user:
    portal_auth current-user row/dict — supplies initials + who-filed."""
    import portal_auth

    missing = [f for f in REQUIRED_FIELDS if not (fields.get(f) or "").strip()]
    if missing:
        raise ValueError(f"Missing required field(s): {', '.join(missing)}")

    initials = portal_auth.initials_for(user)
    today = date.today()
    number = generate_number(initials, today.year)

    row = {k: (fields.get(k) or "").strip() or None for k in EDITABLE_FIELDS}
    db = get_db()
    db.execute(
        f"""
        INSERT INTO items (
            date_added, number, filed_by_user_id, filed_by_username,
            {', '.join(EDITABLE_FIELDS)}
        ) VALUES (?, ?, ?, ?, {', '.join('?' for _ in EDITABLE_FIELDS)})
        """,
        [today.isoformat(), number, user["id"], user["username"]]
        + [row[k] for k in EDITABLE_FIELDS],
    )
    db.commit()
    return get_item(db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])


def get_item(item_id):
    db = get_db()
    row = db.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
    return dict(row) if row else None


def update_item(item_id, fields):
    existing = get_item(item_id)
    if not existing:
        return None
    updates = {k: (fields[k] or "").strip() or None for k in EDITABLE_FIELDS if k in fields}
    if not updates:
        return existing
    missing = [f for f in REQUIRED_FIELDS if f in updates and not updates[f]]
    if missing:
        raise ValueError(f"Missing required field(s): {', '.join(missing)}")
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    db = get_db()
    db.execute(
        f"UPDATE items SET {set_clause}, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        list(updates.values()) + [item_id],
    )
    db.commit()
    return get_item(item_id)


def delete_item(item_id):
    db = get_db()
    cur = db.execute("DELETE FROM items WHERE id = ?", (item_id,))
    db.commit()
    return cur.rowcount > 0


SEARCHABLE_COLUMNS = [
    "number", "ticket_no", "model", "serial", "ra_no", "tracking", "carrier",
    "address", "status", "ussi_resolution", "addtl_info", "origin_company",
    "store_no", "filed_by_username",
]


def list_items(q=None, status=None, limit=500, offset=0):
    """q: free-text search across SEARCHABLE_COLUMNS. status: exact filter.
    Ordered newest-first by date_added, then id, so a filer's own items
    filed the same day stay in filing order."""
    db = get_db()
    where = []
    params = []
    if q:
        like = f"%{q.strip()}%"
        where.append("(" + " OR ".join(f"{c} LIKE ?" for c in SEARCHABLE_COLUMNS) + ")")
        params.extend([like] * len(SEARCHABLE_COLUMNS))
    if status:
        where.append("status = ?")
        params.append(status)
    sql = "SELECT * FROM items"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY date_added DESC, id DESC LIMIT ? OFFSET ?"
    params.extend([limit, offset])
    rows = db.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


def count_items(q=None, status=None):
    db = get_db()
    where = []
    params = []
    if q:
        like = f"%{q.strip()}%"
        where.append("(" + " OR ".join(f"{c} LIKE ?" for c in SEARCHABLE_COLUMNS) + ")")
        params.extend([like] * len(SEARCHABLE_COLUMNS))
    if status:
        where.append("status = ?")
        params.append(status)
    sql = "SELECT COUNT(*) c FROM items"
    if where:
        sql += " WHERE " + " AND ".join(where)
    return db.execute(sql, params).fetchone()["c"]


def distinct_statuses():
    db = get_db()
    rows = db.execute(
        "SELECT DISTINCT status FROM items WHERE status IS NOT NULL AND status != '' ORDER BY status"
    ).fetchall()
    return [r["status"] for r in rows]


# ─────────────────────────────────────────────────────────────────────────
# Export
# ─────────────────────────────────────────────────────────────────────────

EXPORT_COLUMNS = [
    ("date_added", "Date Added"),
    ("number", "Number"),
    ("ticket_no", "Ticket #"),
    ("model", "Model #"),
    ("serial", "Serial #"),
    ("ra_no", "RA #"),
    ("tracking", "Tracking"),
    ("carrier", "Carrier"),
    ("address", "Address"),
    ("status", "Status"),
    ("ussi_resolution", "USSI Resolution Confirmation"),
    ("addtl_info", "Addtl Info"),
    ("origin_company", "Origin Company"),
    ("store_no", "Store #"),
    ("filed_by_username", "Filed By"),
]


def export_xlsx(rows):
    """rows: list of item dicts (as returned by list_items). Returns a
    BytesIO of the finished workbook, ready for send_file()."""
    wb = Workbook()
    ws = wb.active
    ws.title = "NonConforming"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="2F3B4C")
    cell_font = Font(size=10)

    for col_idx, (_, label) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, item in enumerate(rows, start=2):
        for c_idx, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=item.get(key) or "")
            cell.font = cell_font

    for col_idx, (key, label) in enumerate(EXPORT_COLUMNS, start=1):
        width = max(len(label), 12)
        if key in ("address", "addtl_info"):
            width = 38
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────
# Zebra label (ZPL) — 3" x 4" @ 203dpi (609 x 812 dots)
# ─────────────────────────────────────────────────────────────────────────

def build_zpl(item):
    """A simple 3x4 label: big Number (also as a barcode), Model, Serial.
    Kept deliberately plain — easy to hand-tune in Zebra's ZPL if the exact
    layout needs to change later."""
    number = item.get("number") or ""
    model = item.get("model") or ""
    serial = item.get("serial") or ""
    ticket = item.get("ticket_no") or ""

    zpl = f"""^XA
^PW609
^LL812
^CF0,60
^FO40,40^FDSMS NonConforming^FS
^FO40,110^GB529,4,4^FS
^CF0,110
^FO40,150^FD{number}^FS
^BY3,3,120
^FO40,280^BCN,120,Y,N,N^FD{number}^FS
^CF0,40
^FO40,440^FDModel:^FS
^CF0,36
^FO40,480^FD{model}^FS
^CF0,40
^FO40,540^FDSerial:^FS
^CF0,36
^FO40,580^FD{serial}^FS
^CF0,30
^FO40,650^FDTicket #: {ticket}^FS
^XZ"""
    return zpl
