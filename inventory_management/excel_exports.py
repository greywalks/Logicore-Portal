import io
from datetime import datetime

from flask import Response, flash, redirect, request
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from inventory_management.promethean_quality import load_reference
from inventory_management.promethean_quality_v3 import (
    _ensure_override_table,
    _ensure_whitelist_table,
    _reference_path,
    run_inventory_audit,
)
from inventory_management.shipping_history import _shipping_rows, _summary, resolve_range


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True)
SUBTITLE_FONT = Font(size=10, italic=True)
ERROR_FILL = PatternFill("solid", fgColor="FEE2E2")
WARNING_FILL = PatternFill("solid", fgColor="FEF3C7")
REVIEW_FILL = PatternFill("solid", fgColor="E0F2FE")


def _safe_filename(value):
    text = str(value or "").strip()
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in text)
    return safe.strip("_") or "Report"


def _header(ws, row, labels):
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 20


def _autofit(ws, min_width=10, max_width=48):
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = min_width
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def _freeze_filter(ws, header_row):
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{max(header_row, ws.max_row)}"


def _response(workbook, filename):
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(), mimetype=XLSX_MIME, headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def _shipping_workbook(rows, summary, date_range, report_title, model=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipping History"
    ws["A1"] = report_title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Date range: {date_range['start_display']} to {date_range['end_display']}"
    ws["A2"].font = SUBTITLE_FONT
    if model:
        ws["A3"] = f"Model: {model}"
        ws["A3"].font = SUBTITLE_FONT
        summary_row = 5
    else:
        summary_row = 4
    for idx, (label, value) in enumerate([
        ("Shipments", summary["shipments"]), ("Models", summary["models"]),
        ("Unique Serials", summary["serials"]), ("MSOs", summary["msos"]),
    ], start=1):
        ws.cell(row=summary_row, column=idx, value=label).font = HEADER_FONT
        ws.cell(row=summary_row, column=idx).fill = HEADER_FILL
        ws.cell(row=summary_row + 1, column=idx, value=value)
    header_row = summary_row + 3
    _header(ws, header_row, ["Shipped Date", "Model", "Serial Number", "MSO", "Case Number", "Tracking Number", "Sales Order Number", "Pickup Date", "Source File"])
    for r in rows:
        d = r.get("details") or {}
        ws.append([r.get("event_at") or "", r.get("model_number") or "", r.get("serial_number") or "", r.get("mso_number") or "", r.get("case_number") or "", d.get("tracking_number") or "", d.get("sales_order_number") or "", d.get("pickup_date") or "", r.get("original_filename") or ""])
    _freeze_filter(ws, header_row)
    _autofit(ws)
    return wb


def _quality_workbook(audit):
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws["A1"] = "Promethean Current Inventory Audit"
    summary_ws["A1"].font = TITLE_FONT
    summary_ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    summary_ws["A2"].font = SUBTITLE_FONT
    batch = audit.get("batch") or {}
    rows = [
        ("Inventory snapshot", batch.get("original_filename") or "No Inventory Export found"),
        ("Snapshot imported", batch.get("imported_at") or ""),
        ("Inventory rows audited", audit.get("inventory_records", 0)),
        ("Panel units checked", audit.get("units_checked", 0)),
        ("Recognized part rows", audit.get("recognized_parts", 0)),
        ("Whitelisted part/value rows", audit.get("whitelisted_parts", 0)),
        ("Whitelisted model rows", audit.get("whitelisted_models", 0)),
        ("Errors", audit.get("errors", 0)),
        ("Warnings", audit.get("warnings", 0)),
        ("Needs model mapping", audit.get("unmatched_count", 0)),
        ("Serial-specific mappings", len(audit.get("overrides") or [])),
    ]
    _header(summary_ws, 4, ["Metric", "Value"])
    for label, value in rows:
        summary_ws.append([label, value])
    summary_ws["A17"] = "Grade rules"
    summary_ws["A17"].font = Font(bold=True)
    summary_ws["A18"] = "B and R grade units must use the full expected model ending in -NA-R. A stock normally uses the full expected model ending in -NA."
    summary_ws.merge_cells("A18:F19")
    summary_ws["A18"].alignment = Alignment(wrap_text=True, vertical="top")
    _autofit(summary_ws)

    issues_ws = wb.create_sheet("Audit Issues")
    _header(issues_ws, 1, ["Issue Type", "Severity", "Serial Number", "Grade", "Recorded Model", "Expected / Approved Model", "Expectation Source", "Warehouse", "Rack", "Bin", "Item Type", "Source File", "Source Row", "Detail"])
    for item in audit.get("issues") or []:
        issues_ws.append([item.get("issue_type"), item.get("severity"), item.get("serial_number"), item.get("grade"), item.get("model_number"), item.get("expected_model"), item.get("expectation_source"), item.get("warehouse"), item.get("rack"), item.get("bin"), item.get("item_type"), item.get("original_filename"), item.get("source_row_key"), item.get("message")])
        fill = ERROR_FILL if item.get("severity") == "error" else WARNING_FILL
        for cell in issues_ws[issues_ws.max_row]: cell.fill = fill
    _freeze_filter(issues_ws, 1); _autofit(issues_ws)

    mapping_ws = wb.create_sheet("Needs Model Mapping")
    _header(mapping_ws, 1, ["Serial Number", "Grade", "Current Value", "Warehouse", "Rack", "Bin", "Item Type", "Source File", "Source Row", "Reason"])
    for item in audit.get("unmatched") or []:
        mapping_ws.append([item.get("serial_number"), item.get("grade"), item.get("model_number"), item.get("warehouse"), item.get("rack"), item.get("bin"), item.get("item_type"), item.get("original_filename"), item.get("source_row_key"), item.get("message")])
        for cell in mapping_ws[mapping_ws.max_row]: cell.fill = REVIEW_FILL
    _freeze_filter(mapping_ws, 1); _autofit(mapping_ws)

    overrides_ws = wb.create_sheet("Serial Mappings")
    _header(overrides_ws, 1, ["Serial Number", "Approved Model", "Note", "Created", "Updated"])
    for item in audit.get("overrides") or []:
        overrides_ws.append([item.get("serial_number"), item.get("approved_model"), item.get("note") or "", item.get("created_at"), item.get("updated_at")])
    _freeze_filter(overrides_ws, 1); _autofit(overrides_ws)

    whitelist_ws = wb.create_sheet("Global Whitelist")
    _header(whitelist_ws, 1, ["Type", "Approved Value", "Note", "Created", "Updated"])
    whitelist = audit.get("whitelist") or {}
    for kind in ("part", "model"):
        for item in (whitelist.get(kind) or {}).values():
            whitelist_ws.append(["Part / Value" if kind == "part" else "Model", item.get("display_value"), item.get("note") or "", item.get("created_at"), item.get("updated_at")])
    _freeze_filter(whitelist_ws, 1); _autofit(whitelist_ws)
    return wb


def register_excel_exports(app, db_connect, clean, data_dir):
    reference_path = _reference_path(data_dir)

    @app.route("/shipping/model/export.xlsx")
    def shipping_model_export_xlsx():
        date_range = resolve_range(request.args); model = clean(request.args.get("model")); db = db_connect()
        rows = _shipping_rows(db, date_range["start"], date_range["end"], model=model) if model else []; db.close()
        return _response(_shipping_workbook(rows, _summary(rows), date_range, "Model Shipping History", model=model), f"Shipping_History_{_safe_filename(model)}_{date_range['start']}_to_{date_range['end']}.xlsx")

    @app.route("/shipping/all/export.xlsx")
    def shipping_all_export_xlsx():
        date_range = resolve_range(request.args); db = db_connect(); rows = _shipping_rows(db, date_range["start"], date_range["end"]); db.close()
        return _response(_shipping_workbook(rows, _summary(rows), date_range, "All Shipping History"), f"Shipping_History_All_Models_{date_range['start']}_to_{date_range['end']}.xlsx")

    @app.route("/quality/export.xlsx")
    def quality_export_xlsx():
        reference = load_reference(data_dir)
        if not reference:
            flash("Install the Promethean reference data before exporting an audit.", "error")
            return redirect("quality")
        db = db_connect(); _ensure_override_table(db); _ensure_whitelist_table(db)
        audit = run_inventory_audit(db, reference, reference_path); db.close()
        batch = audit.get("batch") or {}; snapshot = _safe_filename(batch.get("original_filename") or "Current_Inventory")
        return _response(_quality_workbook(audit), f"Promethean_Current_Inventory_Audit_{snapshot}.xlsx")
