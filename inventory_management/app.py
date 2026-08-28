import csv
import hashlib
import io
import json
import os
import re
import sqlite3
import uuid
from datetime import date, datetime
from pathlib import Path

from flask import Flask, Response, flash, redirect, render_template, request
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

import portal_auth

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("INVENTORY_DATA_DIR", BASE_DIR / "data"))
UPLOAD_DIR = DATA_DIR / "uploads"
DB_PATH = DATA_DIR / "inventory.db"
SECTION_KEY = "tbd2"  # Legacy permission key retained so existing portal grants keep working.

app = Flask(__name__, template_folder="templates", static_folder=None)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024
app.secret_key = portal_auth.load_or_create_secret()


def _ensure_dirs():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def db_connect():
    _ensure_dirs()
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys = ON")
    return db


def init_db():
    db = db_connect()
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS import_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_name TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            stored_filename TEXT NOT NULL,
            file_hash TEXT NOT NULL UNIQUE,
            imported_at TEXT NOT NULL,
            rows_read INTEGER NOT NULL DEFAULT 0,
            events_added INTEGER NOT NULL DEFAULT 0,
            duplicates_skipped INTEGER NOT NULL DEFAULT 0,
            warnings_json TEXT NOT NULL DEFAULT '[]'
        );

        CREATE TABLE IF NOT EXISTS assets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            serial_number TEXT NOT NULL,
            normalized_serial TEXT NOT NULL UNIQUE,
            last_model TEXT,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS cases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mso_number TEXT NOT NULL,
            normalized_mso TEXT NOT NULL UNIQUE,
            case_number TEXT
        );

        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            asset_id INTEGER REFERENCES assets(id),
            case_id INTEGER REFERENCES cases(id),
            event_type TEXT NOT NULL,
            event_at TEXT,
            model_number TEXT,
            source_type TEXT NOT NULL,
            batch_id INTEGER NOT NULL REFERENCES import_batches(id) ON DELETE CASCADE,
            source_row_key TEXT NOT NULL,
            fingerprint TEXT NOT NULL UNIQUE,
            details_json TEXT NOT NULL DEFAULT '{}'
        );

        CREATE INDEX IF NOT EXISTS idx_assets_serial ON assets(normalized_serial);
        CREATE INDEX IF NOT EXISTS idx_cases_mso ON cases(normalized_mso);
        CREATE INDEX IF NOT EXISTS idx_events_asset ON events(asset_id, event_at);
        CREATE INDEX IF NOT EXISTS idx_events_case ON events(case_id, event_at);
        CREATE INDEX IF NOT EXISTS idx_events_model ON events(model_number);
        CREATE INDEX IF NOT EXISTS idx_events_type ON events(event_type);
        CREATE INDEX IF NOT EXISTS idx_events_date ON events(event_at);
        """
    )
    db.commit()
    db.close()


@app.before_request
def require_access():
    user = portal_auth.get_current_user()
    if not user:
        next_url = (request.environ.get("SCRIPT_NAME", "") or "") + request.path
        return redirect(f"/login?next={next_url}")
    if not portal_auth.has_access(user, SECTION_KEY):
        flash("You don't have permission to view Inventory Management.", "error")
        return redirect("/")


@app.teardown_appcontext
def close_auth_db(exc=None):
    portal_auth.close_db(exc)


@app.context_processor
def nav_context():
    user = portal_auth.get_current_user()
    return {
        "auth_user": user,
        "can_training_tracker": portal_auth.get_role(user, "training-tracker") is not None,
        "script_root": request.environ.get("SCRIPT_NAME", "") or "",
    }


def clean(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = str(value).strip()
    m = re.fullmatch(r'=\"(.*)\"', text, re.S)
    if m:
        text = m.group(1)
    return text.strip()


def norm(value):
    return re.sub(r"\s+", "", clean(value)).upper()


def normalize_mso(value):
    value = norm(value)
    m = re.fullmatch(r"(M\d+)R", value)
    return m.group(1) if m else value


def iso_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat(sep=" ")
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat(sep=" ")
    text = clean(value)
    for fmt in ("%m-%d-%Y", "%m/%d/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).isoformat(sep=" ")
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).replace(microsecond=0).isoformat(sep=" ")
    except ValueError:
        return text or None


def display_date(value):
    if not value:
        return "Unknown date"
    try:
        dt = datetime.fromisoformat(value)
        if dt.hour == 0 and dt.minute == 0 and dt.second == 0:
            return dt.strftime("%b %d, %Y")
        return dt.strftime("%b %d, %Y · %I:%M %p")
    except Exception:
        return value


app.jinja_env.filters["event_date"] = display_date


def get_or_create_asset(db, serial, model=None):
    n = norm(serial)
    if not n:
        return None
    row = db.execute("SELECT id FROM assets WHERE normalized_serial = ?", (n,)).fetchone()
    now = datetime.utcnow().replace(microsecond=0).isoformat(sep=" ")
    if row:
        if clean(model):
            db.execute("UPDATE assets SET last_model = ?, updated_at = ? WHERE id = ?", (clean(model), now, row["id"]))
        return row["id"]
    cur = db.execute(
        "INSERT INTO assets(serial_number, normalized_serial, last_model, updated_at) VALUES(?,?,?,?)",
        (clean(serial), n, clean(model) or None, now),
    )
    return cur.lastrowid


def get_or_create_case(db, mso, case_number=None):
    n = normalize_mso(mso)
    if not n:
        return None
    row = db.execute("SELECT id FROM cases WHERE normalized_mso = ?", (n,)).fetchone()
    if row:
        if clean(case_number):
            db.execute("UPDATE cases SET case_number = COALESCE(case_number, ?) WHERE id = ?", (clean(case_number), row["id"]))
        return row["id"]
    cur = db.execute(
        "INSERT INTO cases(mso_number, normalized_mso, case_number) VALUES(?,?,?)",
        (n, n, clean(case_number) or None),
    )
    return cur.lastrowid


def event_fingerprint(source_type, event_type, serial, mso, event_at, model, details):
    stable = {
        "source": source_type,
        "type": event_type,
        "serial": norm(serial),
        "mso": normalize_mso(mso),
        "at": event_at or "",
        "model": norm(model),
        "details": {k: clean(v) for k, v in sorted(details.items()) if clean(v)},
    }
    return hashlib.sha256(json.dumps(stable, sort_keys=True, separators=(",", ":")).encode()).hexdigest()


def add_event(db, batch_id, source_type, row_key, event_type, event_at=None, serial=None, mso=None, model=None, case_number=None, details=None):
    details = details or {}
    if not any([clean(serial), clean(mso), clean(model)]):
        return 0
    asset_id = get_or_create_asset(db, serial, model) if clean(serial) else None
    case_id = get_or_create_case(db, mso, case_number) if clean(mso) else None
    fp = event_fingerprint(source_type, event_type, serial, mso, event_at, model, details)
    try:
        db.execute(
            """INSERT INTO events(asset_id, case_id, event_type, event_at, model_number, source_type,
               batch_id, source_row_key, fingerprint, details_json) VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (asset_id, case_id, event_type, event_at, clean(model) or None, source_type,
             batch_id, row_key, fp, json.dumps({k: clean(v) for k, v in details.items()}, sort_keys=True)),
        )
        return 1
    except sqlite3.IntegrityError:
        return 0


def source_kind(headers):
    h = {clean(x) for x in headers if clean(x)}
    if {"Received Date", "Model", "Serial Number", "RMA In Tracking"}.issubset(h):
        return "receiving"
    if {"Ticket Number", "Shipped Date", "Model", "Serial Number", "Tracking Number"}.issubset(h):
        return "shipping"
    if {"Transfer Detail ID", "Model", "Serial Number", "Rack", "Warehouse"}.issubset(h):
        return "inventory"
    if {"Timestamp", "Actual Model", "Actual Serial", "Result"}.issubset(h):
        return "repair"
    if {"MSO", "Serial Number", "Outbound Tracking", "Request Date"}.issubset(h):
        return "fedex"
    return None


def dict_rows_from_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        kind = source_kind(headers)
        if not kind:
            raise ValueError("CSV columns do not match a supported Receiving, Shipping, or Inventory export.")
        yield kind, path.name, list(reader)


def dict_rows_from_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    found = False
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            continue
        headers = [clean(v) for v in header]
        kind = source_kind(headers)
        if not kind:
            continue
        found = True
        records = []
        for values in rows:
            if not any(v is not None and clean(v) for v in values):
                continue
            records.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]})
        yield kind, ws.title, records
    if not found:
        raise ValueError("Workbook does not contain a recognized Receiving, Shipping, Inventory, Repair Data, or FedEx sheet.")


def import_record(db, batch_id, kind, row_key, r, observed_at):
    added = 0
    if kind == "receiving":
        mso = r.get("RMA In Tracking") or r.get("Receive Pallet / PGR Tracking")
        details = {"grade": r.get("Grade"), "item_type": r.get("Item Type"), "rma_in_tracking": r.get("RMA In Tracking"), "pgr_tracking": r.get("Receive Pallet / PGR Tracking")}
        added += add_event(db, batch_id, kind, row_key, "received", iso_date(r.get("Received Date")), r.get("Serial Number"), mso, r.get("Model"), details=details)
    elif kind == "shipping":
        details = {"tracking_number": r.get("Tracking Number"), "sales_order_number": r.get("Sales Order Number"), "pickup_date": r.get("Pickup Date")}
        added += add_event(db, batch_id, kind, row_key, "shipped", iso_date(r.get("Shipped Date")), r.get("Serial Number"), r.get("Ticket Number"), r.get("Model"), details=details)
        if clean(r.get("Pickup Date")):
            added += add_event(db, batch_id, kind, row_key, "picked_up", iso_date(r.get("Pickup Date")), r.get("Serial Number"), r.get("Ticket Number"), r.get("Model"), details={"tracking_number": r.get("Tracking Number")})
    elif kind == "inventory":
        details = {"transfer_detail_id": r.get("Transfer Detail ID"), "grade": r.get("Grade"), "rack": r.get("Rack"), "bin": r.get("Bin"), "warehouse": r.get("Warehouse"), "item_type": r.get("Item Type"), "source_received_date": r.get("Received Date")}
        added += add_event(db, batch_id, kind, row_key, "inventory_observed", observed_at, r.get("Serial Number"), None, r.get("Model"), details=details)
    elif kind == "repair":
        result = clean(r.get("Result"))
        rl = result.lower()
        event_type = "repair_pending_parts" if "pending" in rl and "part" in rl else ("repair_scrapped" if "scrap" in rl else "repair")
        details = {"box_model": r.get("Box Model"), "box_serial": r.get("Box Serial"), "result": result, "category": r.get("Category"), "most_recent_ship_date": r.get("Most Recent Ship Date Per Serial")}
        serial = r.get("Actual Serial") or r.get("Box Serial")
        model = r.get("Actual Model") or r.get("Box Model")
        added += add_event(db, batch_id, kind, row_key, event_type, iso_date(r.get("Timestamp")), serial, None, model, details=details)
    elif kind == "fedex":
        serial = r.get("Serial Number") or r.get("Used Serial Number")
        mso = r.get("MSO")
        case_number = r.get("Case Number")
        model = r.get("Part/Component Reported Product Code") or r.get("Part/Component Used Product Code")
        base = {"case_number": case_number, "reported_product_code": r.get("Part/Component Reported Product Code"), "used_product_code": r.get("Part/Component Used Product Code"), "used_serial_number": r.get("Used Serial Number"), "quantity": r.get("Quantity"), "service_company": r.get("Service Company"), "outbound_tracking": r.get("Outbound Tracking"), "return_tracking": r.get("Return Tracking")}
        if clean(r.get("Request Date")):
            added += add_event(db, batch_id, kind, row_key, "fedex_requested", iso_date(r.get("Request Date")), serial, mso, model, case_number, base)
        if clean(r.get("Outbound Tracking")) and not clean(r.get("Fedex Ship Date")):
            added += add_event(db, batch_id, kind, row_key, "fedex_outbound_created", iso_date(r.get("Request Date")), serial, mso, model, case_number, base)
        if clean(r.get("Fedex Ship Date")):
            added += add_event(db, batch_id, kind, row_key, "fedex_outbound_shipped", iso_date(r.get("Fedex Ship Date")), serial, mso, model, case_number, {**base, "outbound_status": r.get("Outbound Status")})
        if clean(r.get("Outbound Delivery Date")):
            added += add_event(db, batch_id, kind, row_key, "fedex_outbound_delivered", iso_date(r.get("Outbound Delivery Date")), serial, mso, model, case_number, {**base, "outbound_status": r.get("Outbound Status")})
        if clean(r.get("Return Tracking")):
            added += add_event(db, batch_id, kind, row_key, "fedex_return_created", iso_date(r.get("Request Date")), serial, mso, model, case_number, base)
        if clean(r.get("Return Tracking Ship Date")):
            added += add_event(db, batch_id, kind, row_key, "fedex_return_shipped", iso_date(r.get("Return Tracking Ship Date")), serial, mso, model, case_number, {**base, "return_status": r.get("Return Status")})
        if clean(r.get("Return Tracking Delivery Date")):
            added += add_event(db, batch_id, kind, row_key, "fedex_return_delivered", iso_date(r.get("Return Tracking Delivery Date")), serial, mso, model, case_number, {**base, "return_status": r.get("Return Status")})
    return added


def import_file(file_storage):
    _ensure_dirs()
    original = secure_filename(file_storage.filename or "")
    if not original:
        raise ValueError("Uploaded file has no filename.")
    ext = Path(original).suffix.lower()
    if ext not in {".csv", ".xlsx", ".xlsm"}:
        raise ValueError(f"Unsupported file type: {ext or 'unknown'}")
    raw = file_storage.read()
    if not raw:
        raise ValueError(f"{original} is empty.")
    digest = hashlib.sha256(raw).hexdigest()
    db = db_connect()
    existing = db.execute("SELECT * FROM import_batches WHERE file_hash = ?", (digest,)).fetchone()
    if existing:
        db.close()
        return {"filename": original, "duplicate_file": True, "rows": existing["rows_read"], "events": 0, "skipped": existing["events_added"]}

    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    stored = f"{stamp}_{uuid.uuid4().hex[:8]}_{original}"
    path = UPLOAD_DIR / stored
    path.write_bytes(raw)
    imported_at = datetime.utcnow().replace(microsecond=0).isoformat(sep=" ")
    cur = db.execute(
        "INSERT INTO import_batches(source_name, original_filename, stored_filename, file_hash, imported_at) VALUES(?,?,?,?,?)",
        ("pending", original, stored, digest, imported_at),
    )
    batch_id = cur.lastrowid
    rows_read = events_added = 0
    sources = []
    try:
        groups = dict_rows_from_csv(path) if ext == ".csv" else dict_rows_from_xlsx(path)
        for kind, sheet, records in groups:
            sources.append(kind)
            for idx, record in enumerate(records, start=2):
                rows_read += 1
                events_added += import_record(db, batch_id, kind, f"{sheet}:{idx}", record, imported_at)
        if not sources:
            raise ValueError("No supported data was found.")
        db.execute(
            "UPDATE import_batches SET source_name=?, rows_read=?, events_added=?, duplicates_skipped=? WHERE id=?",
            (", ".join(sorted(set(sources))), rows_read, events_added, max(0, rows_read - events_added), batch_id),
        )
        db.commit()
    except Exception:
        db.rollback()
        db.execute("DELETE FROM import_batches WHERE id = ?", (batch_id,))
        db.commit()
        try:
            path.unlink()
        except OSError:
            pass
        db.close()
        raise
    db.close()
    return {"filename": original, "duplicate_file": False, "source": ", ".join(sorted(set(sources))), "rows": rows_read, "events": events_added}


def event_rows(db, where, params):
    rows = db.execute(
        f"""SELECT e.*, a.serial_number, a.normalized_serial, c.mso_number, c.case_number,
                   b.original_filename, b.imported_at
            FROM events e
            LEFT JOIN assets a ON a.id=e.asset_id
            LEFT JOIN cases c ON c.id=e.case_id
            JOIN import_batches b ON b.id=e.batch_id
            WHERE {where}
            ORDER BY COALESCE(e.event_at, b.imported_at) ASC, e.id ASC""",
        params,
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["details"] = json.loads(d.pop("details_json") or "{}")
        result.append(d)
    return result


@app.route("/")
def index():
    db = db_connect()
    stats = {
        "assets": db.execute("SELECT COUNT(*) FROM assets").fetchone()[0],
        "events": db.execute("SELECT COUNT(*) FROM events").fetchone()[0],
        "cases": db.execute("SELECT COUNT(*) FROM cases").fetchone()[0],
        "imports": db.execute("SELECT COUNT(*) FROM import_batches").fetchone()[0],
    }
    latest = db.execute("SELECT * FROM import_batches ORDER BY id DESC LIMIT 5").fetchall()
    db.close()
    return render_template("index.html", stats=stats, latest=latest)


@app.route("/search")
def search():
    q = clean(request.args.get("q"))
    kind = clean(request.args.get("kind")) or "auto"
    if not q:
        return redirect("./")
    db = db_connect()
    serial = db.execute("SELECT serial_number FROM assets WHERE normalized_serial = ?", (norm(q),)).fetchone()
    mso = db.execute("SELECT mso_number FROM cases WHERE normalized_mso = ?", (normalize_mso(q),)).fetchone()
    db.close()
    if kind in {"auto", "serial"} and serial:
        return redirect(f"serial/{serial['serial_number']}")
    if kind in {"auto", "mso"} and mso:
        return redirect(f"mso/{mso['mso_number']}")
    return redirect(f"model/{q}")


@app.route("/serial/<path:serial>")
def serial_detail(serial):
    db = db_connect()
    asset = db.execute("SELECT * FROM assets WHERE normalized_serial = ?", (norm(serial),)).fetchone()
    if not asset:
        db.close()
        return render_template("not_found.html", query=serial, kind="Serial Number"), 404
    events = event_rows(db, "e.asset_id = ?", (asset["id"],))
    models = sorted({e["model_number"] for e in events if e["model_number"]})
    msos = sorted({e["mso_number"] for e in events if e["mso_number"]})
    inventory = next((e for e in reversed(events) if e["event_type"] == "inventory_observed"), None)
    db.close()
    return render_template("serial.html", asset=asset, events=events, models=models, msos=msos, inventory=inventory)


@app.route("/mso/<path:mso>")
def mso_detail(mso):
    db = db_connect()
    case = db.execute("SELECT * FROM cases WHERE normalized_mso = ?", (normalize_mso(mso),)).fetchone()
    if not case:
        db.close()
        return render_template("not_found.html", query=mso, kind="MSO"), 404
    events = event_rows(db, "e.case_id = ?", (case["id"],))
    serials = sorted({e["serial_number"] for e in events if e["serial_number"]})
    db.close()
    return render_template("mso.html", case=case, events=events, serials=serials)


@app.route("/model/<path:model>")
def model_detail(model):
    db = db_connect()
    pattern = f"%{clean(model).upper()}%"
    rows = event_rows(db, "UPPER(COALESCE(e.model_number,'')) LIKE ?", (pattern,))
    db.close()
    return render_template("model.html", model=model, events=rows[:2000], total=len(rows))


@app.route("/model/<path:model>/export.csv")
def model_export(model):
    db = db_connect()
    rows = event_rows(db, "UPPER(COALESCE(e.model_number,'')) LIKE ?", (f"%{clean(model).upper()}%",))
    db.close()
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["Event Date", "Event Type", "Model", "Serial Number", "MSO", "Source", "Source File"])
    for e in rows:
        writer.writerow([e["event_at"], e["event_type"], e["model_number"], e["serial_number"], e["mso_number"], e["source_type"], e["original_filename"]])
    return Response(out.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f'attachment; filename="Inventory_Model_{secure_filename(model)}.csv"'})


@app.route("/import", methods=["GET", "POST"])
def import_data():
    results = []
    if request.method == "POST":
        files = [f for f in request.files.getlist("files") if f and f.filename]
        if not files:
            flash("Choose at least one CSV or Excel file.", "error")
        else:
            for f in files:
                try:
                    results.append(import_file(f))
                except Exception as exc:
                    results.append({"filename": f.filename, "error": str(exc)})
    return render_template("import.html", results=results, data_dir=str(DATA_DIR))


@app.route("/imports")
def imports():
    db = db_connect()
    batches = db.execute("SELECT * FROM import_batches ORDER BY id DESC LIMIT 250").fetchall()
    db.close()
    return render_template("imports.html", batches=batches, data_dir=str(DATA_DIR))


init_db()
