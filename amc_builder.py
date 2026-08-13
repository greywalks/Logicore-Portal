"""
amc_builder.py — Builds the AMC Warehouse Invoice
Mirrors the layout of 062026_AMC_Warehouse_Invoice.xlsx (Breakdown / Received / Shipped
tabs), following the same pattern as builder.py / storage_builder.py / philips_builder.py.

Input each month: three raw warehouse-system exports —
  - Receiving Export  (Receiving Export sheet): every unit ever received, one row per
    receipt event. Filtered to the billing period by Received Date for the invoice.
  - Shipping Export   (Shipping Export sheet): every unit ever shipped out, one row per
    ship event. Filtered to the billing period by Shipped Date for the invoice.
  - Inventory Export  (Inventory Export sheet): a point-in-time snapshot of everything
    CURRENTLY on hand. Not date-filtered — like TCL Pallet Storage / Philips Demo sq ft,
    ongoing storage bills on everything on the shelf right now, regardless of when it
    arrived.

Reference data (bundled, not uploaded monthly): amc_dimensions_default.json
  - {MODEL: sq_ft}  — box footprint per TV model, used only for the Storage calc.

Billing rules (mirrors 062026_AMC_Warehouse_Invoice.xlsx, confirmed Aug 2026):
  - Unit Receipt & Processing = count of Receiving Export rows in the billing period,
    $8/each (flat, not sq-ft driven).
  - Storage First 500 sq ft: flat $1,910, qty 1, always billed.
  - Storage Additional Sq Ft, in 50-sq-ft increments = SUM(Dimensions sq ft for every
    unit currently in the Inventory Export snapshot) minus the 500 sq ft credit, rounded
    to the nearest 50, billed at $3.50/unit (same convention as Philips: the "quantity"
    column is the raw sq-ft figure rounded to the nearest 50, not a count of 50-sq-ft
    blocks — see _line() below).
  - Order Fee ($10/each) + Order Out Fee ($8/each) = count of Shipping Export rows in
    the billing period, one of each per shipped row.
  - Tax: 7%.

Models missing a Dimensions entry are flagged (excluded from the sq-ft total) rather
than silently priced at 0 — the review step lets the user supply a value, which is
saved to the live Dimensions store for next time.
"""

import json
import shutil
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils import get_column_letter

DEFAULT_DIMENSIONS_FILE = Path(__file__).parent / "amc_dimensions_default.json"  # bundled seed
DIMENSIONS_FILE          = Path(__file__).parent / "amc_dimensions.json"          # live, editable store
AMC_PRICES_FILE          = Path(__file__).parent / "amc_prices.json"              # live, editable store

ACCT_FMT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
INT_FMT  = '#,##0'
NUM_FMT  = '#,##0.00'
DATE_FMT = "mm-dd-yy"

TAX_RATE  = 0.07
SQFT_UNIT = 50   # billed in 50-sq-ft increments (rounding granularity, not a price)

DEFAULT_PRICES = {
    "unit_receipt":   8.00,     # Unit Receipt & Processing, $/each
    "storage_base":   1910.00,  # Storage, First 500 sq ft — flat
    "base_sqft":      500,      # sq ft included in the flat storage_base charge
    "storage_addl":   3.50,     # Storage, Additional Sq Ft — $ per sq ft (50-ft increments, see _line())
    "order_fee":      10.00,    # Order Fee, $/each shipped
    "order_out_fee":  8.00,     # Order Out Fee, $/each shipped
}


# ══════════════════════════════════════════════════════════════════════════════
# Reference data (Dimensions) — persisted, editable, uploadable
# ══════════════════════════════════════════════════════════════════════════════

def _seed_dimensions():
    return json.loads(DEFAULT_DIMENSIONS_FILE.read_text())


def load_dimensions() -> dict:
    """{MODEL: sq_ft}, upper-cased keys. Seeds from the bundled default on first use."""
    if DIMENSIONS_FILE.exists():
        data = json.loads(DIMENSIONS_FILE.read_text())
    else:
        data = _seed_dimensions()
        save_dimensions(data)
    return {k.upper(): v for k, v in data.items()}


def save_dimensions(dims: dict):
    DIMENSIONS_FILE.write_text(json.dumps(dims, indent=2, sort_keys=True))


def add_dimension(model: str, sqft: float):
    """Append/overwrite one model's footprint and persist immediately."""
    dims = load_dimensions()
    dims[_normalize_model(model)] = float(sqft)
    save_dimensions(dims)
    return dims


def replace_dimensions_from_rows(rows):
    """rows: iterable of (model, sqft) pairs — used when a Dimensions workbook is
    re-uploaded. Replaces the entire store (this becomes the new source of truth)."""
    dims = {}
    for model, sqft in rows:
        if model is None or sqft is None:
            continue
        dims[_normalize_model(model)] = float(sqft)
    save_dimensions(dims)
    return dims


def _normalize_model(model):
    if model is None:
        return ""
    return str(model).strip().upper()


def lookup_sqft(model, dims):
    """Look up box footprint for a model. Exact match only — AMC model codes don't carry
    the same suffix variance Philips does, so no fuzzy stripping here."""
    m = _normalize_model(model)
    if not m:
        return None
    return dims.get(m)


def parse_dimensions_workbook(path):
    """Read an uploaded Dimensions workbook (Model / Sq Ft columns, any sheet name,
    header row auto-detected) and return a list of (model, sqft) pairs."""
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    model_col = next((i for i, h in enumerate(header) if "model" in h), 0)
    sqft_col  = next((i for i, h in enumerate(header) if "sq" in h or "footage" in h or "footprint" in h), 1)
    out = []
    for row in rows[1:]:
        if row is None or len(row) <= max(model_col, sqft_col):
            continue
        model, sqft = row[model_col], row[sqft_col]
        if model is None or sqft is None:
            continue
        try:
            out.append((str(model).strip(), float(sqft)))
        except (TypeError, ValueError):
            continue
    return out


def write_dimensions_workbook(path, dims: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dimensions"
    ws["A1"] = "Model"; ws["B1"] = "Sq Ft"
    ws["A1"].font = _tf(bold=True); ws["B1"].font = _tf(bold=True)
    for i, model in enumerate(sorted(dims), 2):
        ws.cell(row=i, column=1, value=model).font = _tf(size=10)
        ws.cell(row=i, column=2, value=dims[model]).font = _tf(size=10)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    wb.save(path)


# ══════════════════════════════════════════════════════════════════════════════
# Pricing — persisted, editable
# ══════════════════════════════════════════════════════════════════════════════

def load_prices() -> dict:
    import sys
    prices = dict(DEFAULT_PRICES)
    if AMC_PRICES_FILE.exists():
        try:
            prices.update(json.loads(AMC_PRICES_FILE.read_text()))
        except Exception as e:
            print(f"[amc_builder] WARNING: could not parse {AMC_PRICES_FILE} — "
                  f"using default pricing only. Error: {e}", file=sys.stderr)
    return prices


def save_prices(prices: dict):
    AMC_PRICES_FILE.write_text(json.dumps(prices, indent=2))


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — analyze_amc(): load the three raw exports, compute everything
# ══════════════════════════════════════════════════════════════════════════════

def _read_export(path, sheet_name):
    return pd.read_excel(path, sheet_name=sheet_name)


def analyze_amc(receiving_path, shipping_path, inventory_path,
                 period_start, period_end, dims=None, prices=None, log=print):
    if dims is None:
        dims = load_dimensions()
    if prices is None:
        prices = load_prices()

    start = pd.to_datetime(period_start)
    end   = pd.to_datetime(period_end)

    recv = _read_export(receiving_path, "Receiving Export")
    ship = _read_export(shipping_path, "Shipping Export")
    inv  = _read_export(inventory_path, "Inventory Export")
    log(f"Loaded exports — Receiving: {len(recv)}, Shipping: {len(ship)}, Inventory: {len(inv)}")

    # ── Receiving: filter to billing period by Received Date ────────────────────
    recv = recv.copy()
    recv["_date"] = pd.to_datetime(recv["Received Date"], format="%m-%d-%Y", errors="coerce")
    if recv["_date"].isna().any():
        recv.loc[recv["_date"].isna(), "_date"] = pd.to_datetime(
            recv.loc[recv["_date"].isna(), "Received Date"], errors="coerce")
    recv_period = recv[(recv["_date"] >= start) & (recv["_date"] <= end)].copy()
    receipt_count = len(recv_period)
    log(f"Receiving in period {period_start} → {period_end}: {receipt_count} unit(s)")

    # ── Shipping: filter to billing period by Shipped Date ───────────────────────
    ship = ship.copy()
    ship["_date"] = pd.to_datetime(ship["Shipped Date"], format="%m-%d-%Y", errors="coerce")
    if ship["_date"].isna().any():
        ship.loc[ship["_date"].isna(), "_date"] = pd.to_datetime(
            ship.loc[ship["_date"].isna(), "Shipped Date"], errors="coerce")
    ship_period = ship[(ship["_date"] >= start) & (ship["_date"] <= end)].copy()
    ship_count = len(ship_period)
    log(f"Shipping in period {period_start} → {period_end}: {ship_count} unit(s)")

    # ── Storage: ALL units currently on hand (Inventory snapshot), not date-filtered ──
    inv = inv.copy()
    dup = inv["Serial Number"].duplicated()
    if dup.any():
        log(f"WARNING: {int(dup.sum())} duplicate serial(s) in Inventory Export — deduped")
        inv = inv.drop_duplicates(subset=["Serial Number"])

    inv["_sqft"] = inv["Model"].apply(lambda m: lookup_sqft(m, dims))
    missing_df = inv[inv["_sqft"].isna()]
    missing_models = sorted(set(str(m) for m in missing_df["Model"].tolist()))
    if missing_models:
        log(f"WARNING: {len(missing_df)} inventory row(s) across {len(missing_models)} "
            f"model(s) had no Dimensions match — excluded from sq ft total: {missing_models[:10]}")

    total_sqft = float(inv["_sqft"].dropna().sum())
    base_sqft  = float(prices.get("base_sqft", DEFAULT_PRICES["base_sqft"]))
    additional_sqft = max(total_sqft - base_sqft, 0.0)
    additional_qty  = round(additional_sqft / SQFT_UNIT) * SQFT_UNIT

    log(f"Storage — {len(inv)} unit(s) on hand, {total_sqft:.2f} sq ft total, "
        f"{base_sqft:.0f} sq ft credit applied → {additional_qty:.0f} billable "
        f"(rounded to nearest {SQFT_UNIT})")

    excluded_rows = [
        {"Source Tab": "Inventory", "Reason": "No Dimensions match",
         "Model": r.get("Model"), "Serial": r.get("Serial Number"),
         "Rack": r.get("Rack"), "Bin": r.get("Bin")}
        for _, r in missing_df.iterrows()
    ]

    return {
        "receiving_df": recv_period, "shipping_df": ship_period, "inventory_df": inv,
        "receipt_count": receipt_count, "ship_count": ship_count,
        "total_sqft": total_sqft, "base_sqft": base_sqft, "additional_sqft": additional_qty,
        "missing_dimension_models": missing_models,
        "excluded_df": pd.DataFrame(excluded_rows, columns=["Source Tab", "Reason", "Model", "Serial", "Rack", "Bin"]),
        "period_start": period_start, "period_end": period_end,
    }


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — build_amc_invoice(): write the Excel file
# ══════════════════════════════════════════════════════════════════════════════

def _tf(bold=False, size=10):
    return Font(name="Calibri", bold=bold, size=size)

def _hdr_fill():
    f = PatternFill(fill_type="solid")
    f.fgColor = Color(theme=0, tint=-0.25, type="theme")
    return f

def _total_fill():
    f = PatternFill(fill_type="solid")
    f.fgColor = Color(theme=4, tint=0.8, type="theme")
    return f

def _write_header_row(ws, headers, row=1, col_widths=None):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = _tf(bold=True, size=11)
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

def _line(ws, row, label, uom, price, qty_val, qty_formula=None, total_formula=None):
    ws.cell(row=row, column=1, value=label).font = _tf(size=10)
    if uom:
        ws.cell(row=row, column=2, value=uom).font = _tf(size=10)
    if price is not None:
        c = ws.cell(row=row, column=3, value=price)
        c.font = _tf(size=10); c.number_format = ACCT_FMT
    cd = ws.cell(row=row, column=4, value=qty_formula if qty_formula else qty_val)
    cd.font = _tf(size=10); cd.number_format = INT_FMT
    ce = ws.cell(row=row, column=5, value=total_formula if total_formula else f"=D{row}*C{row}")
    ce.font = _tf(size=10); ce.number_format = ACCT_FMT


def build_amc_invoice(analysis, invoice_title, output_path, prices=None, log=print):
    if prices is None:
        prices = load_prices()

    recv = analysis["receiving_df"]
    ship = analysis["shipping_df"]
    receipt_count = analysis["receipt_count"]
    ship_count    = analysis["ship_count"]
    additional_qty = analysis["additional_sqft"]

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Breakdown")
    for col, w in zip("ABCDE", [44.55, 24, 14, 14, 16]):
        ws.column_dimensions[col].width = w

    ws["A1"] = "9145 Ellis Road"
    ws["A2"] = "Melbourne, FL 32904"
    ws["A3"] = "www.ussiglobal.com"
    for r in (1, 2, 3):
        ws.cell(row=r, column=1).font = _tf(size=8)

    ws["A5"] = invoice_title
    ws["A5"].font = _tf(bold=True, size=14)

    def section_hdr(row, label):
        for ci, h in enumerate(["", "Units of Measure", "Unit Price", "Quantity", "Total Amount"], 1):
            c = ws.cell(row=row, column=ci, value=h if ci > 1 else label)
            c.fill = _hdr_fill()
            c.font = _tf(bold=True, size=10)
            c.alignment = Alignment(horizontal="center" if ci > 1 else None)

    row = 7
    section_hdr(row, "Warehouse Storage, Ins, and Outs"); row += 1
    _line(ws, row, "Unit Receipt & Processing", "Each", prices["unit_receipt"], receipt_count,
          # receipt_count = len(recv_period): every row that passed the period
          # date filter, regardless of whether Model is populated. COUNTA must
          # therefore anchor on a column that's guaranteed non-blank for every
          # such row — "Receive Date" (col C), which is exactly what the filter
          # ran on — not "Model" (col A), which can be blank on a given row
          # while the date isn't, silently undercounting vs. receipt_count.
          qty_formula="=COUNTA(Received!C2:C1048576)"); row += 1
    _line(ws, row, "Storage First 500sq Ft.", "Each", prices["storage_base"], 1); row += 1
    _line(ws, row, "Storage Addition Sq Ft (50ft Increments)", "Each", prices["storage_addl"], additional_qty); row += 2

    _line(ws, row, "Order Fee", "Each", prices["order_fee"], ship_count,
          qty_formula="=COUNTA(Shipped!A2:A1048576)"); row += 1
    _line(ws, row, "Order Out Fee", "Each", prices["order_out_fee"], ship_count,
          qty_formula="=COUNTA(Shipped!A2:A1048576)"); row += 2

    last_line = row - 1
    ws.cell(row=row, column=4, value="Subtotal").font = _tf(bold=True)
    ws.cell(row=row, column=5, value=f"=SUM(E8:E{last_line})").number_format = ACCT_FMT
    ws.cell(row=row, column=5).font = _tf(bold=True)
    subtotal_row = row; row += 1
    ws.cell(row=row, column=4, value="Tax").font = _tf(bold=True)
    ws.cell(row=row, column=5, value=f"=E{subtotal_row}*7%").number_format = ACCT_FMT
    ws.cell(row=row, column=5).font = _tf(bold=True)
    tax_row = row; row += 1
    ws.cell(row=row, column=4, value="Total").font = _tf(bold=True)
    tc = ws.cell(row=row, column=5, value=f"=SUM(E{subtotal_row}:E{tax_row})")
    tc.number_format = ACCT_FMT; tc.font = _tf(bold=True); tc.fill = _total_fill()

    # ── Raw data tabs ─────────────────────────────────────────────────────────
    _write_received_tab(wb, recv, prices["unit_receipt"])
    _write_shipped_tab(wb, ship, prices["order_fee"], prices["order_out_fee"])
    _write_excluded_tab(wb, analysis.get("excluded_df"))

    wb.save(output_path)
    log(f"Invoice saved → {Path(output_path).name}")

    subtotal = (
        receipt_count * prices["unit_receipt"]
        + prices["storage_base"]
        + additional_qty * prices["storage_addl"]
        + ship_count * prices["order_fee"]
        + ship_count * prices["order_out_fee"]
    )
    tax = round(subtotal * TAX_RATE, 2)
    total = round(subtotal + tax, 2)
    return {"subtotal": round(subtotal, 2), "tax": tax, "total": total}


def _write_received_tab(wb, recv, charge):
    ws = wb.create_sheet("Received")
    headers = ["Model", "Serial", "Receive Date", "Charge"]
    _write_header_row(ws, headers, col_widths=[22, 20, 18, 10])
    for i, (_, r) in enumerate(recv.iterrows(), 2):
        vals = [r.get("Model"), r.get("Serial Number"), r.get("_date"), charge]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=v)
            c.font = _tf(size=10)
            if ci == 3 and isinstance(v, (pd.Timestamp, datetime)):
                c.number_format = DATE_FMT
            if ci == 4:
                c.number_format = ACCT_FMT


def _write_shipped_tab(wb, ship, order_fee, out_fee):
    ws = wb.create_sheet("Shipped")
    headers = ["Shipped Date", "Ticket #", "Model", "Serial Number", "Tracking Number",
               "Return Tracking", "Order Fee", "Out Fee"]
    _write_header_row(ws, headers, col_widths=[14, 16, 22, 20, 20, 20, 10, 10])
    for i, (_, r) in enumerate(ship.iterrows(), 2):
        vals = [r.get("_date"), r.get("Ticket Number"), r.get("Model"), r.get("Serial Number"),
                r.get("Tracking Number"), r.get("Return Tracking"), order_fee, out_fee]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=v)
            c.font = _tf(size=10)
            if ci == 1 and isinstance(v, (pd.Timestamp, datetime)):
                c.number_format = DATE_FMT
            if ci in (7, 8):
                c.number_format = ACCT_FMT


def _write_excluded_tab(wb, excluded_df):
    """Inventory rows with no Dimensions match — excluded from the sq-ft total so
    nothing silently disappears from the invoice."""
    ws = wb.create_sheet("Excluded Items")
    headers = ["Source Tab", "Reason", "Model", "Serial", "Rack", "Bin"]
    _write_header_row(ws, headers, col_widths=[12, 24, 22, 20, 12, 12])
    if excluded_df is None or len(excluded_df) == 0:
        ws.cell(row=2, column=1, value="No excluded items this period.").font = _tf(size=10)
        return
    for i, (_, r) in enumerate(excluded_df.iterrows(), 2):
        for ci, col in enumerate(headers, 1):
            ws.cell(row=i, column=ci, value=r.get(col, "")).font = _tf(size=10)
