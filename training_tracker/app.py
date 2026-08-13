import os
import sys
import io
import re
import socket
import secrets
import threading
import time
import webbrowser
import sqlite3
import zipfile
from functools import wraps
from datetime import datetime, date

from flask import (
    Flask, render_template, request, redirect, url_for, flash, g, jsonify, session,
    send_file, abort
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

# ---------------------------------------------------------------------------
# Paths — aware of running as a normal script vs. a PyInstaller-bundled .exe.
# When frozen, templates are extracted to a temp folder (sys._MEIPASS) but
# the database must live next to the .exe itself so it survives restarts
# and can be found/backed-up by whoever is hosting it.
# ---------------------------------------------------------------------------

IS_FROZEN = getattr(sys, "frozen", False)

if IS_FROZEN:
    BASE_DIR = os.path.dirname(sys.executable)
    TEMPLATE_DIR = os.path.join(getattr(sys, "_MEIPASS", BASE_DIR), "templates")
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    TEMPLATE_DIR = os.path.join(BASE_DIR, "templates")

DB_PATH = os.path.join(BASE_DIR, "training_planner.db")
HOST = "0.0.0.0"          # listen on all network interfaces so others on the LAN can reach it
PORT = int(os.environ.get("PORT", 5050))

# When hosted on a platform with a persistent disk (Render, etc.), set the
# DATA_DIR environment variable to that disk's mount path so the database
# and secret key survive redeploys instead of living in the ephemeral code
# directory. Falls back to BASE_DIR for local/LAN use, unchanged from before.
DATA_DIR = os.environ.get("DATA_DIR", BASE_DIR)
os.makedirs(DATA_DIR, exist_ok=True)
DB_PATH = os.path.join(DATA_DIR, "training_planner.db")

# Scanned/photographed physical sign-off sheets — stored on the same
# persistent disk as the database so they survive redeploys. One file per
# week; re-uploading replaces the previous one.
SIGNOFF_UPLOAD_DIR = os.path.join(DATA_DIR, "signoff_uploads")
os.makedirs(SIGNOFF_UPLOAD_DIR, exist_ok=True)
ALLOWED_SIGNOFF_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "gif", "webp", "heic"}

# static_folder=None: this app is mounted inside the Logicore Portal at
# /training-tracker/ (see portal-level app.py), and its templates reference
# the portal's own shared assets directly via root-relative "/static/..."
# paths (logo, fonts) rather than this app's own static endpoint — it has
# no static assets of its own, so there's nothing to serve here.
app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=None)
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024  # 20MB — generous for a phone-photo scan


def friendly(d, month_format="%B"):
    """Format a date/datetime as 'Month D, YYYY' (or 'Mon D, YYYY') without
    relying on the %-d / %#d strftime flags, which aren't portable across
    platforms (Linux/macOS use %-d, Windows uses %#d — neither works
    everywhere)."""
    return f"{d.strftime(month_format)} {d.day}, {d.year}"
SECRET_KEY_PATH = os.path.join(DATA_DIR, ".secret_key")


def _load_or_create_secret_key():
    if os.path.exists(SECRET_KEY_PATH):
        with open(SECRET_KEY_PATH, "r") as f:
            key = f.read().strip()
            if key:
                return key
    key = secrets.token_hex(32)
    try:
        with open(SECRET_KEY_PATH, "w") as f:
            f.write(key)
    except Exception:
        pass
    return key


app.config["SECRET_KEY"] = _load_or_create_secret_key()


# ---------------------------------------------------------------------------
# Editable content: every user-facing string on every page lives here as a
# key -> default value. Admins can override any of them from /admin without
# touching code; whatever isn't overridden falls back to these defaults.
# ---------------------------------------------------------------------------

CONTENT_SCHEMA = [
    {
        "category": "Branding",
        "fields": [
            {"key": "site.name", "label": "Site name", "type": "text", "default": "Training Ledger"},
            {"key": "site.tagline", "label": "Header tagline badge", "type": "text", "default": "weekly curriculum"},
            {"key": "site.footer", "label": "Footer line", "type": "text", "default": "Kept like a ledger — every week logged, every signature accounted for."},
            {"key": "nav.weeks", "label": "Nav link: Weeks", "type": "text", "default": "Weeks"},
            {"key": "nav.roster", "label": "Nav link: Roster", "type": "text", "default": "Roster"},
            {"key": "nav.admin", "label": "Nav link: Edit content", "type": "text", "default": "Edit Content"},
        ],
    },
    {
        "category": "Dashboard (weeks list)",
        "fields": [
            {"key": "index.heading", "label": "Page heading", "type": "text", "default": "The Curriculum, Week by Week"},
            {"key": "index.entries_label", "label": "Entry count label (singular)", "type": "text", "default": "entry"},
            {"key": "index.entries_label_plural", "label": "Entry count label (plural)", "type": "text", "default": "entries"},
            {"key": "index.roster_count_label", "label": "Roster count suffix", "type": "text", "default": "on roster"},
            {"key": "index.new_week_btn", "label": "\u201cNew week\u201d button", "type": "text", "default": "+ New Week"},
            {"key": "index.empty_title", "label": "Empty state title", "type": "text", "default": "No entries yet"},
            {"key": "index.empty_body", "label": "Empty state body", "type": "text", "default": "Start the ledger with your first training week."},
            {"key": "index.topics_label", "label": "Topic count suffix", "type": "text", "default": "topics"},
            {"key": "index.signed_off_label", "label": "Sign-off progress suffix", "type": "text", "default": "signed off"},
            {"key": "index.dialog_title", "label": "New week dialog title", "type": "text", "default": "New Week"},
            {"key": "index.field_title", "label": "Field label: Title", "type": "text", "default": "Title"},
            {"key": "index.field_title_placeholder", "label": "Title placeholder", "type": "text", "default": "e.g. Customer Discovery Fundamentals"},
            {"key": "index.field_date", "label": "Field label: Start date", "type": "text", "default": "Start date"},
            {"key": "index.cancel_btn", "label": "Cancel button", "type": "text", "default": "Cancel"},
            {"key": "index.create_btn", "label": "Create button", "type": "text", "default": "Create"},
        ],
    },
    {
        "category": "Week detail",
        "fields": [
            {"key": "week.back_link", "label": "Back link", "type": "text", "default": "All weeks"},
            {"key": "week.edit_btn", "label": "Edit week button", "type": "text", "default": "Edit Week"},
            {"key": "week.delete_btn", "label": "Delete week button", "type": "text", "default": "Delete"},
            {"key": "week.delete_confirm", "label": "Delete week confirmation", "type": "text", "default": "Delete this entire week? This cannot be undone."},
            {"key": "week.topics_heading", "label": "Topics section heading", "type": "text", "default": "Topics & Lesson Plans"},
            {"key": "week.add_topic_btn", "label": "Add topic button", "type": "text", "default": "+ Add Topic"},
            {"key": "week.no_topics", "label": "No-topics message", "type": "text", "default": "No topics yet. Add the first one for this week."},
            {"key": "week.edit_topic_btn", "label": "Edit topic link", "type": "text", "default": "Edit"},
            {"key": "week.delete_topic_btn", "label": "Delete topic link", "type": "text", "default": "Delete"},
            {"key": "week.delete_topic_confirm", "label": "Delete topic confirmation", "type": "text", "default": "Delete this topic?"},
            {"key": "week.key_points_label", "label": "Key points label", "type": "text", "default": "Key points to cover"},
            {"key": "week.videos_label", "label": "Videos label", "type": "text", "default": "Linked videos"},
            {"key": "week.video_title_placeholder", "label": "Video title placeholder", "type": "text", "default": "Video label (optional)"},
            {"key": "week.video_url_placeholder", "label": "Video URL placeholder", "type": "text", "default": "https://..."},
            {"key": "week.link_btn", "label": "Link video button", "type": "text", "default": "Link"},
            {"key": "week.attendance_heading", "label": "Attendance heading", "type": "text", "default": "Attendance"},
            {"key": "week.attendance_subheading", "label": "Attendance subheading", "type": "text", "default": "Check who attended, then send the sign-off sheet"},
            {"key": "week.no_roster", "label": "No-roster message", "type": "text", "default": "No one on the roster yet."},
            {"key": "week.add_people_link", "label": "\u201cAdd people\u201d link", "type": "text", "default": "Add people"},
            {"key": "week.signed_badge", "label": "\u201cSigned\u201d badge", "type": "text", "default": "Signed"},
            {"key": "week.pending_badge", "label": "\u201cPending\u201d badge", "type": "text", "default": "Pending"},
            {"key": "week.signoff_btn", "label": "Open sign-off sheet button", "type": "text", "default": "Open Sign-off Sheet"},
            {"key": "week.new_topic_dialog_title", "label": "New topic dialog title", "type": "text", "default": "New Topic"},
            {"key": "week.topic_title_label", "label": "Topic title field label", "type": "text", "default": "Title"},
            {"key": "week.topic_title_placeholder", "label": "Topic title placeholder", "type": "text", "default": "e.g. Objection Handling"},
            {"key": "week.lesson_plan_label", "label": "Lesson plan field label", "type": "text", "default": "Lesson plan / description"},
            {"key": "week.lesson_plan_placeholder", "label": "Lesson plan placeholder", "type": "text", "default": "What will this session cover?"},
            {"key": "week.key_points_input_label", "label": "Key points field label", "type": "text", "default": "Key points (one per line)"},
            {"key": "week.key_points_placeholder", "label": "Key points placeholder", "type": "text", "default": "One key point per line"},
            {"key": "week.add_topic_submit", "label": "Add topic submit button", "type": "text", "default": "Add Topic"},
            {"key": "week.edit_week_dialog_title", "label": "Edit week dialog title", "type": "text", "default": "Edit Week"},
            {"key": "week.notes_label", "label": "Notes field label", "type": "text", "default": "Notes"},
            {"key": "week.save_btn", "label": "Save button", "type": "text", "default": "Save"},
            {"key": "week.cancel_btn", "label": "Cancel button", "type": "text", "default": "Cancel"},
        ],
    },
    {
        "category": "Sessions",
        "fields": [
            {"key": "session.heading", "label": "Sessions section heading", "type": "text", "default": "Sessions"},
            {"key": "session.subheading", "label": "Sessions subheading", "type": "text", "default": "Each trainer running this week gets their own attendance, digital sign-off, and physical sign-off sheet"},
            {"key": "session.new_btn", "label": "\u201cNew session\u201d button", "type": "text", "default": "+ New Session"},
            {"key": "session.no_sessions", "label": "No-sessions message", "type": "text", "default": "No sessions scheduled yet for this week."},
            {"key": "session.trainer_label", "label": "Trainer field label", "type": "text", "default": "Trainer"},
            {"key": "session.date_label", "label": "Session date field label", "type": "text", "default": "Session date"},
            {"key": "session.location_label", "label": "Location field label", "type": "text", "default": "Location (optional)"},
            {"key": "session.edit_btn", "label": "Edit session link", "type": "text", "default": "Edit"},
            {"key": "session.delete_btn", "label": "Delete session link", "type": "text", "default": "Delete"},
            {"key": "session.delete_confirm", "label": "Delete session confirmation", "type": "text", "default": "Delete this session? Its attendance, signatures, and uploaded sheet will be lost."},
            {"key": "session.attended_suffix", "label": "Attended-count suffix", "type": "text", "default": "attended"},
            {"key": "session.signed_suffix", "label": "Signed-count suffix", "type": "text", "default": "signed"},
            {"key": "session.open_signoff_btn", "label": "Open digital sign-off sheet button", "type": "text", "default": "Open Sign-off Sheet"},
            {"key": "session.download_signoff_btn", "label": "Download PDF template button", "type": "text", "default": "Download Sign-off Template"},
            {"key": "session.upload_signoff_label", "label": "Upload section label", "type": "text", "default": "Signed sheet"},
            {"key": "session.upload_signoff_btn", "label": "Upload signed sheet button", "type": "text", "default": "Upload Signed Sheet"},
            {"key": "session.signoff_replace_btn", "label": "Replace uploaded sheet button", "type": "text", "default": "Replace"},
            {"key": "session.signoff_uploaded_label", "label": "Uploaded-file view link", "type": "text", "default": "View uploaded sheet"},
            {"key": "session.signoff_remove_btn", "label": "Remove uploaded sheet button", "type": "text", "default": "Remove"},
            {"key": "session.signoff_remove_confirm", "label": "Remove confirmation", "type": "text", "default": "Remove the uploaded signed sheet for this session?"},
            {"key": "session.new_dialog_title", "label": "New session dialog title", "type": "text", "default": "New Session"},
            {"key": "session.edit_dialog_title", "label": "Edit session dialog title", "type": "text", "default": "Edit Session"},
        ],
    },
    {
        "category": "Roster",
        "fields": [
            {"key": "people.heading", "label": "Page heading", "type": "text", "default": "Roster"},
            {"key": "people.count_label", "label": "Count suffix", "type": "text", "default": "people"},
            {"key": "people.empty", "label": "Empty state message", "type": "text", "default": "No one added yet. Add trainees or team members with the form."},
            {"key": "people.remove_btn", "label": "Remove button", "type": "text", "default": "Remove"},
            {"key": "people.remove_confirm", "label": "Remove confirmation (use {name})", "type": "text", "default": "Remove {name} from the roster?"},
            {"key": "people.add_heading", "label": "Add-person card heading", "type": "text", "default": "Add Person"},
            {"key": "people.name_label", "label": "Name field label", "type": "text", "default": "Name"},
            {"key": "people.email_label", "label": "Email field label", "type": "text", "default": "Email (optional)"},
            {"key": "people.add_btn", "label": "Add-to-roster button", "type": "text", "default": "Add to Roster"},
        ],
    },
    {
        "category": "Login & account",
        "fields": [
            {"key": "login.heading", "label": "Login heading", "type": "text", "default": "Welcome Back"},
            {"key": "login.subheading", "label": "Login subheading", "type": "text", "default": "Sign in to access the training ledger."},
            {"key": "login.username_label", "label": "Username field label", "type": "text", "default": "Username"},
            {"key": "login.password_label", "label": "Password field label", "type": "text", "default": "Password"},
            {"key": "login.submit_btn", "label": "Log in button", "type": "text", "default": "Log In"},
            {"key": "nav.account", "label": "Nav: account link", "type": "text", "default": "Account"},
            {"key": "nav.logout", "label": "Nav: log out button", "type": "text", "default": "Log Out"},
            {"key": "nav.login", "label": "Nav: log in link", "type": "text", "default": "Log In"},
            {"key": "account.heading", "label": "Account page heading", "type": "text", "default": "My Account"},
            {"key": "account.current_password_label", "label": "Current password field label", "type": "text", "default": "Current password"},
            {"key": "account.new_password_label", "label": "New password field label", "type": "text", "default": "New password"},
            {"key": "account.confirm_password_label", "label": "Confirm password field label", "type": "text", "default": "Confirm new password"},
            {"key": "account.submit_btn", "label": "Update password button", "type": "text", "default": "Update Password"},
        ],
    },
    {
        "category": "Sign-off sheet",
        "fields": [
            {"key": "sign.back_link", "label": "Back link", "type": "text", "default": "Back to week"},
            {"key": "sign.eyebrow", "label": "Eyebrow label", "type": "text", "default": "Sign-off sheet"},
            {"key": "sign.subheading", "label": "Subheading", "type": "text", "default": "Attendees confirm their attendance below by signing"},
            {"key": "sign.empty", "label": "Empty state message", "type": "text", "default": "No one is marked as attended yet. Check attendance on the week page first."},
            {"key": "sign.not_linked", "label": "Account not linked to roster message", "type": "text", "default": "Your account isn't linked to a roster profile yet, so there's nothing for you to sign here. Ask an admin to link your account under Edit Content \u2192 Users."},
            {"key": "sign.signed_prefix", "label": "\u201cSigned\u201d timestamp prefix", "type": "text", "default": "Signed"},
            {"key": "sign.awaiting", "label": "Awaiting-signature label", "type": "text", "default": "Awaiting signature"},
            {"key": "sign.sign_btn", "label": "Sign button", "type": "text", "default": "Sign"},
            {"key": "sign.dialog_subheading", "label": "Signature dialog subheading", "type": "text", "default": "Draw your signature below to confirm attendance"},
            {"key": "sign.clear_btn", "label": "Clear button", "type": "text", "default": "Clear"},
            {"key": "sign.confirm_btn", "label": "Confirm button", "type": "text", "default": "Confirm & Sign"},
            {"key": "sign.cancel_btn", "label": "Cancel button", "type": "text", "default": "Cancel"},
        ],
    },
    {
        "category": "Physical sign-off sheet",
        "fields": [
            {"key": "week.download_signoff_btn", "label": "Download template button", "type": "text", "default": "Download Sign-off Template"},
            {"key": "week.upload_signoff_label", "label": "Upload section label", "type": "text", "default": "Signed sheet"},
            {"key": "week.upload_signoff_btn", "label": "Upload button", "type": "text", "default": "Upload Signed Sheet"},
            {"key": "week.signoff_uploaded_label", "label": "Uploaded-file view link", "type": "text", "default": "View uploaded sheet"},
            {"key": "week.signoff_replace_btn", "label": "Replace uploaded sheet button", "type": "text", "default": "Replace"},
            {"key": "week.signoff_remove_btn", "label": "Remove uploaded sheet button", "type": "text", "default": "Remove"},
            {"key": "week.signoff_remove_confirm", "label": "Remove confirmation", "type": "text", "default": "Remove the uploaded signed sheet for this week?"},
            {"key": "signoff.pdf_lesson_label", "label": "PDF: Lesson label", "type": "text", "default": "Lesson"},
            {"key": "signoff.pdf_date_label", "label": "PDF: Date label", "type": "text", "default": "Date"},
            {"key": "signoff.pdf_presenter_label", "label": "PDF: Presenter label", "type": "text", "default": "Presenter"},
            {"key": "signoff.pdf_signature_header", "label": "PDF: Signature column header", "type": "text", "default": "Signature"},
            {"key": "signoff.pdf_print_name_header", "label": "PDF: Print Name column header", "type": "text", "default": "Print Name"},
            {"key": "signoff.pdf_date_header", "label": "PDF: Date column header", "type": "text", "default": "Date"},
            {"key": "signoff.pdf_no_attendees", "label": "PDF: no-attendees note", "type": "text", "default": "No attendees marked yet — check attendance on the week page first, then re-download for pre-filled names."},
        ],
    },
    {
        "category": "Reports",
        "fields": [
            {"key": "reports.nav_link", "label": "Nav link: Reports", "type": "text", "default": "Reports"},
            {"key": "reports.heading", "label": "Page heading", "type": "text", "default": "Reports"},
            {"key": "reports.subheading", "label": "Page subheading", "type": "text", "default": "Export signed sheets and attendance for a range of weeks"},
            {"key": "reports.start_week_label", "label": "Start week field label", "type": "text", "default": "Start week"},
            {"key": "reports.end_week_label", "label": "End week field label", "type": "text", "default": "End week"},
            {"key": "reports.zip_heading", "label": "ZIP export card heading", "type": "text", "default": "Signed Sheets (ZIP)"},
            {"key": "reports.zip_body", "label": "ZIP export card body", "type": "text", "default": "Every uploaded signed sheet in the selected range, bundled into one .zip file."},
            {"key": "reports.zip_btn", "label": "ZIP export button", "type": "text", "default": "Download ZIP"},
            {"key": "reports.zip_none", "label": "No signed sheets error", "type": "text", "default": "No signed sheets have been uploaded for that range yet."},
            {"key": "reports.matrix_heading", "label": "Attendance matrix card heading", "type": "text", "default": "Attendance Matrix (xlsx)"},
            {"key": "reports.matrix_body", "label": "Attendance matrix card body", "type": "text", "default": "A spreadsheet with everyone on the roster as rows and each week as a column, marking who attended."},
            {"key": "reports.matrix_btn", "label": "Attendance matrix button", "type": "text", "default": "Download Spreadsheet"},
            {"key": "reports.range_error", "label": "Missing range error", "type": "text", "default": "Choose both a start and end week."},
        ],
    },
]

# Flat lookup of key -> default value, and key -> field type
CONTENT_DEFAULTS = {}
CONTENT_TYPES = {}
for _group in CONTENT_SCHEMA:
    for _f in _group["fields"]:
        CONTENT_DEFAULTS[_f["key"]] = _f["default"]
        CONTENT_TYPES[_f["key"]] = _f.get("type", "text")

# Colors below match the Logicore Portal's own dark design system (see
# static/css/app.css --accent / --bg / --text tokens at the portal level) so
# Training Tracker reads as part of the same product rather than a bolted-on
# third-party app. "ink" and "parchment" kept their historical names from
# the original light/paper theme, but their roles are now inverted: ink is
# the LIGHT tone (text on the dark canvas + light-on-dark button faces),
# parchment is the DARK tone (the canvas itself, dominant background). See
# build_theme() below — the light/dark derivation direction was flipped to
# match, not just the base hex values.
THEME_FIELDS = [
    {"key": "ink", "label": "Ink (text / light-on-dark surfaces)", "default": "#EEF2F6"},
    {"key": "parchment", "label": "Parchment (page background)", "default": "#0A0D12"},
    {"key": "amber", "label": "Amber (primary accent)", "default": "#2FD8A6"},
    {"key": "sage", "label": "Sage (success / signed)", "default": "#22C55E"},
    {"key": "rose", "label": "Rose (danger / delete)", "default": "#EF4444"},
]
THEME_DEFAULTS = {f["key"]: f["default"] for f in THEME_FIELDS}


# ---------------------------------------------------------------------------
# Color helpers — derive light/dark shades from the 5 admin-editable colors
# so the whole palette (used across every template) stays consistent.
# ---------------------------------------------------------------------------

def _hex_to_rgb(h):
    h = h.lstrip("#")
    if len(h) == 3:
        h = "".join(c * 2 for c in h)
    return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))


def _rgb_to_hex(rgb):
    return "#" + "".join(f"{max(0, min(255, round(c))):02X}" for c in rgb)


def _mix(hex_color, target, amount):
    r1, g1, b1 = _hex_to_rgb(hex_color)
    r2, g2, b2 = target
    return _rgb_to_hex((
        r1 + (r2 - r1) * amount,
        g1 + (g2 - g1) * amount,
        b1 + (b2 - b1) * amount,
    ))


def lighten(hex_color, amount):
    return _mix(hex_color, (255, 255, 255), amount)


def darken(hex_color, amount):
    return _mix(hex_color, (0, 0, 0), amount)


def build_theme(colors):
    ink = colors.get("ink", THEME_DEFAULTS["ink"])
    parchment = colors.get("parchment", THEME_DEFAULTS["parchment"])
    amber = colors.get("amber", THEME_DEFAULTS["amber"])
    sage = colors.get("sage", THEME_DEFAULTS["sage"])
    rose = colors.get("rose", THEME_DEFAULTS["rose"])
    # NOTE on direction: "ink" is now the light/text tone, so its "faint"
    # (muted secondary text) shade should read as a dimmer gray, not a
    # brighter near-white — hence darken() here, not lighten() as in the
    # original light-paper theme. Same logic in reverse for "parchment":
    # it's now the dark canvas color, so "dim"/"line" (card surfaces and
    # borders that need to stand out a little from the base canvas) use
    # lighten(), not darken().
    return {
        "ink": {"DEFAULT": ink, "light": lighten(ink, 0.16), "faint": darken(ink, 0.34)},
        "parchment": {"DEFAULT": parchment, "dim": lighten(parchment, 0.06), "line": lighten(parchment, 0.16)},
        "amber": {"DEFAULT": amber, "dark": darken(amber, 0.17), "light": lighten(amber, 0.22)},
        "sage": {"DEFAULT": sage, "dark": darken(sage, 0.17)},
        "rose": {"DEFAULT": rose, "dark": darken(rose, 0.17)},
    }


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


SCHEMA = """
CREATE TABLE IF NOT EXISTS weeks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    week_number INTEGER NOT NULL,
    title TEXT NOT NULL,
    start_date TEXT,
    notes TEXT,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS topics (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    week_id INTEGER NOT NULL REFERENCES weeks(id) ON DELETE CASCADE,
    title TEXT NOT NULL,
    lesson_plan TEXT,
    key_points TEXT,
    sort_order INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS videos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    topic_id INTEGER NOT NULL REFERENCES topics(id) ON DELETE CASCADE,
    title TEXT,
    url TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS people (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    email TEXT,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS sessions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    week_id INTEGER NOT NULL REFERENCES weeks(id) ON DELETE CASCADE,
    trainer_name TEXT NOT NULL DEFAULT 'Unassigned',
    session_date TEXT,
    location TEXT,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS attendance (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    session_id INTEGER NOT NULL REFERENCES sessions(id) ON DELETE CASCADE,
    person_id INTEGER NOT NULL REFERENCES people(id) ON DELETE CASCADE,
    attended INTEGER DEFAULT 0,
    signature TEXT,
    signed_at TEXT,
    UNIQUE(session_id, person_id)
);

CREATE TABLE IF NOT EXISTS content_strings (
    key TEXT PRIMARY KEY,
    value TEXT
);

CREATE TABLE IF NOT EXISTS theme_settings (
    key TEXT PRIMARY KEY,
    value TEXT
);

CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT UNIQUE NOT NULL,
    password_hash TEXT NOT NULL,
    role TEXT NOT NULL DEFAULT 'trainee',
    person_id INTEGER REFERENCES people(id) ON DELETE SET NULL,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS signoff_uploads (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    session_id INTEGER NOT NULL UNIQUE REFERENCES sessions(id) ON DELETE CASCADE,
    filename TEXT NOT NULL,
    original_filename TEXT,
    content_type TEXT,
    uploaded_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
    uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
);
"""


def migrate_to_sessions(db):
    """One-time upgrade for databases created before the 'sessions' table
    existed. Old installs key `attendance` and `signoff_uploads` directly
    by week_id; this rebuilds both keyed by session_id instead, creating
    one 'Unassigned' session per week that had data so every existing
    checkbox, signature, and uploaded scan is preserved — just parked
    under a placeholder trainer until someone renames it from the UI.
    Shares one session per week across both tables so a week's existing
    attendance and its existing uploaded scan land in the same session.
    No-ops on a fresh install or a database that's already migrated.
    """
    db.row_factory = sqlite3.Row
    att_cols = [r["name"] for r in db.execute("PRAGMA table_info(attendance)").fetchall()]
    upload_cols = [r["name"] for r in db.execute("PRAGMA table_info(signoff_uploads)").fetchall()]
    attendance_needs_migration = "week_id" in att_cols and "session_id" not in att_cols
    uploads_needs_migration = "week_id" in upload_cols and "session_id" not in upload_cols

    if not attendance_needs_migration and not uploads_needs_migration:
        return

    print("Migrating attendance and sign-off uploads to per-trainer sessions...")

    week_ids = set()
    if attendance_needs_migration:
        week_ids.update(r["week_id"] for r in db.execute("SELECT DISTINCT week_id FROM attendance").fetchall())
    if uploads_needs_migration:
        week_ids.update(r["week_id"] for r in db.execute("SELECT DISTINCT week_id FROM signoff_uploads").fetchall())

    session_for_week = {}
    for week_id in week_ids:
        week = db.execute("SELECT * FROM weeks WHERE id = ?", (week_id,)).fetchone()
        session_date = week["start_date"] if week else None
        cur = db.execute(
            "INSERT INTO sessions (week_id, trainer_name, session_date) VALUES (?, ?, ?)",
            (week_id, "Unassigned", session_date),
        )
        session_for_week[week_id] = cur.lastrowid

    if attendance_needs_migration:
        db.execute("ALTER TABLE attendance RENAME TO attendance_legacy")
        db.execute(
            """
            CREATE TABLE attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id INTEGER NOT NULL REFERENCES sessions(id) ON DELETE CASCADE,
                person_id INTEGER NOT NULL REFERENCES people(id) ON DELETE CASCADE,
                attended INTEGER DEFAULT 0,
                signature TEXT,
                signed_at TEXT,
                UNIQUE(session_id, person_id)
            )
            """
        )
        for week_id, session_id in session_for_week.items():
            db.execute(
                """
                INSERT INTO attendance (session_id, person_id, attended, signature, signed_at)
                SELECT ?, person_id, attended, signature, signed_at
                FROM attendance_legacy WHERE week_id = ?
                """,
                (session_id, week_id),
            )
        db.execute("DROP TABLE attendance_legacy")

    if uploads_needs_migration:
        db.execute("ALTER TABLE signoff_uploads RENAME TO signoff_uploads_legacy")
        db.execute(
            """
            CREATE TABLE signoff_uploads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id INTEGER NOT NULL UNIQUE REFERENCES sessions(id) ON DELETE CASCADE,
                filename TEXT NOT NULL,
                original_filename TEXT,
                content_type TEXT,
                uploaded_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        for week_id, session_id in session_for_week.items():
            db.execute(
                """
                INSERT INTO signoff_uploads (session_id, filename, original_filename, content_type, uploaded_by, uploaded_at)
                SELECT ?, filename, original_filename, content_type, uploaded_by, uploaded_at
                FROM signoff_uploads_legacy WHERE week_id = ?
                """,
                (session_id, week_id),
            )
        db.execute("DROP TABLE signoff_uploads_legacy")

    db.commit()
    print(f"Migration complete — created {len(session_for_week)} default session(s).")


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")
    db.executescript(SCHEMA)
    migrate_to_sessions(db)
    existing_users = db.execute("SELECT COUNT(*) c FROM users").fetchone()[0]
    if existing_users == 0:
        db.execute(
            "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
            ("admin", generate_password_hash("admin"), "admin"),
        )
    db.commit()
    db.close()


# ---------------------------------------------------------------------------
# Content / theme loading (available to every template)
# ---------------------------------------------------------------------------

def load_content():
    db = get_db()
    rows = db.execute("SELECT key, value FROM content_strings").fetchall()
    overrides = {r["key"]: r["value"] for r in rows if r["value"] not in (None, "")}
    merged = dict(CONTENT_DEFAULTS)
    merged.update(overrides)
    return merged


def load_theme_colors():
    db = get_db()
    rows = db.execute("SELECT key, value FROM theme_settings").fetchall()
    overrides = {r["key"]: r["value"] for r in rows if r["value"]}
    merged = dict(THEME_DEFAULTS)
    merged.update(overrides)
    return merged


# ---------------------------------------------------------------------------
# Authentication — simple username/password login backed by the users table.
# Roles: 'admin' (everything, incl. content/appearance/user management),
# 'editor' (create/edit courses, roster, attendance — not appearance/users),
# 'trainee' (view-only, can sign off their own attendance).
# ---------------------------------------------------------------------------

def get_current_user():
    if "user" in g:
        return g.user
    user_id = session.get("user_id")
    if not user_id:
        g.user = None
        return None
    db = get_db()
    g.user = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if g.user is None:
        session.clear()
    return g.user


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not get_current_user():
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


def roles_required(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            user = get_current_user()
            if not user:
                return redirect(url_for("login", next=request.path))
            if user["role"] not in roles:
                flash("You don't have permission to do that.", "error")
                return redirect(url_for("index"))
            return view(*args, **kwargs)
        return wrapped
    return decorator


@app.context_processor
def inject_globals():
    content = load_content()

    def c(key, **kwargs):
        text = content.get(key, CONTENT_DEFAULTS.get(key, key))
        if kwargs:
            try:
                return text.format(**kwargs)
            except (KeyError, IndexError):
                return text
        return text

    theme = build_theme(load_theme_colors())
    user = get_current_user()
    is_admin = bool(user and user["role"] == "admin")
    can_manage_courses = bool(user and user["role"] in ("admin", "editor"))
    using_default_admin_password = bool(
        user and user["username"] == "admin" and check_password_hash(user["password_hash"], "admin")
    )
    return {
        "c": c,
        "theme": theme,
        "user": user,
        "is_admin": is_admin,
        "can_manage_courses": can_manage_courses,
        "using_default_admin_password": using_default_admin_password,
    }


# ---------------------------------------------------------------------------
# Small utilities
# ---------------------------------------------------------------------------

def get_week_or_404(db, week_id):
    week = db.execute("SELECT * FROM weeks WHERE id = ?", (week_id,)).fetchone()
    if week is None:
        from flask import abort
        abort(404)
    return week


def get_session_or_404(db, session_id):
    sess = db.execute("SELECT * FROM sessions WHERE id = ?", (session_id,)).fetchone()
    if sess is None:
        from flask import abort
        abort(404)
    return sess


def session_progress(db, session_id):
    """Attended/signed counts for one trainer's session."""
    attended = db.execute(
        "SELECT COUNT(*) c FROM attendance WHERE session_id = ? AND attended = 1", (session_id,)
    ).fetchone()["c"]
    signed = db.execute(
        "SELECT COUNT(*) c FROM attendance WHERE session_id = ? AND attended = 1 AND signature IS NOT NULL",
        (session_id,),
    ).fetchone()["c"]
    return attended, signed


def week_progress(db, week_id):
    """Roll-up across every session of a week, for the dashboard. Counts
    distinct people so someone who attended more than one session of the
    same week isn't double-counted."""
    attended = db.execute(
        """
        SELECT COUNT(DISTINCT a.person_id) c
        FROM attendance a JOIN sessions s ON s.id = a.session_id
        WHERE s.week_id = ? AND a.attended = 1
        """,
        (week_id,),
    ).fetchone()["c"]
    signed = db.execute(
        """
        SELECT COUNT(DISTINCT a.person_id) c
        FROM attendance a JOIN sessions s ON s.id = a.session_id
        WHERE s.week_id = ? AND a.attended = 1 AND a.signature IS NOT NULL
        """,
        (week_id,),
    ).fetchone()["c"]
    return attended, signed


# ---------------------------------------------------------------------------
# Physical sign-off sheets — PDF template, uploaded scans, and reports
# (ZIP of signed sheets, attendance-matrix spreadsheet).
# ---------------------------------------------------------------------------

def get_signoff_upload(db, session_id):
    return db.execute("SELECT * FROM signoff_uploads WHERE session_id = ?", (session_id,)).fetchone()


def safe_filename_part(text):
    text = re.sub(r"[^A-Za-z0-9 _-]", "", text or "").strip()
    return text or "Untitled"


def build_signoff_pdf(week, session, attendee_names, content):
    """Generates a printable sign-off sheet for one trainer's session:
    roster names pre-filled (if attendance has been checked already) with
    blank Signature/Date columns. If no one's been checked off yet, ships
    a handful of blank rows instead."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    trainer_name = session["trainer_name"] if session and session["trainer_name"] != "Unassigned" else ""

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
        title=f"Week {week['week_number']}: {week['title']}"
        + (f" — {trainer_name}" if trainer_name else "")
        + " — Sign-off Sheet",
        author=content.get("site.name", "Training Ledger"),
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TLTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "TLMeta", parent=styles["Normal"], fontSize=10, textColor=rl_colors.HexColor("#444444"), spaceAfter=2,
    )
    note_style = ParagraphStyle(
        "TLNote", parent=styles["Normal"], fontSize=9, textColor=rl_colors.HexColor("#B5533C"), spaceBefore=8,
    )

    lesson_label = content.get("signoff.pdf_lesson_label", "Lesson")
    date_label = content.get("signoff.pdf_date_label", "Date")
    presenter_label = content.get("signoff.pdf_presenter_label", "Presenter")
    session_date = session["session_date"] if session and session["session_date"] else week["start_date"]
    display_date = friendly_date(session_date) if session_date else ""
    presenter_value = trainer_name or "________________________________"

    elements = [
        Paragraph(f"Week {week['week_number']}: {week['title']}", title_style),
        Paragraph(f"{lesson_label}: {week['title']}", meta_style),
        Paragraph(f"{date_label}: {display_date}", meta_style),
        Paragraph(f"{presenter_label}: {presenter_value}", meta_style),
        Spacer(1, 0.25 * inch),
    ]

    sig_header = content.get("signoff.pdf_signature_header", "Signature")
    name_header = content.get("signoff.pdf_print_name_header", "Print Name")
    date_header = content.get("signoff.pdf_date_header", "Date")

    data = [[name_header, sig_header, date_header]]
    if attendee_names:
        for name in attendee_names:
            data.append([name, "", ""])
    else:
        data.extend([["", "", ""]] * 12)

    table = Table(data, colWidths=[2.3 * inch, 3.0 * inch, 1.4 * inch], repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("LINEBELOW", (0, 0), (-1, 0), 1, rl_colors.HexColor("#1B2430")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.6, rl_colors.HexColor("#CBBFA3")),
        ("TOPPADDING", (0, 1), (-1, -1), 14),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 14),
    ]))
    elements.append(table)

    if not attendee_names:
        elements.append(Paragraph(
            content.get(
                "signoff.pdf_no_attendees",
                "No attendees marked yet — check attendance on the week page first, then re-download for pre-filled names.",
            ),
            note_style,
        ))

    doc.build(elements)
    buf.seek(0)
    return buf


def parse_week_range(db):
    start_id = request.args.get("start_week", type=int)
    end_id = request.args.get("end_week", type=int)
    if not start_id or not end_id:
        return None, None, "reports.range_error"
    start_week = db.execute("SELECT * FROM weeks WHERE id = ?", (start_id,)).fetchone()
    end_week = db.execute("SELECT * FROM weeks WHERE id = ?", (end_id,)).fetchone()
    if not start_week or not end_week:
        return None, None, "reports.range_error"
    lo, hi = sorted([start_week["week_number"], end_week["week_number"]])
    return lo, hi, None


def build_signoff_zip(db, start_week_number, end_week_number):
    weeks = db.execute(
        "SELECT * FROM weeks WHERE week_number BETWEEN ? AND ? ORDER BY week_number ASC",
        (start_week_number, end_week_number),
    ).fetchall()
    buf = io.BytesIO()
    count = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for w in weeks:
            sessions = db.execute(
                "SELECT * FROM sessions WHERE week_id = ? ORDER BY session_date ASC, id ASC", (w["id"],)
            ).fetchall()
            for s in sessions:
                upload = get_signoff_upload(db, s["id"])
                if not upload:
                    continue
                path = os.path.join(SIGNOFF_UPLOAD_DIR, upload["filename"])
                if not os.path.exists(path):
                    continue
                ext = upload["filename"].rsplit(".", 1)[-1] if "." in upload["filename"] else "dat"
                trainer_part = f" - {safe_filename_part(s['trainer_name'])}" if s["trainer_name"] != "Unassigned" else ""
                arcname = f"Week {w['week_number']:02d} - {safe_filename_part(w['title'])}{trainer_part}.{ext}"
                zf.write(path, arcname)
                count += 1
    buf.seek(0)
    return buf, count


def build_attendance_matrix_xlsx(db, start_week_number, end_week_number):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    weeks = db.execute(
        "SELECT * FROM weeks WHERE week_number BETWEEN ? AND ? ORDER BY week_number ASC",
        (start_week_number, end_week_number),
    ).fetchall()
    people = db.execute("SELECT * FROM people ORDER BY name ASC").fetchall()

    attended_set = set()
    if weeks:
        placeholders = ",".join("?" for _ in weeks)
        rows = db.execute(
            f"""
            SELECT s.week_id AS week_id, a.person_id AS person_id
            FROM attendance a JOIN sessions s ON s.id = a.session_id
            WHERE a.attended = 1 AND s.week_id IN ({placeholders})
            """,
            [w["id"] for w in weeks],
        ).fetchall()
        attended_set = {(r["week_id"], r["person_id"]) for r in rows}

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B2430")
    center = Alignment(horizontal="center")

    ws.cell(row=1, column=1, value="Name").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    for col, w in enumerate(weeks, start=2):
        cell = ws.cell(row=1, column=col, value=f"Week {w['week_number']:02d}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row, p in enumerate(people, start=2):
        ws.cell(row=row, column=1, value=p["name"])
        for col, w in enumerate(weeks, start=2):
            mark = "X" if (w["id"], p["id"]) in attended_set else ""
            ws.cell(row=row, column=col, value=mark).alignment = center

    ws.column_dimensions["A"].width = 28
    for col in range(2, len(weeks) + 2):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Routes: Dashboard
# ---------------------------------------------------------------------------

@app.route("/")
@login_required
def index():
    db = get_db()
    weeks = db.execute("SELECT * FROM weeks ORDER BY week_number ASC").fetchall()
    weeks_data = []
    for w in weeks:
        topic_count = db.execute(
            "SELECT COUNT(*) c FROM topics WHERE week_id = ?", (w["id"],)
        ).fetchone()["c"]
        attended, signed = week_progress(db, w["id"])
        weeks_data.append(
            {"row": w, "topic_count": topic_count, "attended": attended, "signed": signed}
        )
    people_count = db.execute("SELECT COUNT(*) c FROM people").fetchone()["c"]
    return render_template("index.html", weeks=weeks_data, people_count=people_count)


@app.route("/week/new", methods=["POST"])
@roles_required("admin", "editor")
def new_week():
    db = get_db()
    next_number = db.execute("SELECT COALESCE(MAX(week_number), 0) + 1 n FROM weeks").fetchone()["n"]
    title = request.form.get("title", "").strip() or f"Week {next_number}"
    start_date = request.form.get("start_date") or None
    cur = db.execute(
        "INSERT INTO weeks (week_number, title, start_date) VALUES (?, ?, ?)",
        (next_number, title, start_date),
    )
    db.commit()
    return redirect(url_for("week_detail", week_id=cur.lastrowid))


@app.route("/week/<int:week_id>/edit", methods=["POST"])
@roles_required("admin", "editor")
def edit_week(week_id):
    db = get_db()
    title = request.form.get("title", "").strip()
    start_date = request.form.get("start_date") or None
    notes = request.form.get("notes", "").strip()
    db.execute(
        "UPDATE weeks SET title = ?, start_date = ?, notes = ? WHERE id = ?",
        (title, start_date, notes, week_id),
    )
    db.commit()
    return redirect(url_for("week_detail", week_id=week_id))


@app.route("/week/<int:week_id>/delete", methods=["POST"])
@roles_required("admin", "editor")
def delete_week(week_id):
    db = get_db()
    db.execute("DELETE FROM weeks WHERE id = ?", (week_id,))
    db.commit()
    return redirect(url_for("index"))


# ---------------------------------------------------------------------------
# Routes: Week detail (topics, videos, attendance)
# ---------------------------------------------------------------------------

@app.route("/week/<int:week_id>")
@login_required
def week_detail(week_id):
    db = get_db()
    week = get_week_or_404(db, week_id)
    topics = db.execute(
        "SELECT * FROM topics WHERE week_id = ? ORDER BY sort_order ASC, id ASC", (week_id,)
    ).fetchall()
    topics_data = []
    for t in topics:
        videos = db.execute("SELECT * FROM videos WHERE topic_id = ? ORDER BY id ASC", (t["id"],)).fetchall()
        key_points = [p for p in (t["key_points"] or "").split("\n") if p.strip()]
        topics_data.append({"row": t, "videos": videos, "key_points": key_points})

    people = db.execute("SELECT * FROM people ORDER BY name ASC").fetchall()
    people_total = len(people)

    sessions = db.execute(
        "SELECT * FROM sessions WHERE week_id = ? ORDER BY session_date ASC, id ASC", (week_id,)
    ).fetchall()
    sessions_data = []
    for s in sessions:
        attendance_rows = db.execute("SELECT * FROM attendance WHERE session_id = ?", (s["id"],)).fetchall()
        attendance_map = {a["person_id"]: a for a in attendance_rows}
        session_roster = []
        for p in people:
            a = attendance_map.get(p["id"])
            session_roster.append({"person": p, "attended": bool(a["attended"]) if a else False})
        s_attended, s_signed = session_progress(db, s["id"])
        sessions_data.append(
            {
                "row": s,
                "attended": s_attended,
                "signed": s_signed,
                "roster_total": people_total,
                "roster": session_roster,
                "upload": get_signoff_upload(db, s["id"]),
            }
        )

    attended, signed = week_progress(db, week_id)

    return render_template(
        "week_detail.html",
        week=week,
        topics=topics_data,
        sessions=sessions_data,
        attended=attended,
        signed=signed,
        people_total=people_total,
    )


# ---- Sessions (one per trainer/cohort delivering a given week) ----

@app.route("/week/<int:week_id>/session/new", methods=["POST"])
@roles_required("admin", "editor")
def new_session(week_id):
    db = get_db()
    get_week_or_404(db, week_id)
    trainer_name = request.form.get("trainer_name", "").strip() or "Unassigned"
    session_date = request.form.get("session_date") or None
    location = request.form.get("location", "").strip() or None
    db.execute(
        "INSERT INTO sessions (week_id, trainer_name, session_date, location) VALUES (?, ?, ?, ?)",
        (week_id, trainer_name, session_date, location),
    )
    db.commit()
    return redirect(url_for("week_detail", week_id=week_id))


@app.route("/session/<int:session_id>/edit", methods=["POST"])
@roles_required("admin", "editor")
def edit_session(session_id):
    db = get_db()
    sess = get_session_or_404(db, session_id)
    trainer_name = request.form.get("trainer_name", "").strip() or "Unassigned"
    session_date = request.form.get("session_date") or None
    location = request.form.get("location", "").strip() or None
    db.execute(
        "UPDATE sessions SET trainer_name = ?, session_date = ?, location = ? WHERE id = ?",
        (trainer_name, session_date, location, session_id),
    )
    db.commit()
    return redirect(url_for("week_detail", week_id=sess["week_id"]))


@app.route("/session/<int:session_id>/delete", methods=["POST"])
@roles_required("admin", "editor")
def delete_session(session_id):
    db = get_db()
    sess = get_session_or_404(db, session_id)
    week_id = sess["week_id"]
    upload = get_signoff_upload(db, session_id)
    if upload:
        path = os.path.join(SIGNOFF_UPLOAD_DIR, upload["filename"])
        if os.path.exists(path):
            os.remove(path)
    db.execute("DELETE FROM sessions WHERE id = ?", (session_id,))
    db.commit()
    return redirect(url_for("week_detail", week_id=week_id))


# ---- Topics ----

@app.route("/week/<int:week_id>/topic/new", methods=["POST"])
@roles_required("admin", "editor")
def new_topic(week_id):
    db = get_db()
    title = request.form.get("title", "").strip() or "Untitled topic"
    lesson_plan = request.form.get("lesson_plan", "").strip()
    key_points = request.form.get("key_points", "").strip()
    max_order = db.execute(
        "SELECT COALESCE(MAX(sort_order), 0) + 1 n FROM topics WHERE week_id = ?", (week_id,)
    ).fetchone()["n"]
    db.execute(
        "INSERT INTO topics (week_id, title, lesson_plan, key_points, sort_order) VALUES (?, ?, ?, ?, ?)",
        (week_id, title, lesson_plan, key_points, max_order),
    )
    db.commit()
    return redirect(url_for("week_detail", week_id=week_id))


@app.route("/topic/<int:topic_id>/edit", methods=["POST"])
@roles_required("admin", "editor")
def edit_topic(topic_id):
    db = get_db()
    topic = db.execute("SELECT * FROM topics WHERE id = ?", (topic_id,)).fetchone()
    title = request.form.get("title", "").strip() or "Untitled topic"
    lesson_plan = request.form.get("lesson_plan", "").strip()
    key_points = request.form.get("key_points", "").strip()
    db.execute(
        "UPDATE topics SET title = ?, lesson_plan = ?, key_points = ? WHERE id = ?",
        (title, lesson_plan, key_points, topic_id),
    )
    db.commit()
    return redirect(url_for("week_detail", week_id=topic["week_id"]))


@app.route("/topic/<int:topic_id>/delete", methods=["POST"])
@roles_required("admin", "editor")
def delete_topic(topic_id):
    db = get_db()
    topic = db.execute("SELECT * FROM topics WHERE id = ?", (topic_id,)).fetchone()
    week_id = topic["week_id"]
    db.execute("DELETE FROM topics WHERE id = ?", (topic_id,))
    db.commit()
    return redirect(url_for("week_detail", week_id=week_id))


# ---- Videos ----

@app.route("/topic/<int:topic_id>/video/add", methods=["POST"])
@roles_required("admin", "editor")
def add_video(topic_id):
    db = get_db()
    topic = db.execute("SELECT * FROM topics WHERE id = ?", (topic_id,)).fetchone()
    url = request.form.get("url", "").strip()
    title = request.form.get("video_title", "").strip() or url
    if url:
        db.execute("INSERT INTO videos (topic_id, title, url) VALUES (?, ?, ?)", (topic_id, title, url))
        db.commit()
    return redirect(url_for("week_detail", week_id=topic["week_id"]))


@app.route("/video/<int:video_id>/delete", methods=["POST"])
@roles_required("admin", "editor")
def delete_video(video_id):
    db = get_db()
    video = db.execute(
        "SELECT videos.*, topics.week_id as week_id FROM videos JOIN topics ON topics.id = videos.topic_id WHERE videos.id = ?",
        (video_id,),
    ).fetchone()
    db.execute("DELETE FROM videos WHERE id = ?", (video_id,))
    db.commit()
    return redirect(url_for("week_detail", week_id=video["week_id"]))


# ---------------------------------------------------------------------------
# Routes: People / roster
# ---------------------------------------------------------------------------

@app.route("/people")
@login_required
def people_page():
    db = get_db()
    people = db.execute("SELECT * FROM people ORDER BY name ASC").fetchall()
    return render_template("people.html", people=people)


@app.route("/people/add", methods=["POST"])
@roles_required("admin", "editor")
def add_person():
    db = get_db()
    name = request.form.get("name", "").strip()
    email = request.form.get("email", "").strip()
    if name:
        db.execute("INSERT INTO people (name, email) VALUES (?, ?)", (name, email))
        db.commit()
    return redirect(request.referrer or url_for("people_page"))


@app.route("/people/<int:person_id>/delete", methods=["POST"])
@roles_required("admin", "editor")
def delete_person(person_id):
    db = get_db()
    db.execute("DELETE FROM people WHERE id = ?", (person_id,))
    db.commit()
    return redirect(request.referrer or url_for("people_page"))


# ---------------------------------------------------------------------------
# Routes: Attendance + sign-off
# ---------------------------------------------------------------------------

@app.route("/session/<int:session_id>/attendance/toggle", methods=["POST"])
@roles_required("admin", "editor")
def toggle_attendance(session_id):
    db = get_db()
    get_session_or_404(db, session_id)
    person_id = int(request.form.get("person_id"))
    attended = 1 if request.form.get("attended") == "1" else 0
    existing = db.execute(
        "SELECT * FROM attendance WHERE session_id = ? AND person_id = ?", (session_id, person_id)
    ).fetchone()
    if existing:
        db.execute(
            "UPDATE attendance SET attended = ? WHERE session_id = ? AND person_id = ?",
            (attended, session_id, person_id),
        )
    else:
        db.execute(
            "INSERT INTO attendance (session_id, person_id, attended) VALUES (?, ?, ?)",
            (session_id, person_id, attended),
        )
    db.commit()
    return jsonify({"ok": True})


@app.route("/session/<int:session_id>/signoff")
@login_required
def signoff_sheet(session_id):
    db = get_db()
    sess = get_session_or_404(db, session_id)
    week = get_week_or_404(db, sess["week_id"])
    user = get_current_user()
    people = db.execute("SELECT * FROM people ORDER BY name ASC").fetchall()
    attendance_rows = db.execute("SELECT * FROM attendance WHERE session_id = ?", (session_id,)).fetchall()
    attendance_map = {a["person_id"]: a for a in attendance_rows}
    roster = []
    for p in people:
        a = attendance_map.get(p["id"])
        if not a or not a["attended"]:
            continue
        # Trainees only see/sign their own row; admins and editors can see
        # and sign for anyone (useful for a shared kiosk at a training).
        if user["role"] == "trainee" and p["id"] != user["person_id"]:
            continue
        roster.append(
            {"person": p, "signed": bool(a["signature"]), "signature": a["signature"], "signed_at": a["signed_at"]}
        )
    today = friendly(date.today())
    return render_template("sign.html", week=week, session=sess, roster=roster, today=today)


@app.route("/session/<int:session_id>/sign/<int:person_id>", methods=["POST"])
@login_required
def submit_signature(session_id, person_id):
    db = get_db()
    get_session_or_404(db, session_id)
    user = get_current_user()
    if user["role"] == "trainee" and person_id != user["person_id"]:
        flash("You can only sign for yourself.", "error")
        return redirect(url_for("signoff_sheet", session_id=session_id))
    signature = request.form.get("signature", "")
    if not signature:
        flash("Please provide a signature before submitting.", "error")
        return redirect(url_for("signoff_sheet", session_id=session_id))
    signed_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    existing = db.execute(
        "SELECT * FROM attendance WHERE session_id = ? AND person_id = ?", (session_id, person_id)
    ).fetchone()
    if existing:
        db.execute(
            "UPDATE attendance SET signature = ?, signed_at = ?, attended = 1 WHERE session_id = ? AND person_id = ?",
            (signature, signed_at, session_id, person_id),
        )
    else:
        db.execute(
            "INSERT INTO attendance (session_id, person_id, attended, signature, signed_at) VALUES (?, ?, 1, ?, ?)",
            (session_id, person_id, signature, signed_at),
        )
    db.commit()
    return redirect(url_for("signoff_sheet", session_id=session_id))


@app.route("/session/<int:session_id>/signoff/download")
@roles_required("admin", "editor")
def download_signoff_template(session_id):
    db = get_db()
    sess = get_session_or_404(db, session_id)
    week = get_week_or_404(db, sess["week_id"])
    people = db.execute("SELECT * FROM people ORDER BY name ASC").fetchall()
    attendance_rows = db.execute(
        "SELECT * FROM attendance WHERE session_id = ? AND attended = 1", (session_id,)
    ).fetchall()
    attended_ids = {a["person_id"] for a in attendance_rows}
    attendee_names = [p["name"] for p in people if p["id"] in attended_ids]
    content = load_content()
    buf = build_signoff_pdf(week, sess, attendee_names, content)
    trainer_part = f" - {safe_filename_part(sess['trainer_name'])}" if sess["trainer_name"] != "Unassigned" else ""
    filename = f"Week {week['week_number']:02d} Sign-off Template{trainer_part}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=filename)


@app.route("/session/<int:session_id>/signoff/upload", methods=["POST"])
@roles_required("admin", "editor")
def upload_signoff_scan(session_id):
    db = get_db()
    sess = get_session_or_404(db, session_id)
    file = request.files.get("signoff_file")
    if not file or not file.filename:
        flash("Choose a file to upload.", "error")
        return redirect(url_for("week_detail", week_id=sess["week_id"]))
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ALLOWED_SIGNOFF_EXTENSIONS:
        flash("That file type isn't supported. Use PDF, PNG, JPG, GIF, WEBP, or HEIC.", "error")
        return redirect(url_for("week_detail", week_id=sess["week_id"]))

    user = get_current_user()
    existing = get_signoff_upload(db, session_id)
    if existing:
        old_path = os.path.join(SIGNOFF_UPLOAD_DIR, existing["filename"])
        if os.path.exists(old_path):
            os.remove(old_path)

    stored_name = f"session_{session_id}_{secrets.token_hex(8)}.{ext}"
    file.save(os.path.join(SIGNOFF_UPLOAD_DIR, stored_name))

    if existing:
        db.execute(
            "UPDATE signoff_uploads SET filename = ?, original_filename = ?, content_type = ?, "
            "uploaded_by = ?, uploaded_at = CURRENT_TIMESTAMP WHERE session_id = ?",
            (stored_name, secure_filename(file.filename), file.content_type, user["id"], session_id),
        )
    else:
        db.execute(
            "INSERT INTO signoff_uploads (session_id, filename, original_filename, content_type, uploaded_by) "
            "VALUES (?, ?, ?, ?, ?)",
            (session_id, stored_name, secure_filename(file.filename), file.content_type, user["id"]),
        )
    db.commit()
    flash("Signed sheet uploaded.", "success")
    return redirect(url_for("week_detail", week_id=sess["week_id"]))


@app.route("/session/<int:session_id>/signoff/view")
@roles_required("admin", "editor")
def view_signoff_scan(session_id):
    db = get_db()
    get_session_or_404(db, session_id)
    upload = get_signoff_upload(db, session_id)
    if not upload:
        abort(404)
    path = os.path.join(SIGNOFF_UPLOAD_DIR, upload["filename"])
    if not os.path.exists(path):
        abort(404)
    return send_file(path, download_name=upload["original_filename"] or upload["filename"])


@app.route("/session/<int:session_id>/signoff/remove", methods=["POST"])
@roles_required("admin", "editor")
def remove_signoff_scan(session_id):
    db = get_db()
    sess = get_session_or_404(db, session_id)
    upload = get_signoff_upload(db, session_id)
    if upload:
        path = os.path.join(SIGNOFF_UPLOAD_DIR, upload["filename"])
        if os.path.exists(path):
            os.remove(path)
        db.execute("DELETE FROM signoff_uploads WHERE session_id = ?", (session_id,))
        db.commit()
        flash("Signed sheet removed.", "success")
    return redirect(url_for("week_detail", week_id=sess["week_id"]))


# ---------------------------------------------------------------------------
# Routes: Reports (ZIP of signed sheets, attendance-by-week spreadsheet)
# ---------------------------------------------------------------------------

@app.route("/reports")
@roles_required("admin", "editor")
def reports_page():
    db = get_db()
    weeks = db.execute("SELECT * FROM weeks ORDER BY week_number ASC").fetchall()
    return render_template("reports.html", weeks=weeks)


@app.route("/reports/signoff-zip")
@roles_required("admin", "editor")
def report_signoff_zip():
    db = get_db()
    lo, hi, error_key = parse_week_range(db)
    if error_key:
        flash(load_content().get(error_key, "Choose both a start and end week."), "error")
        return redirect(url_for("reports_page"))
    buf, count = build_signoff_zip(db, lo, hi)
    if count == 0:
        flash(
            load_content().get("reports.zip_none", "No signed sheets have been uploaded for that range yet."),
            "error",
        )
        return redirect(url_for("reports_page"))
    filename = f"Signoff Sheets Week {lo:02d}-{hi:02d}.zip"
    return send_file(buf, mimetype="application/zip", as_attachment=True, download_name=filename)


@app.route("/reports/attendance-matrix")
@roles_required("admin", "editor")
def report_attendance_matrix():
    db = get_db()
    lo, hi, error_key = parse_week_range(db)
    if error_key:
        flash(load_content().get(error_key, "Choose both a start and end week."), "error")
        return redirect(url_for("reports_page"))
    buf = build_attendance_matrix_xlsx(db, lo, hi)
    filename = f"Attendance Week {lo:02d}-{hi:02d}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ---------------------------------------------------------------------------
# Routes: Authentication
# ---------------------------------------------------------------------------

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        db = get_db()
        row = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        if row and check_password_hash(row["password_hash"], password):
            session.clear()
            session["user_id"] = row["id"]
            session.permanent = True
            next_url = request.form.get("next") or request.args.get("next") or url_for("index")
            return redirect(next_url)
        flash("Incorrect username or password.", "error")
    next_url = request.args.get("next", "")
    return render_template("login.html", next_url=next_url)


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    flash("Signed out.", "success")
    return redirect(url_for("login"))


@app.route("/account", methods=["GET", "POST"])
@login_required
def account():
    user = get_current_user()
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        if not check_password_hash(user["password_hash"], current_password):
            flash("Current password is incorrect.", "error")
        elif len(new_password) < 4:
            flash("New password must be at least 4 characters.", "error")
        elif new_password != confirm_password:
            flash("New passwords don't match.", "error")
        else:
            db = get_db()
            db.execute(
                "UPDATE users SET password_hash = ? WHERE id = ?",
                (generate_password_hash(new_password), user["id"]),
            )
            db.commit()
            flash("Password updated.", "success")
            return redirect(url_for("account"))
    return render_template("account.html")


# ---------------------------------------------------------------------------
# Routes: Admin — content, appearance & user management
# ---------------------------------------------------------------------------

@app.route("/admin")
@roles_required("admin")
def admin_content():
    db = get_db()
    content = load_content()
    colors = load_theme_colors()
    users = db.execute("SELECT * FROM users ORDER BY username ASC").fetchall()
    people = db.execute("SELECT * FROM people ORDER BY name ASC").fetchall()
    return render_template(
        "admin.html",
        schema=CONTENT_SCHEMA,
        values=content,
        theme_fields=THEME_FIELDS,
        colors=colors,
        users=users,
        people=people,
    )


@app.route("/admin/content/save", methods=["POST"])
@roles_required("admin")
def admin_save_content():
    db = get_db()
    for key in CONTENT_DEFAULTS:
        if key in request.form:
            value = request.form.get(key, "").strip()
            db.execute(
                "INSERT INTO content_strings (key, value) VALUES (?, ?) "
                "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                (key, value),
            )
    db.commit()
    flash("Content updated.", "success")
    return redirect(url_for("admin_content") + "#content")


@app.route("/admin/appearance/save", methods=["POST"])
@roles_required("admin")
def admin_save_appearance():
    db = get_db()
    for field in THEME_FIELDS:
        key = field["key"]
        value = request.form.get(key, "").strip()
        if value:
            db.execute(
                "INSERT INTO theme_settings (key, value) VALUES (?, ?) "
                "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                (key, value),
            )
    db.commit()
    flash("Appearance updated.", "success")
    return redirect(url_for("admin_content") + "#appearance")


@app.route("/admin/reset", methods=["POST"])
@roles_required("admin")
def admin_reset():
    db = get_db()
    scope = request.form.get("scope")
    if scope == "content":
        db.execute("DELETE FROM content_strings")
    elif scope == "appearance":
        db.execute("DELETE FROM theme_settings")
    db.commit()
    flash("Reset to defaults.", "success")
    return redirect(url_for("admin_content"))


@app.route("/admin/users/new", methods=["POST"])
@roles_required("admin")
def admin_new_user():
    db = get_db()
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    role = request.form.get("role", "trainee")
    if role not in ("admin", "editor", "trainee"):
        role = "trainee"
    person_id = request.form.get("person_id") or None
    if not username or not password:
        flash("Username and password are required.", "error")
        return redirect(url_for("admin_content") + "#users")
    if len(password) < 4:
        flash("Password must be at least 4 characters.", "error")
        return redirect(url_for("admin_content") + "#users")
    taken = db.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
    if taken:
        flash(f'The username "{username}" is already taken.', "error")
        return redirect(url_for("admin_content") + "#users")
    db.execute(
        "INSERT INTO users (username, password_hash, role, person_id) VALUES (?, ?, ?, ?)",
        (username, generate_password_hash(password), role, person_id),
    )
    db.commit()
    flash(f'User "{username}" created.', "success")
    return redirect(url_for("admin_content") + "#users")


@app.route("/admin/users/<int:user_id>/update", methods=["POST"])
@roles_required("admin")
def admin_update_user(user_id):
    db = get_db()
    role = request.form.get("role", "trainee")
    if role not in ("admin", "editor", "trainee"):
        role = "trainee"
    person_id = request.form.get("person_id") or None
    new_password = request.form.get("password", "").strip()

    target = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not target:
        flash("User not found.", "error")
        return redirect(url_for("admin_content") + "#users")

    admin_count = db.execute("SELECT COUNT(*) c FROM users WHERE role = 'admin'").fetchone()["c"]
    if target["role"] == "admin" and role != "admin" and admin_count <= 1:
        flash("Can't demote the last remaining admin.", "error")
        return redirect(url_for("admin_content") + "#users")

    db.execute("UPDATE users SET role = ?, person_id = ? WHERE id = ?", (role, person_id, user_id))
    if new_password:
        if len(new_password) < 4:
            flash("Password must be at least 4 characters — role/roster link were saved, password was not.", "error")
            return redirect(url_for("admin_content") + "#users")
        db.execute(
            "UPDATE users SET password_hash = ? WHERE id = ?",
            (generate_password_hash(new_password), user_id),
        )
    db.commit()
    flash(f'User "{target["username"]}" updated.', "success")
    return redirect(url_for("admin_content") + "#users")


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@roles_required("admin")
def admin_delete_user(user_id):
    db = get_db()
    current = get_current_user()
    target = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not target:
        flash("User not found.", "error")
        return redirect(url_for("admin_content") + "#users")
    if target["id"] == current["id"]:
        flash("You can't delete your own account while logged in.", "error")
        return redirect(url_for("admin_content") + "#users")
    admin_count = db.execute("SELECT COUNT(*) c FROM users WHERE role = 'admin'").fetchone()["c"]
    if target["role"] == "admin" and admin_count <= 1:
        flash("Can't delete the last remaining admin account.", "error")
        return redirect(url_for("admin_content") + "#users")
    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()
    flash(f'User "{target["username"]}" removed.', "success")
    return redirect(url_for("admin_content") + "#users")


# ---------------------------------------------------------------------------

@app.template_filter("friendly_date")
def friendly_date(value):
    if not value:
        return ""
    try:
        d = datetime.strptime(value, "%Y-%m-%d")
        return friendly(d, month_format="%b")
    except ValueError:
        return value


# ---------------------------------------------------------------------------
# Networking helpers — find the LAN-facing IP so we can tell the host what
# address to share with teammates.
# ---------------------------------------------------------------------------

def get_lan_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        # Doesn't actually send anything (UDP) — just asks the OS which
        # local interface it would use to reach an external address.
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
    except Exception:
        ip = "127.0.0.1"
    finally:
        s.close()
    return ip


def _open_browser_when_ready(port):
    time.sleep(1.2)
    try:
        webbrowser.open(f"http://127.0.0.1:{port}/")
    except Exception:
        pass


def _print_banner(port, lan_ip):
    line = "=" * 62
    print(line)
    print("  Training Ledger is running")
    print(line)
    print(f"  On this PC:        http://localhost:{port}/")
    print(f"  On your network:   http://{lan_ip}:{port}/")
    print()
    print("  Share the second address with anyone on the same network —")
    print("  everything they enter is saved here and visible to everyone.")
    print()
    print("  Keep this window open. Closing it stops the app for everyone.")
    print(line)


def _running_on_hosted_platform():
    """True when running on a cloud host (Render, Railway, Heroku-style
    platforms all set one of these), as opposed to a local/LAN machine."""
    return bool(
        os.environ.get("RENDER")
        or os.environ.get("RAILWAY_ENVIRONMENT")
        or os.environ.get("DYNO")
        or os.environ.get("FLY_APP_NAME")
    )


if __name__ == "__main__":
    init_db()

    if _running_on_hosted_platform():
        print(f"Training Ledger starting on port {PORT} (DATA_DIR={DATA_DIR})...")
    else:
        lan_ip = get_lan_ip()
        _print_banner(PORT, lan_ip)
        threading.Thread(target=_open_browser_when_ready, args=(PORT,), daemon=True).start()

    try:
        from waitress import serve
        serve(app, host=HOST, port=PORT, threads=8)
    except ImportError:
        # Fallback if waitress isn't installed — fine for quick local testing,
        # not recommended for multiple simultaneous LAN users.
        app.run(host=HOST, port=PORT, debug=False)
