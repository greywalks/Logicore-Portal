from datetime import datetime
import sqlite3

import pandas as pd
from openpyxl import Workbook, load_workbook

# nonconforming.init_db() currently assumes named SQLite rows during app import,
# although its raw connection does not set row_factory. Keep this PR's tests
# focused on billing/access-control by supplying named rows only during import.
_real_sqlite_connect = sqlite3.connect


def _row_sqlite_connect(*args, **kwargs):
    connection = _real_sqlite_connect(*args, **kwargs)
    connection.row_factory = sqlite3.Row
    return connection


sqlite3.connect = _row_sqlite_connect
try:
    import app as portal_app
finally:
    sqlite3.connect = _real_sqlite_connect

import builder
import storage_builder


def _storage_template(path):
    wb = Workbook()
    wb.active.title = "Breakdown"
    for name in (
        "Unit Storage",
        "Unit Receiving",
        "Units Shipped",
        "Part Testing & Programming",
        "Small Parts Check In",
    ):
        wb.create_sheet(name)
    wb.save(path)


def test_fedex_small_part_picks_sum_quantity(tmp_path):
    source = tmp_path / "fedex.xlsx"
    pd.DataFrame({
        "Request Date": ["2026-08-01", "2026-08-02", "2026-08-03", "2026-08-04"],
        "MSO": ["M1", "M1", "M2", "M3"],
        "Outbound Tracking": ["T1", "T2", "T3", "T4"],
        "Part/Component Reported Product Code": ["AP10A-65PSU"] * 4,
        "Serial Number": ["S1", "S2", "S3", "S4"],
        "Quantity": [3, None, 0, 2],
    }).to_excel(source, index=False)

    result = storage_builder.analyze_fedex(
        source,
        datetime(2026, 8, 1),
        datetime(2026, 8, 31),
        {"PSU": 7},
        8,
        log=lambda _msg: None,
    )

    assert result["small_part_picks"] == 7
    assert result["programming_df"]["Quantity"].tolist() == [3, 1, 1, 2]
    assert result["part_type_totals"] == {"PSU": 7}


def test_storage_invoice_excludes_parts_testing_charges(tmp_path, monkeypatch):
    template = tmp_path / "storage-template.xlsx"
    output = tmp_path / "storage-output.xlsx"
    _storage_template(template)

    line_prices = dict(storage_builder.DEFAULT_LINE_PRICES)
    monkeypatch.setattr(storage_builder, "TEMPLATE_FILE", template)
    monkeypatch.setattr(
        storage_builder,
        "load_prices",
        lambda: ({"PSU": 999.0}, line_prices),
    )

    analysis = {
        "unit_storage_df": pd.DataFrame([
            {"ActualModel": "MODEL", "Actual Serial": "INV-1", "Storage": 8.0}
        ]),
        "units_received": pd.DataFrame([
            {"Model": "MODEL", "Serial Number": "RECV-1"}
        ]),
        "programming_df": pd.DataFrame([
            {
                "MSO": "M1",
                "Request Date": pd.Timestamp("2026-08-01"),
                "Outbound Tracking": "TRACK",
                "Part #": "AP10A-65PSU",
                "Type": "PSU",
                "Serial": "PART-1",
                "Quantity": 3,
                "Individual Part Fee": 999.0,
                "Total Programming Fee": 2997.0,
                "Part Pick Fee": 8.0,
            }
        ]),
        "part_type_totals": {"PSU": 3},
        "ship_month": pd.DataFrame([
            {
                "Ticket Number": "MSO-1",
                "Pickup Date": pd.Timestamp("2026-08-01"),
                "Model": "MODEL",
                "Serial Number": "SHIP-1",
                "Tracking Number": "TRACK",
                "Sales Order Number": "SO-1",
            }
        ]),
        "unit_picks_count": 1,
        "small_part_picks": 3,
        "pallet_count": 1,
        "auto_spc_rows": [],
        "unmatched_df": pd.DataFrame(columns=["_Source"]),
    }

    result = storage_builder.build_storage_invoice(
        analysis=analysis,
        invoice_date=datetime(2026, 8, 31),
        completed_date=datetime(2026, 8, 31),
        call_id="CALL",
        customer="Promethean",
        output_path=output,
        log=lambda _msg: None,
    )

    expected = 8.0 + 23.5 + 15.0 + 8.0 + (3 * 8.0)
    assert result["subtotal"] == expected

    wb = load_workbook(output, data_only=False)
    ws = wb["Breakdown"]
    labels = [ws.cell(row=row, column=1).value for row in range(1, ws.max_row + 1)]
    assert not any("Parts Testing" in str(label) for label in labels if label)
    assert ws["D17"].value == "=SUM('Part Testing & Programming'!G:G)"
    assert ws["A19"].value is None
    assert ws["E31"].value == "=SUM(E8:E30)"


def test_workshop_rebuild_restores_complete_corrected_issue_row():
    raw_df = pd.DataFrame({
        "Date Integer": [pd.Timestamp("2026-08-05")],
        "Actual Model": ["RAW-MODEL"],
        "Actual Serial": ["SERIAL-1"],
        "Result": ["Repaired"],
        "Category": ["Refurbished"],
        "Derive Size": ["unknown"],
        "_clean_model": ["AP10-B75-NA-R"],
        "_size": [None],
        "Size": [None],
        "_Type": ["Depot Repair Tab"],
        "_Type2": ["Basic"],
    }, index=[4])

    clean_df, corrected_raw = portal_app._rebuild_workshop_clean_df(
        raw_df,
        {"4": {"field": "Derive Size", "value": "75"}},
        issue_indices=[4],
    )

    assert list(clean_df.index) == [4]
    row = clean_df.loc[4]
    assert row["Actual Model"] == "AP10-B75-NA-R"
    assert row["Actual Serial"] == "SERIAL-1"
    assert row["Result"] == "Repaired"
    assert row["Category"] == "Refurbished"
    assert row["Type"] == "Depot Repair Tab"
    assert row["Type2"] == "Basic"
    assert row["Size"] == "Small"
    assert corrected_raw.loc[4, "Derive Size"] == "75"

    unresolved, _ = portal_app._rebuild_workshop_clean_df(
        raw_df, {}, issue_indices=[4])
    assert unresolved.empty


def test_apply_dedup_returns_excluded_serial_audit(tmp_path):
    master = tmp_path / "master.xlsx"
    shipping = tmp_path / "shipping.csv"

    with pd.ExcelWriter(master, engine="openpyxl") as writer:
        pd.DataFrame([
            {
                "Month": pd.Timestamp("2026-01-01"),
                "Model": "MODEL",
                "Serial": "DUP-1",
                "Type": "Basic",
            }
        ]).to_excel(writer, sheet_name="Repair Log", index=False)
        pd.DataFrame(columns=["Date", "Model", "Serial", "Type"]).to_excel(
            writer, sheet_name="Triage Log", index=False)

    pd.DataFrame([
        {"Serial Number": "OTHER", "Shipped Date": "01-15-2026"}
    ]).to_csv(shipping, index=False)

    repair_df = pd.DataFrame([
        {
            "Date Integer": pd.Timestamp("2026-02-10"),
            "Actual Model": "MODEL",
            "Actual Serial": "DUP-1",
            "Type": "Depot Repair Tab",
            "Type2": "Basic",
            "Size": "Small",
            "Category": "Refurbished",
            "Result": "Repaired",
        }
    ])

    depot, triage, excluded = portal_app.apply_dedup(
        repair_df,
        master,
        shipping,
        log=lambda _msg: None,
        billing_start=datetime(2026, 2, 1),
    )

    assert depot.empty
    assert triage.empty
    assert len(excluded) == 1
    assert excluded.iloc[0]["Serial"] == "DUP-1"
    assert excluded.iloc[0]["Source Tab"] == "Depot Repair"
    assert "no later shipment" in excluded.iloc[0]["Reason"]


def test_finish_passes_excluded_dataframe_to_builder(tmp_path, monkeypatch):
    template = tmp_path / "template.xlsx"
    template.touch()
    captured = {}

    def fake_build(**kwargs):
        captured.update(kwargs)

    monkeypatch.setattr(builder, "TEMPLATE_FILE", template)
    monkeypatch.setattr(builder, "build", fake_build)

    empty_depot = pd.DataFrame(columns=["Type2", "Size", "was_prev_triaged"])
    empty_triage = pd.DataFrame(columns=["Type2", "Size", "was_prev_triaged"])
    excluded = pd.DataFrame([{"Serial": "DUP-1"}])
    done_payload = {}

    portal_app._finish(
        empty_depot,
        empty_triage,
        tmp_path / "invoice.xlsx",
        datetime(2026, 8, 31),
        datetime(2026, 8, 31),
        "CALL",
        "Promethean",
        excluded_df=excluded,
        done=lambda success, payload: done_payload.update(
            {"success": success, **payload}),
        log=lambda _msg: None,
    )

    assert captured["excluded_df"] is excluded
    assert done_payload["success"] is True


def test_session_store_registers_output_module():
    portal_app._OUTPUT_ACCESS.clear()
    store = portal_app.SessionStore("amc")
    q = store.new_queue("session")
    _log, done = portal_app.SessionStore.make_logger(q)
    done(True, {"filename": "AMC_Invoice.xlsx"})
    assert portal_app._registered_output_subsection("AMC_Invoice.xlsx") == "amc"
    portal_app._OUTPUT_ACCESS.clear()


def test_download_authorization_uses_registered_module(tmp_path, monkeypatch):
    output = tmp_path / "amc.xlsx"
    output.write_bytes(b"test workbook")
    monkeypatch.setattr(portal_app, "OUTPUT_DIR", tmp_path)
    portal_app._OUTPUT_ACCESS.clear()
    portal_app._register_output(output, "amc")

    monkeypatch.setattr(
        portal_app.portal_auth,
        "get_current_user",
        lambda: {"id": 1, "username": "amc-user", "is_superadmin": 0},
    )
    monkeypatch.setattr(
        portal_app.portal_auth,
        "has_access",
        lambda _user, section, subsection=None: (
            section == "invoice-generator" and subsection == "amc"),
    )

    with portal_app.app.test_client() as client:
        allowed = client.get("/download/amc.xlsx")
        assert allowed.status_code == 200

        monkeypatch.setattr(
            portal_app.portal_auth,
            "has_access",
            lambda _user, _section, _subsection=None: False,
        )
        denied = client.get("/download/amc.xlsx")
        assert denied.status_code == 302

        unknown = client.get("/download/unregistered.xlsx")
        assert unknown.status_code == 404

    portal_app._OUTPUT_ACCESS.clear()


def test_config_endpoints_require_config_permission():
    config_endpoints = {
        "get_serial_rules",
        "save_serial_rules",
        "get_storage_prices",
        "set_storage_prices",
        "get_fedex_shipment_defaults",
        "set_fedex_shipment_defaults",
        "get_philips_dimensions",
        "upload_philips_dimensions",
        "download_philips_dimensions",
        "get_philips_repair_cost",
        "set_philips_repair_cost",
        "get_amc_dimensions",
        "upload_amc_dimensions",
        "download_amc_dimensions",
        "get_amc_prices",
        "set_amc_prices",
    }

    for endpoint in config_endpoints:
        assert portal_app.ROUTE_SECTIONS[endpoint] == (
            "invoice-generator", "config")

    assert "download" not in portal_app.ROUTE_SECTIONS
    assert portal_app.ROUTE_SECTIONS["analyze_amc_route"] == (
        "invoice-generator", "amc")
    assert portal_app.ROUTE_SECTIONS["analyze_philips_route"] == (
        "invoice-generator", "philips")
