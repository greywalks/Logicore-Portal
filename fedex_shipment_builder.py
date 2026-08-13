"""
fedex_shipment_builder.py — Builds the FedEx Shipment Upload (Promethean)

Not an invoice like the other modules — this transforms FedEx's raw monthly
billing export into a bulk "call" import file for the ticketing system
(NGSC-style import — see the template's own "Instructions - please read"
sheet). Each FedEx shipment line becomes one call/ticket billed to Promethean
at a markup over what FedEx actually charged.

Input each month: the raw FedEx billing export (one sheet, one row per
tracking ID, ~210 columns — only a handful are used).

Pricing rule (confirmed against USSI's actual June 2026 upload, Aug 2026):
  Event1Price = floor(Net Charge Amount / margin_divisor, 2 decimals)
  Default margin_divisor = 0.85 (i.e. a ~15% markup on FedEx's net charge).
  This is a FLOOR/truncation, not a round-half-up — verified against all 260
  rows of the June 2026 sample with zero mismatches at 2-decimal truncation
  vs. several mismatches under naive rounding.

Row → column mapping (verified 1:1 against the June 2026 sample, 260/260
rows matched exactly):
  - SiteName / MainContact / Address / Address2 / City / State / ZipCode
    come from the raw file's Recipient Company / Name / Address Line 1 /
    Address Line 2 / City / State / Zip Code. When Recipient Company (and
    Name) are blank — this happens for shipments routed back to USSI's own
    dock — they fall back to the static site-info defaults below (USSI's own
    office), which is what the raw file's Recipient Address already pointed
    at in every observed case.
  - CustomerPO comes from "Original Customer Reference" (falls back to
    "Original Ref#3/PO Number" if that's blank — not observed in the sample,
    but the two matched in every case where both were populated).
  - Summary = "<period_label> FedEx Invoices - <tracking id>".
  - Every other destination column is a static value that doesn't vary by
    shipment (OrgID, CustomerID, BillCustomerID, CallType, CallRcvdTime,
    Status, QueueID, DUE, Tech, Event1) — see fedex_shipment_defaults.json.
  - CallRcvd / Due By Date are NOT derived from the raw file (its own
    "Invoice Date"/"Shipment Date" columns vary per FedEx sub-invoice within
    the same month) — they're a single date supplied per run, matching the
    June sample where every row shared one date.

ZIP codes: the raw file sometimes ships US zip+4 concatenated with no
separator (e.g. "32940754199" instead of "32940-7541") — truncated to the
first 5 digits when the value is longer than 5 chars and purely numeric.
Non-numeric (e.g. Canadian) postal codes are passed through unchanged.
"""

import json
import math
import shutil
from pathlib import Path
from datetime import datetime, date, time

import pandas as pd
from openpyxl import load_workbook

DEFAULTS_FILE = Path(__file__).parent / "fedex_shipment_defaults.json"
TEMPLATE_FILE = Path(__file__).parent / "template" / "FedEx_Shipment_Upload_Template.xlsx"

DEFAULT_DEFAULTS = {
    "margin_divisor": 0.85,
    "event1_code": "PMTH-SHIP",
    "site_name": "UNITED SERVICE SOURCE",
    "org_id": "Promethean",
    "customer_id": "Promethean",
    "address": "7195 WAELTI DR",
    "address2": "STE 101",
    "city": "MELBOURNE",
    "state": "FL",
    "zip_code": "32940",
    "main_contact": "MATT SHAW",
    "bill_customer_id": "Promethean",
    "call_type": "DEPOT",
    "call_rcvd_time": "17:00:00",
    "status": "CMP",
    "queue_id": "DIGIMED",
    "due": "BY",
    "tech": "E-ERWE",
}

RAW_COLS = {
    "tracking":     "Express or Ground Tracking ID",
    "net_charge":   "Net Charge Amount",
    "payor":        "Payor",
    "recip_company": "Recipient Company",
    "recip_name":   "Recipient Name",
    "recip_addr1":  "Recipient Address Line 1",
    "recip_addr2":  "Recipient Address Line 2",
    "recip_city":   "Recipient City",
    "recip_state":  "Recipient State",
    "recip_zip":    "Recipient Zip Code",
    "orig_cust_ref": "Original Customer Reference",
    "orig_ref3":    "Original Ref#3/PO Number",
}


def load_defaults() -> dict:
    """Merged config: bundled defaults + any live overrides in
    fedex_shipment_defaults.json (same pattern as amc_builder.load_prices())."""
    import sys
    cfg = dict(DEFAULT_DEFAULTS)
    if DEFAULTS_FILE.exists():
        try:
            cfg.update(json.loads(DEFAULTS_FILE.read_text()))
        except Exception as e:
            print(f"[fedex_shipment_builder] WARNING: could not parse {DEFAULTS_FILE} — "
                  f"using bundled defaults only. Error: {e}", file=sys.stderr)
    return cfg


def save_defaults(cfg: dict):
    DEFAULTS_FILE.write_text(json.dumps(cfg, indent=2))


def _s(v) -> str:
    """Safe string coercion — pandas leaves truly empty cells as float NaN
    even under dtype=str, which has no .strip()."""
    if v is None:
        return ""
    try:
        if isinstance(v, float) and pd.isna(v):
            return ""
    except TypeError:
        pass
    return str(v).strip()


def _clean_zip(z) -> str:
    """Truncates to the leading run of digits, capped at 5 (handles both
    zip+4 concatenated with no separator, e.g. "32940754199", and stray
    trailing junk like "32940-7541-26"). Non-numeric-leading values (e.g.
    Canadian postal codes) pass through unchanged."""
    s = _s(z)
    if s.endswith(".0"):  # pandas sometimes reads a numeric zip as float
        s = s[:-2]
    lead_digits = ""
    for ch in s:
        if ch.isdigit():
            lead_digits += ch
        else:
            break
    if len(lead_digits) >= 5:
        return lead_digits[:5]
    return s


def _parse_call_date(d):
    """Accepts a date/datetime already, or a 'YYYY-MM-DD' string (what the
    frontend's <input type=date> sends), and returns a datetime — the
    template's own June sample stores this as a real Excel datetime, not
    text, so the ticketing system's importer likely expects that type."""
    if isinstance(d, datetime):
        return d
    if isinstance(d, date):
        return datetime(d.year, d.month, d.day)
    return datetime.strptime(str(d).strip(), "%Y-%m-%d")


def _parse_time(t):
    """Accepts a time already, or an 'HH:MM:SS'/'HH:MM' string."""
    if isinstance(t, time):
        return t
    s = str(t).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Unrecognized time format: {t!r}")


def analyze_fedex_shipment(raw_path, period_label, call_date, defaults=None, log=print):
    """Reads the raw FedEx export and computes what will be billed, without
    writing anything. Returns a dict with the built rows plus review info —
    rows where recipient info was blank (defaulted to USSI's own office) and
    any rows that had to be skipped (no tracking ID or no charge to bill)."""
    if defaults is None:
        defaults = load_defaults()

    log(f"Reading raw FedEx export: {Path(raw_path).name}")
    df = pd.read_excel(raw_path, sheet_name=0, dtype=str)

    missing = [c for c in RAW_COLS.values() if c not in df.columns]
    if missing:
        raise ValueError(
            "Raw FedEx file is missing expected column(s): " + ", ".join(missing) +
            ". Make sure this is the raw FedEx billing export, not an already-transformed file.")

    log(f"{len(df)} row(s) in raw export.")

    divisor = float(defaults.get("margin_divisor", 0.85))
    call_dt = _parse_call_date(call_date)
    call_time = _parse_time(defaults["call_rcvd_time"])
    out_rows = []
    defaulted_rows = []
    skipped_rows = []
    seen_tids = set()

    for i, row in df.iterrows():
        tid = _s(row.get(RAW_COLS["tracking"]))
        net_raw = _s(row.get(RAW_COLS["net_charge"]))
        try:
            net = float(net_raw) if net_raw else None
        except (TypeError, ValueError):
            net = None

        if not tid:
            skipped_rows.append({"row": int(i) + 2, "reason": "No tracking ID"})
            continue
        if net is None or net == 0:
            skipped_rows.append({"row": int(i) + 2, "tracking": tid, "reason": "No Net Charge Amount"})
            continue
        if tid in seen_tids:
            skipped_rows.append({"row": int(i) + 2, "tracking": tid, "reason": "Duplicate tracking ID — already billed once this run"})
            continue
        seen_tids.add(tid)

        company = _s(row.get(RAW_COLS["recip_company"]))
        name    = _s(row.get(RAW_COLS["recip_name"]))
        addr1   = _s(row.get(RAW_COLS["recip_addr1"]))
        # "used_default" = shipment has no recipient company/name at all (routed
        # back to USSI's own dock rather than out to an end customer) — the
        # WHOLE site-info block gets replaced with USSI's own canonical office
        # info. Verified against the June 2026 sample: when company+name ARE
        # present (even for other USSI-addressed shipments, e.g. "USSI -
        # PROMETHEAN"), the raw recipient fields are passed through as typed —
        # no casing/name normalization is applied. Don't try to detect
        # "is this USSI's own office" from the address/zip; the real upload
        # only normalizes when both company and name are blank.
        used_default = not company and not name

        po = _s(row.get(RAW_COLS["orig_cust_ref"]))
        if not po:
            po = _s(row.get(RAW_COLS["orig_ref3"]))

        price = math.floor(net / divisor * 100) / 100  # truncate to 2 decimals, not round

        if used_default:
            site_name, main_contact = defaults["site_name"], defaults["main_contact"]
            address, address2 = defaults["address"], defaults["address2"]
            city, state = defaults["city"], defaults["state"]
            zip_code = defaults["zip_code"]
        else:
            site_name, main_contact = company, name
            address = addr1 or defaults["address"]
            address2 = _s(row.get(RAW_COLS["recip_addr2"]))
            city = _s(row.get(RAW_COLS["recip_city"])) or defaults["city"]
            state = _s(row.get(RAW_COLS["recip_state"])) or defaults["state"]
            zip_code = _clean_zip(row.get(RAW_COLS["recip_zip"])) or defaults["zip_code"]
            # Fields that case-insensitively match USSI's own canonical
            # site info get normalized to that canonical casing, even when
            # the rest of the row's recipient info is passed through as
            # typed. Confirmed on two independent fields in the June sample
            # (MainContact "Matt Shaw"→"MATT SHAW" on a row where SiteName
            # stayed raw "USSI - PROMETHEAN"; SiteName "United Service
            # Source"→"UNITED SERVICE SOURCE" on another row) — so this is
            # applied per-field, not as an all-or-nothing site match.
            if site_name.strip().lower() == defaults["site_name"].strip().lower():
                site_name = defaults["site_name"]
            if main_contact.strip().lower() == defaults["main_contact"].strip().lower():
                main_contact = defaults["main_contact"]
            if address.strip().lower() == defaults["address"].strip().lower():
                address = defaults["address"]
            if address2.strip().lower() == defaults["address2"].strip().lower():
                address2 = defaults["address2"]
            if city.strip().lower() == defaults["city"].strip().lower():
                city = defaults["city"]
            if state.strip().lower() == defaults["state"].strip().lower():
                state = defaults["state"]

        built = {
            "SiteID": None,
            "SiteName": site_name,
            "OrgID": defaults["org_id"],
            "CustomerID": defaults["customer_id"],
            "Address": address,
            "Address2": address2,
            "City": city,
            "State": state,
            "ZipCode": zip_code,
            "MainContact": main_contact,
            "BillCustomerID": defaults["bill_customer_id"],
            "CustomerPO": po,
            "CallType": defaults["call_type"],
            "CallRcvd": call_dt,
            "CallRcvdTime": call_time,
            "Status": defaults["status"],
            "QueueID": defaults["queue_id"],
            "DUE": defaults["due"],
            "Due By Date": call_dt,
            "Summary": f"{period_label} FedEx Invoices - {tid}",
            "Tech": defaults["tech"],
            "Event1": defaults["event1_code"],
            "Event1Price": price,
        }
        out_rows.append(built)
        if used_default:
            defaulted_rows.append({"tracking": tid, "po": po})

    log(f"Built {len(out_rows)} call row(s); {len(defaulted_rows)} used the default USSI site info; "
        f"{len(skipped_rows)} skipped.")

    return {
        "rows": out_rows,
        "defaulted_rows": defaulted_rows,
        "skipped_rows": skipped_rows,
        "total_price": round(sum(r["Event1Price"] for r in out_rows), 2),
        "period_label": period_label,
        "call_date": call_date,
    }


def build_fedex_shipment_upload(analysis, output_path, log=print):
    """Writes analysis['rows'] into a copy of the upload template."""
    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE_FILE}")

    log("Copying upload template…")
    shutil.copy(TEMPLATE_FILE, output_path)

    wb = load_workbook(output_path)
    ws = wb["_uploadsheet"]
    header_row = [c.value for c in ws[1]]
    col_of = {}
    for idx, h in enumerate(header_row, start=1):
        if h and h not in col_of:   # first occurrence wins (template has a duplicate "MainContact" header)
            col_of[h] = idx

    rows = analysis["rows"]
    log(f"Writing {len(rows)} row(s)…")
    date_cols = {"CallRcvd", "Due By Date"}
    time_cols = {"CallRcvdTime"}
    for r_offset, built in enumerate(rows, start=2):
        for field, value in built.items():
            col = col_of.get(field)
            if col is None:
                continue
            cell = ws.cell(row=r_offset, column=col, value=value)
            if field in date_cols:
                cell.number_format = "mm-dd-yy"
            elif field in time_cols:
                cell.number_format = "hh:mm:ss"

    wb.save(output_path)
    log(f"Saved: {Path(output_path).name}")

    return {
        "row_count": len(rows),
        "total_price": analysis["total_price"],
        "defaulted_count": len(analysis["defaulted_rows"]),
        "skipped_count": len(analysis["skipped_rows"]),
    }
