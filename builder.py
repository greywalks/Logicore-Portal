"""
Invoice builder — pixel-perfect match to Sample_Promethean_Workshop_Invoice.xlsx
"""

import shutil
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.formatting.rule import ColorScaleRule
import sys as _sys
_sys.path.insert(0, str(Path(__file__).parent))
from normalizer import normalize_result
# Parts Testing & Configuration now bills on the Workshop invoice (moved from
# Storage — see PROJECT_BRIEF.md). Reuse Storage's FedEx classification/pricing
# helpers and its "Part Testing & Programming" sheet writer so both invoices
# stay byte-for-byte consistent on that shared data.
import storage_builder

TEMPLATE_FILE = Path(__file__).parent / "template" / "Sample_Promethean_Workshop_Invoice.xlsx"

# Exact accounting format from original
ACCT_FMT  = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
INT_FMT   = '#,##0'
MONEY_FMT = '"$"#,##0.00'

# Theme-based fills matching the original exactly
def _theme_fill(theme: int, tint: float = 0.0):
    c = Color(theme=theme, tint=tint, type="theme")
    f = PatternFill(fill_type="solid")
    f.fgColor = c
    return f

HDR_FILL   = _theme_fill(0, tint=-0.249977111117893)   # row 10 dark header
TOTAL_FILL = _theme_fill(4, tint=0.7999816888943144)   # row 47 orange-ish total

def _theme_font(theme=1, bold=False, size=10, underline=None):
    c = Color(theme=theme, type="theme")
    return Font(name="Calibri", bold=bold, size=size, color=c,
                underline=underline)


def build(depot_df, triage_df, output_path: Path,
          invoice_date: datetime, completed_date: datetime,
          call_id: str, customer: str, excluded_df=None,
          programming_df=None, part_prices=None):

    shutil.copy(TEMPLATE_FILE, output_path)
    wb = load_workbook(output_path)

    if programming_df is None:
        programming_df = pd.DataFrame(columns=[
            "MSO","Request Date","Outbound Tracking","Part #","Type","Serial",
            "Quantity","Individual Part Fee","Total Programming Fee","Part Pick Fee"])
    if part_prices is None:
        part_prices, _ = storage_builder.load_prices()

    _build_breakdown(wb, depot_df, triage_df,
                     invoice_date, completed_date, call_id, customer, part_prices)
    _build_depot(wb, depot_df)
    _build_triage(wb, triage_df)
    storage_builder._build_part_testing(wb, programming_df)
    _build_excluded(wb, excluded_df)

    # Expand Excel Table objects to cover all written rows
    from openpyxl.utils import get_column_letter
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            # Get current data extent on this sheet
            last_data_row = ws.max_row
            if last_data_row < 2:
                continue
            # Parse current ref to get start cell and column range
            import re as _re
            m = _re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', tbl.ref)
            if m:
                col_start, row_start, col_end, _ = m.groups()
                tbl.ref = f"{col_start}{row_start}:{col_end}{last_data_row}"

    # Strip all conditional formatting — output must be client-ready
    for sheet in wb.worksheets:
        sheet.conditional_formatting._cf_rules.clear()
    wb.save(output_path)


# ── Breakdown ─────────────────────────────────────────────────────────────────

def _build_breakdown(wb, depot_df, triage_df,
                     invoice_date, completed_date, call_id, customer, part_prices):
    ws = wb["Breakdown"]
    ws.delete_rows(1, ws.max_row)

    # ── Column widths (exact from original) ───────────────────────────────────
    ws.column_dimensions["A"].width = 44.5546875
    ws.column_dimensions["B"].width = 14.5546875
    ws.column_dimensions["C"].width = 17.44140625
    ws.column_dimensions["D"].width = 16.44140625
    ws.column_dimensions["E"].width = 16.0

    # ── Address block (rows 1-6) ───────────────────────────────────────────────
    def addr(row, val, col="A"):
        c = ws.cell(row=row, column=ord(col)-64, value=val)
        c.font = _theme_font(theme=1, size=8)
        return c

    def label(row, col, val):
        c = ws.cell(row=row, column=ord(col)-64, value=val)
        c.font = _theme_font(theme=1, bold=True, size=9)
        c.alignment = Alignment(horizontal="right")
        return c

    def value(row, col, val, fmt=None, as_date=False):
        c = ws.cell(row=row, column=ord(col)-64, value=val)
        c.font = _theme_font(theme=1, size=9)
        if as_date:
            c.number_format = "mm-dd-yy"
        elif fmt:
            c.number_format = fmt
        return c

    addr(1, "9145 Ellis Road")
    addr(2, "Melbourne, FL 32904")
    # www with hyperlink styling
    a3 = addr(3, "www.ussiglobal.com")
    a3.font = _theme_font(theme=10, size=8, underline="single")

    label(1, "D", "Call ID:");    value(1, "E", call_id)
    label(2, "D", "Currency:");   value(2, "E", "USD")
    label(3, "D", "Invoice Date:");    value(3, "E", invoice_date, as_date=True)
    label(4, "D", "Completed Date:"); value(4, "E", completed_date, as_date=True)
    label(5, "D", "Customer:");   value(5, "E", customer)
    label(6, "D", "Terms:");      value(6, "E", "Net Amt Due in 45")

    # ── Title row 8 (merged A8:E8, formula, centered) ─────────────────────────
    ws.row_dimensions[8].height = 21.0
    ws.merge_cells("A8:E8")
    a8 = ws["A8"]
    a8.value = '=CONCATENATE("PROFORMA INVOICE: ",TEXT(E4,"MMMM YYYY"))'
    a8.font = _theme_font(theme=1, bold=True, size=16)
    a8.alignment = Alignment(horizontal="center", vertical="center")

    # ── Column header row 10 ──────────────────────────────────────────────────
    headers = ["Workshop", "Units of Measure", "Unit Price", "Quantity", "Total Amount"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=10, column=ci, value=h)
        c.fill = HDR_FILL
        c.font = _theme_font(theme=1, bold=True, size=10)
        c.alignment = Alignment(horizontal="center" if ci > 1 else None)

    # ── Line items ─────────────────────────────────────────────────────────────
    # Layout: rows 11-14 T&R, blank 15, 16-19 Triage, blank 20,
    # 21-24 Repair-only, blank 25, 26 Reboxing, blank 27,
    # 28-29 Scrap, blank 30, 31 Special, blank 32,
    # 33 Parts Testing & Configuration header, 34-43 part-type lines, blank 44,
    # totals 45-47.

    LINES = [
        # (row, label, uom, price, qty_formula)
        (11, "Standard \u2018Basic\u2019 Panel Triage & Repair Small (65\u201d-75\u201d)", "Each", 110,
         "=COUNTIFS('Depot Repair'!C:C,\"BASIC\",'Depot Repair'!D:D,\"Small\")"),
        (12, "Standard \u2018Basic\u2019 Panel Triage & Repair Large (86\u201d)", "Each", 135,
         "=COUNTIFS('Depot Repair'!C:C,\"BASIC\",'Depot Repair'!D:D,\"LARGE\")"),
        (13, "Advanced \u2018Heavy\u2019 Panel Triage & Repairs Small (65\u201d-75\u201d)", "Each", 220,
         "=COUNTIFS('Depot Repair'!C:C,\"Heavy\",'Depot Repair'!D:D,\"Small\")"),
        (14, "Advanced \u2018Heavy\u2019 Panel Triage & Repairs Large (86\u201d)", "Each", 268,
         "=COUNTIFS('Depot Repair'!C:C,\"Heavy\",'Depot Repair'!D:D,\"Large\")"),

        (16, "Standard \u2018Basic\u2019 Panel Triage Small (65\u201d-75\u201d)", "Each", 86,
         "=COUNTIFS('Triage Units'!D:D,\"BasicSmall-Triage\")"),
        (17, "Standard \u2018Basic\u2019 Panel Triage Large (86\u201d)", "Each", 101,
         "=COUNTIFS('Triage Units'!D:D,\"BasicLarge-Triage\")"),
        (18, "Advanced \u2018Heavy\u2019 Panel Triage Small (65\u201d-75\u201d)", "Each", 152,
         "=COUNTIFS('Triage Units'!D:D,\"HeavySmall-Triage\")"),
        (19, "Advanced \u2018Heavy\u2019 Panel Triage Large (86\u201d)", "Each", 181,
         "=COUNTIFS('Triage Units'!D:D,\"HeavyLarge-Triage\")"),

        (21, "Standard \u2018Basic\u2019 Panel Repair Small (65\u201d-75\u201d)", "Each", 64,
         "=COUNTIFS('Depot Repair'!C:C,\"BasicSmall - Previously Triaged\")"),
        (22, "Standard \u2018Basic\u2019 Panel Repair Large (86\u201d)", "Each", 74,
         "=COUNTIFS('Depot Repair'!C:C,\"BasicLarge - Previously Triaged\")"),
        (23, "Advanced \u2018Heavy\u2019 Panel Repair Small (65\u201d-75\u201d)", "Each", 108,
         "=COUNTIFS('Depot Repair'!C:C,\"HeavySmall - Previously Triaged\")"),
        (24, "Advanced \u2018Heavy\u2019 Panel Repair Large (86\u201d)", "Each", 127,
         "=COUNTIFS('Depot Repair'!C:C,\"HeavyLarge - Previously Triaged\")"),

        (26, "Reboxing Fee", "Each", 50, None),          # hardcoded 0
        (28, "Immediate Scrap", "Each", 19,
         "=COUNTIFS('Depot Repair'!C:C,\"Immediate Scrap\")"),
        (29, "Salvage of Hardware and Scrap", "Each", 28,
         "=COUNTIFS('Depot Repair'!C:C,\"Salvage of Hardware and Scrap\")"),
        (31, "Special Warehouse Projects", "Hourly", 75, None),  # hardcoded 0
    ]

    for (row, lbl, uom, price, qty_fml) in LINES:
        # A: label
        ca = ws.cell(row=row, column=1, value=lbl)
        ca.font = _theme_font(size=10)

        # B: uom
        cb = ws.cell(row=row, column=2, value=uom)
        cb.font = _theme_font(size=10)

        # C: unit price
        cc = ws.cell(row=row, column=3, value=price)
        cc.font = _theme_font(size=10)
        cc.number_format = ACCT_FMT

        # D: quantity
        cd = ws.cell(row=row, column=4)
        if qty_fml:
            cd.value = qty_fml
        else:
            cd.value = 0
        cd.font = _theme_font(size=10)
        cd.number_format = INT_FMT

        # E: total
        ce = ws.cell(row=row, column=5, value=f"=D{row}*C{row}")
        ce.font = _theme_font(size=10)
        ce.number_format = ACCT_FMT

    # ── Parts Testing & Configuration (moved here from the Storage invoice) ────
    # Row 33 section header, matching the row-10 header style.
    pt_headers = ["Parts Testing & Configuration", "Units of Measure", "Unit Price", "Quantity", "Total Amount"]
    for ci, h in enumerate(pt_headers, 1):
        c = ws.cell(row=33, column=ci, value=h)
        c.fill = HDR_FILL
        c.font = _theme_font(theme=1, bold=True, size=10)
        c.alignment = Alignment(horizontal="center" if ci > 1 else None)

    # Quantities are live SUMIFS formulas against the TblPartTesting table
    # (written by storage_builder._build_part_testing on this same workbook),
    # not hardcoded Python-computed values — so the sheet recalculates if the
    # underlying Part Testing & Programming data is ever hand-edited.
    PT_LINES = [
        (34, "PSU - Testing",                              "Each", "PSU"),
        (35, "Mainboard Configure for Dispatch - Testing", "Each", "Mainboard Configure for Dispatch"),
        (36, "Mainboard - Testing",                        "Each", "Mainboard"),
        (37, "AC-PCA - Testing",                           "Each", "AC-PCA"),
        (38, "Keypad - Testing",                           "Each", "Keypad"),
        (39, "Maintouch - Testing",                        "Each", "Maintouch"),
        (40, "Ext-Input - Testing",                        "Each", "EXT-INPUT"),
        (41, "OPS-PCA - Testing",                          "Each", "OPS-PCA"),
        (42, "USB - Testing",                              "Each", "USB"),
        (43, "Speaker Testing",                            "Each", "SPEAKER"),
    ]
    for row, lbl, uom, ptype in PT_LINES:
        price = part_prices.get(ptype, 0)
        qty_fml = f'=SUMIFS(TblPartTesting[Quantity],TblPartTesting[Type],"{ptype}")'

        ca = ws.cell(row=row, column=1, value=lbl); ca.font = _theme_font(size=10)
        cb = ws.cell(row=row, column=2, value=uom); cb.font = _theme_font(size=10)
        cc = ws.cell(row=row, column=3, value=price)
        cc.font = _theme_font(size=10); cc.number_format = ACCT_FMT
        cd = ws.cell(row=row, column=4, value=qty_fml)
        cd.font = _theme_font(size=10); cd.number_format = INT_FMT
        ce = ws.cell(row=row, column=5, value=f"=D{row}*C{row}")
        ce.font = _theme_font(size=10); ce.number_format = ACCT_FMT

    # ── Totals (rows 45-47) ─────────────────────────────────────────────────────
    # Subtotal
    ws["D45"].value = "Subtotal"
    ws["D45"].font  = _theme_font(bold=True, size=10)
    ws["E45"].value = "=SUM(E11:E43)"
    ws["E45"].font  = _theme_font(size=10)
    ws["E45"].number_format = ACCT_FMT

    # Tax
    ws["D46"].value = "Tax"
    ws["D46"].font  = _theme_font(bold=True, size=10)
    ws["E46"].value = "=E45*7%"
    ws["E46"].font  = _theme_font(size=10)
    ws["E46"].number_format = ACCT_FMT

    # Total
    ws["D47"].value = "Total"
    ws["D47"].font  = _theme_font(bold=True, size=10)
    ws["E47"].value = "=SUM(E45:E46)"
    ws["E47"].font  = _theme_font(bold=True, size=10)
    ws["E47"].fill  = TOTAL_FILL
    ws["E47"].number_format = ACCT_FMT


# ── Depot Repair sheet ────────────────────────────────────────────────────────

def _build_depot(wb, depot_df):
    wd = wb["Depot Repair"]
    wd.delete_rows(1, wd.max_row)

    # Exact column widths from original
    for col, w in zip("ABCDEF", [21.44140625, 26.21875, 31.21875,
                                   13.33203125, 20.5546875, 35.44140625]):
        wd.column_dimensions[col].width = w

    # Header row — NOT bold, sz=11, no fill (matches original)
    for ci, h in enumerate(["Model","Serial","Type","Size","Repair Cost","Parts Summary"], 1):
        c = wd.cell(row=1, column=ci, value=h)
        c.font = _theme_font(theme=1, bold=False, size=11)

    # Build type label matching the COUNTIFS keys in Breakdown
    def depot_type(row):
        t2, sz, pt = row["Type2"], row["Size"], row["was_prev_triaged"]
        if t2 == "Salvage of Hardware and Scrap":
            return "Salvage of Hardware and Scrap"
        if pt:
            prefix = "Heavy" if t2 == "Heavy" else "Basic"
            return f"{prefix}{sz} - Previously Triaged"
        # Full triage+repair: COUNTIFS checks "BASIC"/"Heavy" + size
        return t2   # "Basic" or "Heavy"

    df = depot_df.copy()
    df["_type_label"] = df.apply(depot_type, axis=1)
    df = df.sort_values(["_type_label", "Size", "Actual Model"])

    for i, (_, r) in enumerate(df.iterrows(), 2):
        vals = [r["Actual Model"], r["Actual Serial"], r["_type_label"],
                r["Size"], r["Unit Price"], normalize_result(r["Result"])]
        for ci, v in enumerate(vals, 1):
            c = wd.cell(row=i, column=ci, value=v)
            c.font = _theme_font(size=11)
            c.alignment = Alignment(horizontal="left")
            if ci == 5:
                c.number_format = ACCT_FMT


# ── Triage Units sheet ────────────────────────────────────────────────────────

def _build_triage(wb, triage_df):
    wt = wb["Triage Units"]
    wt.delete_rows(1, wt.max_row)

    # Exact column widths from original
    for col, w in zip("ABCDEF", [21.6640625, 25.88671875, 44.88671875,
                                   17.88671875, 15.109375, 13.5546875]):
        wt.column_dimensions[col].width = w

    for ci, h in enumerate(["Model","Serial","Triage","Derived Type","Size","Cost"], 1):
        c = wt.cell(row=1, column=ci, value=h)
        c.font = _theme_font(theme=1, bold=False, size=11)

    df = triage_df.copy()
    df["_derived"] = df.apply(
        lambda r: f"{'Heavy' if r['Type2']=='Heavy' else 'Basic'}{r['Size']}-Triage", axis=1)
    df = df.sort_values(["_derived", "Actual Model"])

    for i, (_, r) in enumerate(df.iterrows(), 2):
        vals = [r["Actual Model"], r["Actual Serial"], normalize_result(r["Result"]),
                r["_derived"], r["Size"], r["Unit Price"]]
        for ci, v in enumerate(vals, 1):
            c = wt.cell(row=i, column=ci, value=v)
            c.font = _theme_font(size=11)
            c.alignment = Alignment(horizontal="left")
            if ci == 6:
                c.number_format = ACCT_FMT


# ── Excluded Serials sheet ────────────────────────────────────────────────────

def _build_excluded(wb, excluded_df):
    """
    Lists every serial that was pulled out of this run and NOT invoiced,
    with a plain-English reason (currently: previously-invoiced duplicates
    that either never shipped again, or shipped before the prior invoice
    date and were never reshipped since).
    """
    SHEET_NAME = "Excluded Serials"
    if SHEET_NAME in wb.sheetnames:
        we = wb[SHEET_NAME]
        we.delete_rows(1, we.max_row)
    else:
        we = wb.create_sheet(SHEET_NAME)

    headers = ["Model", "Serial", "Source Tab", "Category", "Result", "Original Date", "Reason"]
    widths  = [22, 26, 14, 16, 38, 14, 55]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = we.cell(row=1, column=ci, value=h)
        c.font = _theme_font(theme=1, bold=True, size=11)
        we.column_dimensions[chr(64 + ci)].width = w

    if excluded_df is None or len(excluded_df) == 0:
        c = we.cell(row=2, column=1, value="No serials were excluded this run.")
        c.font = _theme_font(size=11)
        return

    for i, (_, r) in enumerate(excluded_df.iterrows(), 2):
        date_val = r.get("Original Date", "")
        date_str = str(date_val)[:10] if date_val not in (None, "") else ""
        vals = [r.get("Model", ""), r.get("Serial", ""), r.get("Source Tab", ""),
                r.get("Category", ""), r.get("Result", ""), date_str, r.get("Reason", "")]
        for ci, v in enumerate(vals, 1):
            c = we.cell(row=i, column=ci, value=v)
            c.font = _theme_font(size=11)
            c.alignment = Alignment(horizontal="left")
