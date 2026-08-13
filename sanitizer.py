"""
sanitizer.py — Derives Type, Type2, Size, and clean Model from raw Repair Data.
Returns (clean_df, issues_list) where issues_list contains rows needing human review.
Also exports a corrected copy of the original workbook.
"""

import re
import shutil
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Size rule ──────────────────────────────────────────────────────────────────
# Everything < 86" is Small; 86" and above is Large.
# 70" panels ARE billable (treated as Small).
BILLABLE_SIZES = {'55', '65', '70', '75', '86'}
LARGE_SIZES    = {'86'}

# ── Result keywords → Heavy ────────────────────────────────────────────────────
HEAVY_KEYWORDS = [
    'lcd', 'lcm', 'pending lcd', 'pending lcm',
    'retaped deflector', 'retape deflector',
    'deflector sheet', 'backlight deflector', 'reflector sheet',
    'overlay replaced', 'lcd and', 'lcm and',
    'lcd, ', 'lcm, ', 'obf lcd', 'pending overlay',
]

# ── Category → Type ────────────────────────────────────────────────────────────
CATEGORY_TYPE = {
    'refurbished':  'Depot Repair Tab',
    'scrap':        'Depot Repair Tab',
    'pending parts':'Triage Tab',
}

# ── Serial prefix → (model_base, size_digits) ─────────────────────────────────
# model_base will have -NA-R appended.
# For 775T / 770T / 786T the -02 suffix is determined by year character.

YEAR_CHAR_02_THRESHOLD = 'L'   # L=2021 onward → -02 revision

# Year character is position 5 in 775T-XNNN... serials (index 5 after the dash)
# i.e. serial[5] for "775T-L..."

def _is_02_revision(serial: str, pos: int = 5) -> bool:
    """Return True if the year letter in the serial indicates -02 revision (L/M/N/...)."""
    # Works for 775T-Xxx, 770T-Xxx, 786T-Xxx patterns
    if len(serial) < 6:
        return False
    year_char = serial[pos].upper() if len(serial) > pos else ""
    return year_char >= YEAR_CHAR_02_THRESHOLD


# ── Load serial rules from config (hot-reloadable) ────────────────────────────
import json as _json
import sys as _sys_warn

def _load_rules():
    cfg = Path(__file__).parent / "serial_rules.json"
    if cfg.exists():
        try:
            return _json.loads(cfg.read_text()).get("rules", [])
        except Exception as e:
            # Falling back to the hardcoded prefix table below is a real,
            # user-visible behavior change (config-driven models/sizes stop
            # applying) — silently swallowing this made a bad hand-edit in
            # the Config page invisible until someone noticed wrong pricing.
            print(f"[sanitizer] WARNING: could not parse {cfg} — "
                  f"falling back to hardcoded serial rules only. Error: {e}",
                  file=_sys_warn.stderr)
    return []


def _serial_to_model_and_size(serial: str):
    """
    Returns (clean_model_with_NA_R, size_str) from a serial number.
    Returns (None, None) if unrecognized.
    """
    s = str(serial).strip()
    su = s.upper()

    # ── Config-driven prefix lookup (longest prefix wins) ─────────────────────
    rules = sorted(_load_rules(), key=lambda r: len(r.get("prefix","")), reverse=True)
    for rule in rules:
        prefix = rule.get("prefix","").upper()
        if not prefix or not su.startswith(prefix):
            continue
        model_base = rule.get("model_base","")
        size       = str(rule.get("size","75"))
        o2_rule    = rule.get("o2_rule","never")
        year_pos   = int(rule.get("year_pos", 5))
        if o2_rule == "always":
            suffix = "-02-NA-R"
        elif o2_rule == "never":
            suffix = "-NA-R"
        else:  # year_pos5 or year_pos6
            suffix = "-02-NA-R" if _is_02_revision(s, pos=year_pos) else "-NA-R"
        # Special cases that need extra logic beyond prefix
        if su.startswith("9A75"):
            if "V" in su[6:10]: return "AP9-A75-V-NA-R", size
        if su.startswith("9B75") and su[4:6] == "GP":
            return "AP9-B75-02-NA-R", size
        return f"{model_base}{suffix}", size

    # ── 4-char prefix dispatch (fallback hardcoded) ───────────────────────────
    p4 = su[:4]
    p3 = su[:3]
    p2 = su[:2]

    # AP6 legacy
    if p4 == '686P':
        return 'AP6-86-4K-R', '86'
    if p4 == '675F':
        return 'AP6-75-4K-R', '75'

    # AP7 U-series (65W/75W/86W)
    if su[:3] == '65W':
        suffix = '-02-NA-R' if _is_02_revision(s, pos=6) else '-NA-R'
        return f'AP7-U65{suffix}', '65'
    if su[:3] == '75W':
        suffix = '-02-NA-R' if _is_02_revision(s, pos=6) else '-NA-R'
        return f'AP7-U75{suffix}', '75'
    if su[:3] == '86W':
        suffix = '-02-NA-R' if _is_02_revision(s, pos=6) else '-NA-R'
        return f'AP7-U86{suffix}', '86'

    # AP7 B-series (770T / 775T / 786T)
    if p4 == '770T' or p4 == '77OT':   # O vs 0 typo
        suffix = '-02-NA-R' if _is_02_revision(s) else '-NA-R'
        return f'AP7-B70{suffix}', '70'
    if p4 == '775T' or p4 == '7751':
        suffix = '-02-NA-R' if _is_02_revision(s) else '-NA-R'
        return f'AP7-B75{suffix}', '75'
    if p4 == '786T':
        suffix = '-02-NA-R' if _is_02_revision(s) else '-NA-R'
        return f'AP7-B86{suffix}', '86'

    # AP9-A series (9A6x / 9A7x / 9A8x)
    if su[:4] == '9A65': return 'AP9-A65-NA-R', '65'
    if su[:4] == '9A76': return 'AP9-A75-NA-R', '75'  # data quirk
    if su[:4] == '9A75':
        # V-variant: serial has NA+V in positions ~4-7 e.g. 9A752NA43V...
        if 'V' in su[6:10]: return 'AP9-A75-V-NA-R', '75'
        return 'AP9-A75-NA-R', '75'
    if su[:4] == '9A86': return 'AP9-A86-NA-R', '86'

    # AP9-B series (9B6x / 9B7x / 9B8x)
    if su[:4] == '9B65': return 'AP9-B65-NA-R', '65'
    if su[:4] == '9B75':
        # GP in positions 4-5 → -02 variant
        if su[4:6] == 'GP': return 'AP9-B75-02-NA-R', '75'
        return 'AP9-B75-NA-R', '75'
    if su[:4] == '9B86' or su[:4] == '9G86': return 'AP9-B86-NA-R', '86'
    if su[:4] == '9G65': return 'AP9-B65-NA-R', '65'  # data entry typo

    # AP10-A series (AA6x / AA7x / AA8x)
    if su[:4] == 'AA65': return 'AP10-A65-NA-R', '65'
    if su[:4] == 'AA75': return 'AP10-A75-NA-R', '75'
    if su[:4] == 'AA86': return 'AP10-A86-NA-R', '86'

    # AP10-B series (AB5x / AB6x / AB7x / AB8x)
    if su[:4] == 'AB55': return 'AP10-B55-NA-R', '55'
    if su[:4] == 'AB65': return 'AP10-B65-NA-R', '65'
    if su[:4] == 'AB75': return 'AP10-B75-NA-R', '75'
    if su[:4] in ('AB86','AB83'): return 'AP10-B86-NA-R', '86'

    # APLE series (BE6x / BE7x / BE8x)
    if su[:4] == 'BE65': return 'APLE-65-NA-R', '65'
    if su[:4] == 'BE75': return 'APLE-75-NA-R', '75'
    if su[:4] == 'BE86': return 'APLE-86-NA-R', '86'

    # APLX series (LX6x / LX7x / LX8x)
    if su[:4] == 'LX65': return 'APLX-65-NA-R', '65'
    if su[:4] in ('LX75','LX57'): return 'APLX-75-NA-R', '75'
    if su[:4] == 'LX86': return 'APLX-86-NA-R', '86'

    # VTP
    if su[:3] == 'V65': return 'VTP-65-NA-R', '65'
    if su[:3] == 'V75': return 'VTP-75-NA-R', '75'

    # AP5
    if su[:3] in ('P70', 'P5-'): return None, None  # AP5 legacy, not billable

    return None, None


def _clean_model(model: str) -> str:
    """
    Normalize a raw model string to end with -NA-R.
    Strips trailing junk, corrects common typos, appends -NA-R if missing.
    """
    if not isinstance(model, str):
        return model
    m = model.strip()
    # Lower-case fix
    m = m  # keep case, we'll normalize endings

    # Remove known suffixes that shouldn't be there.
    # Loop to a fixed point (not a single pass) — a raw model can carry more
    # than one bad suffix at once (e.g. a trailing space *after* "-NA-2"), and
    # a single top-to-bottom pass only strips whichever one happens to be
    # outermost, silently leaving the rest (e.g. "Foo-NA-2 " -> "Foo-NA-2"
    # instead of "Foo").
    stripped = True
    while stripped:
        stripped = False
        for bad in [' - Harvest', '-NA-2', '-NA—R', '-R-EU', '-2', ' ']:
            if m.endswith(bad):
                m = m[:-len(bad)]
                stripped = True

    # Standardize ending to -NA-R
    if m.endswith('-NA-R'):
        return m
    if m.endswith('-NA'):
        return m + '-R'
    if m.endswith('-R') and not m.endswith('-NA-R'):
        # e.g. AP6-86-4K-R stays, but AP9-A75 → add -NA-R
        if '-4K-R' in m or '-4k-R' in m:
            return m
        return m[:-2] + '-NA-R'
    # No suffix at all
    return m + '-NA-R'


def _classify_type2(result: str, category: str) -> str:
    cat = str(category).lower().strip()
    if 'scrap' in cat:
        return 'Salvage of Hardware and Scrap'
    res = str(result).lower()
    for kw in HEAVY_KEYWORDS:
        if kw in res:
            return 'Heavy'
    return 'Basic'


def _derive_type(category: str):
    cat = str(category).lower().strip()
    for k, v in CATEGORY_TYPE.items():
        if k in cat:
            return v
    return None


# ── Main sanitize function ────────────────────────────────────────────────────

def sanitize(raw_path, date_from: datetime, date_to: datetime):
    """
    Load raw production file, derive Type/Type2/Size/CleanModel, filter by date.
    Returns: (clean_df, issues, raw_df)
    """
    df = pd.read_excel(raw_path, sheet_name='Repair Data')
    df['Date Integer'] = pd.to_datetime(df['Date Integer'], errors='coerce')

    # Date filter
    mask = (
        (df['Date Integer'] >= pd.Timestamp(date_from)) &
        (df['Date Integer'] <= pd.Timestamp(date_to))
    )
    df = df[mask].copy().reset_index(drop=True)

    # ── Derive model and size from serial ──────────────────────────────────────
    derived = df['Actual Serial'].apply(
        lambda s: pd.Series(_serial_to_model_and_size(s),
                            index=['_derived_model', '_derived_size']))
    df = pd.concat([df, derived], axis=1)

    # ── Clean model name ───────────────────────────────────────────────────────
    # Use derived model if available; otherwise clean the raw model
    df['_clean_model'] = df.apply(
        lambda r: r['_derived_model'] if pd.notna(r['_derived_model'])
                  else _clean_model(r['Actual Model']), axis=1)

    # ── Resolve size ───────────────────────────────────────────────────────────
    raw_size = df['Derive Size'].astype(str).str.strip()

    SIZE_MAP = {
        '5t':'75','6t':'86','0t':'70','7o':'70',
        'b7':'75','77':'75','76':'75','5w':'75',
        '66':'65','68':'86','57':'75','83':'86',
    }
    df['_size'] = raw_size.str.lower().map(
        lambda s: s if s in ('55','65','70','75','86')
                  else SIZE_MAP.get(s)
    )
    # Fill from derived serial lookup
    serial_has_size = df['_derived_size'].notna()
    df.loc[serial_has_size, '_size'] = df.loc[serial_has_size, '_derived_size']

    # Fall back: parse size digits from clean model name
    still_missing = df['_size'].isna()
    def size_from_model(m):
        if not isinstance(m, str): return None
        hit = re.search(r'[-](\d{2})[-]', m)
        if hit and hit.group(1) in ('55','65','70','75','86'):
            return hit.group(1)
        return None
    df.loc[still_missing, '_size'] = df.loc[still_missing, '_clean_model'].apply(size_from_model)

    # ── Derive Type and Type2 ──────────────────────────────────────────────────
    df['_Type']  = df['Category'].apply(_derive_type)
    df['_Type2'] = df.apply(lambda r: _classify_type2(r['Result'], r['Category']), axis=1)

    # ── Size → billing size label ──────────────────────────────────────────────
    # Anything < 86" is Small; 86"+ is Large
    df['Size'] = df['_size'].apply(
        lambda s: 'Large' if s in LARGE_SIZES else ('Small' if s in BILLABLE_SIZES else None))

    # ── Collect issues ─────────────────────────────────────────────────────────
    issues = []
    issue_idx = set()

    # Unresolvable size
    for idx in df[df['_size'].isna()].index:
        row = df.loc[idx]
        issue_idx.add(idx)
        issues.append({
            'row_index':       int(idx),
            'issue_type':      'unresolved_size',
            'description':     f"Cannot determine panel size from serial '{row['Actual Serial']}' "
                               f"or model '{row['Actual Model']}'",
            'field':           'Derive Size',
            'current_value':   str(row['Derive Size']),
            'Actual Model':    str(row['Actual Model']),
            'Actual Serial':   str(row['Actual Serial']),
            'Result':          str(row['Result']),
            'Category':        str(row['Category']),
            'Date':            str(row['Date Integer'])[:10],
            'suggested_values':['65','70','75','86','EXCLUDE'],
        })

    # Suspect serial (pure number / very short)
    for idx in df[df['Actual Serial'].astype(str).str.match(r'^\d{1,5}$', na=False)].index:
        if idx not in issue_idx:
            row = df.loc[idx]
            issue_idx.add(idx)
            issues.append({
                'row_index':       int(idx),
                'issue_type':      'suspect_serial',
                'description':     f"Serial '{row['Actual Serial']}' looks invalid",
                'field':           'Actual Serial',
                'current_value':   str(row['Actual Serial']),
                'Actual Model':    str(row['Actual Model']),
                'Actual Serial':   str(row['Actual Serial']),
                'Result':          str(row['Result']),
                'Category':        str(row['Category']),
                'Date':            str(row['Date Integer'])[:10],
                'suggested_values':[],
            })

    # Unknown category
    for idx in df[df['_Type'].isna()].index:
        if idx not in issue_idx:
            row = df.loc[idx]
            issue_idx.add(idx)
            issues.append({
                'row_index':       int(idx),
                'issue_type':      'unknown_category',
                'description':     f"Unknown category '{row['Category']}'",
                'field':           'Category',
                'current_value':   str(row['Category']),
                'Actual Model':    str(row['Actual Model']),
                'Actual Serial':   str(row['Actual Serial']),
                'Result':          str(row['Result']),
                'Category':        str(row['Category']),
                'Date':            str(row['Date Integer'])[:10],
                'suggested_values':['Refurbished','Pending Parts','Scrap','EXCLUDE'],
            })

    # ── Build auto-corrections summary (for UI transparency) ────────────────
    auto_corrections = []
    for idx, row in df.iterrows():
        model_changed = (
            pd.notna(row.get('_derived_model')) and
            str(row.get('_derived_model','')) != str(row.get('Actual Model',''))
        )
        size_raw = str(row.get('Derive Size','')).strip()
        size_changed = (
            row.get('_size') is not None and
            size_raw.lower() not in ('55','65','70','75','86','nan','none','') and
            size_raw != row.get('_size','')
        )
        if model_changed or size_changed:
            auto_corrections.append({
                'row_index':     int(idx),
                'Actual Model':  str(row.get('Actual Model','')),
                'Actual Serial': str(row.get('Actual Serial','')),
                'Date':          str(row.get('Date Integer',''))[:10],
                'Result':        str(row.get('Result','')),
                'Category':      str(row.get('Category','')),
                'old_model':     str(row.get('Actual Model','')),
                'new_model':     str(row.get('_derived_model','')) if model_changed else str(row.get('Actual Model','')),
                'old_size':      size_raw,
                'new_size':      str(row.get('_size','')),
                'model_changed': model_changed,
                'size_changed':  size_changed,
            })

    # ── Build clean dataframe ──────────────────────────────────────────────────
    non_billable_idx = set(
        df[(df['_size'].notna()) & (~df['_size'].isin(BILLABLE_SIZES))].index
    )
    exclude_all = issue_idx | non_billable_idx

    clean_df = df[~df.index.isin(exclude_all)].copy()
    clean_df['Type']         = clean_df['_Type']
    clean_df['Type2']        = clean_df['_Type2']
    clean_df['Actual Model'] = clean_df['_clean_model']

    return clean_df, issues, df, auto_corrections


def apply_corrections(df: pd.DataFrame, corrections: dict) -> pd.DataFrame:
    df = df.copy()
    for idx_str, corr in corrections.items():
        idx = int(idx_str)
        if corr == 'EXCLUDE' or (isinstance(corr, dict) and corr.get('value') == 'EXCLUDE'):
            df.loc[idx, '_exclude'] = True
            continue
        field = corr.get('field') if isinstance(corr, dict) else None
        value = corr.get('value') if isinstance(corr, dict) else corr
        if not field or not value:
            continue
        if field == 'Derive Size':
            df.loc[idx, '_size']    = value
            df.loc[idx, 'Size']     = 'Large' if value in LARGE_SIZES else 'Small'
            df.loc[idx, 'Derive Size'] = value
        elif field == 'Category':
            df.loc[idx, 'Category'] = value
            df.loc[idx, '_Type']    = _derive_type(value)
            df.loc[idx, 'Type']     = _derive_type(value)
            df.loc[idx, '_Type2']   = _classify_type2(df.loc[idx, 'Result'], value)
            df.loc[idx, 'Type2']    = df.loc[idx, '_Type2']
        elif field == 'Actual Serial':
            df.loc[idx, 'Actual Serial'] = value
            model, size = _serial_to_model_and_size(value)
            if model:
                df.loc[idx, '_clean_model'] = model
                df.loc[idx, 'Actual Model'] = model
            if size:
                df.loc[idx, '_size'] = size
                df.loc[idx, 'Size']  = 'Large' if size in LARGE_SIZES else 'Small'
        df.loc[idx, '_corrected'] = True
    return df


def export_corrected_workbook(raw_path, df_with_corrections: pd.DataFrame, output_path):
    raw_path    = Path(raw_path)
    output_path = Path(output_path)
    shutil.copy(raw_path, output_path)
    wb = load_workbook(output_path)

    if 'Sanitized Data' in wb.sheetnames:
        del wb['Sanitized Data']
    ws = wb.create_sheet('Sanitized Data')

    out_cols = ['Date Integer','Actual Model','Actual Serial','Type','Type2',
                'Result','Category','Size','Derive Size','_notes']

    HDR_FILL = PatternFill('solid', fgColor='1F4E79')
    FIX_FILL = PatternFill('solid', fgColor='FFF2CC')
    EXC_FILL = PatternFill('solid', fgColor='FFE0E0')

    col_widths = [14,26,24,20,14,38,16,8,12,40]
    for ci, (h, w) in enumerate(zip(
        ['Date','Clean Model','Serial','Type','Type2','Result','Category','Size','Raw Size','Notes'],
        col_widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64+ci)].width = w

    for ri, (_, row) in enumerate(df_with_corrections.iterrows(), 2):
        excluded  = row.get('_exclude', False)
        corrected = row.get('_corrected', False)
        fill = EXC_FILL if excluded else (FIX_FILL if corrected else None)
        vals = [
            str(row.get('Date Integer',''))[:10],
            row.get('Actual Model',''),
            row.get('Actual Serial',''),
            row.get('_Type') or row.get('Type',''),
            row.get('_Type2') or row.get('Type2',''),
            row.get('Result',''),
            row.get('Category',''),
            row.get('Size',''),
            row.get('Derive Size',''),
            row.get('_notes',''),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=str(v) if pd.notna(v) else '')
            c.font = Font(name='Calibri', size=10)
            if fill:
                c.fill = fill

    wb.save(output_path)
