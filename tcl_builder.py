"""
tcl_builder.py — Builds the TCL (TTE Technology) Warehouse Invoice.

Source file: a single "Inventory Export" workbook (raw, unprocessed) with columns:
    Transfer Detail ID, Model, Serial Number, Grade, Rack, Bin, Warehouse, Received Date

There is no "Item Type" column like the Promethean storage file — units (TVs) vs.
non-serialized parts are distinguished by warehouse location:
    Rack == 'MAIN' / Bin == 'RCV'   -> whole unit (TV), billed as Pallet Storage
    everything else                  -> non-serialized part, billed as a box "In Fee"

Neither pallet groupings nor box groupings exist anywhere in the source file — the
warehouse doesn't track them digitally. So this builder works in two steps, same
shape as the Storage module:

    STEP 1 - analyze_tcl()   Loads the file, splits units vs parts, groups each into
                              batches (units by Received Date, parts by Model+Received
                              Date), and returns those batches as "input needed" groups.
                              Nothing is billed yet.

    STEP 2 - build_tcl_invoice()  Takes the user's pallet/box breakdowns for every
                              group (validated to sum to the group's quantity), prices
                              them per the Depot Services rate sheet, and writes the
                              Excel invoice (Invoice + Line Items sheets).

Pricing (Depot Services rate sheet):
    Pallet Storage   <=10 total pallets : $75.00 / pallet
                     >=11 total pallets : $60.00 / pallet
    Non-Serialized In Fee, per box, by parts-in-box:
        2-5   parts -> $3.85
        6-10  parts -> $7.70
        11-15 parts -> $11.55
        16-20 parts -> $15.00
    A box of exactly 1 part bills at the Serialized In Fee rate ($3.85/part) instead
    of a box tier.
"""

import shutil
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Color

TEMPLATE_FILE = Path(__file__).parent / "template" / "TCL_Warehouse_Invoice_Template.xlsx"

ACCT_FMT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
DATE_FMT = "mm-dd-yy"
TAX_RATE = 0.07

PALLET_RATE_LOW  = 75.00   # <= 10 total pallets
PALLET_RATE_HIGH = 60.00   # >= 11 total pallets
PALLET_BREAK     = 10

SERIALIZED_IN_FEE = 3.85   # single-part "box" (qty == 1)

BOX_TIERS = [
    (2, 5,   3.85,  "2-5 Parts in Box"),
    (6, 10,  7.70,  "6-10 Parts in Box"),
    (11, 15, 11.55, "11-15 Parts in Box"),
    (16, 20, 15.00, "16-20 Parts in Box"),
]

DEFAULT_BILL_TO = "TTE Technology Inc\n189 Technology Dr.\nIrvine, CA 92618"


def box_tier(qty: int):
    """Return (rate, label) for a box of this size, or (None, None) if invalid/out of range."""
    if qty == 1:
        return SERIALIZED_IN_FEE, "Serialized In Fee (1 part)"
    for lo, hi, rate, label in BOX_TIERS:
        if lo <= qty <= hi:
            return rate, label
    return None, None


def pallet_rate(total_pallets: int) -> float:
    return PALLET_RATE_LOW if total_pallets <= PALLET_BREAK else PALLET_RATE_HIGH


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — analyze_tcl(): load the raw inventory file, split units/parts, group
# ══════════════════════════════════════════════════════════════════════════════

def analyze_tcl(inventory_path, date_from, date_to, log=print):
    """
    Returns a dict:
      - unit_groups: [{key, received_date, quantity, rows:[...]}]   (needs pallet breakdown)
      - part_groups: [{key, model, grade, received_date, quantity, rows:[...]}] (needs box breakdown, period-only)
      - period_start / period_end
      - raw counts for logging
    Nothing is priced yet — the caller must collect a breakdown for every group
    (a list of ints summing to that group's quantity) before calling build_tcl_invoice().
    """
    inv = pd.read_excel(inventory_path, sheet_name=0)
    inv.columns = [str(c).strip() for c in inv.columns]
    required = ["Model", "Serial Number", "Grade", "Rack", "Bin", "Received Date"]
    missing = [c for c in required if c not in inv.columns]
    if missing:
        raise ValueError(f"Inventory file is missing expected column(s): {missing}")

    inv["Received Date"] = pd.to_datetime(inv["Received Date"], errors="coerce")
    log(f"Inventory Export: {len(inv)} rows")

    is_unit = (inv["Rack"].astype(str).str.strip().str.upper() == "MAIN") | \
              (inv["Bin"].astype(str).str.strip().str.upper() == "RCV")
    units = inv[is_unit].copy()
    parts = inv[~is_unit].copy()
    log(f"Classified by Rack/Bin — Units (whole product, pallet storage): {len(units)}, "
        f"Parts (non-serialized, box in-fee): {len(parts)}")

    # ── Units: ALL units currently in inventory are billed for pallet storage
    #    every period (it's ongoing storage rent), regardless of when received.
    #    Grouped by Received Date batch — that's the natural physical grouping,
    #    since pallet assignment isn't tracked in the source file.
    unit_groups = []
    if len(units):
        units_sorted = units.sort_values("Received Date")
        for recv_date, grp in units_sorted.groupby(units_sorted["Received Date"].dt.date):
            key = f"unit::{recv_date}"
            unit_groups.append({
                "key": key,
                "received_date": str(recv_date),
                "quantity": len(grp),
                "rows": grp[["Model", "Serial Number", "Received Date"]].to_dict("records"),
            })
    log(f"Unit batches needing pallet assignment: {len(unit_groups)} "
        f"({sum(g['quantity'] for g in unit_groups)} units total)")

    # ── Parts: only parts RECEIVED IN THE BILLING PERIOD get an In Fee this
    #    invoice (it's a one-time fee charged when the box checks in — a part
    #    received last month was already billed on last month's invoice).
    #    Grouped by Model + Received Date, per the user's own packing convention
    #    (same part# received the same day = candidate for the same box).
    period_start = pd.Timestamp(date_from)
    period_end   = pd.Timestamp(date_to)
    parts_period = parts[
        (parts["Received Date"] >= period_start) & (parts["Received Date"] <= period_end)
    ].copy()
    log(f"Parts received in billing period {period_start.date()} to {period_end.date()}: {len(parts_period)}")

    part_groups = []
    if len(parts_period):
        for (model, grade, recv_date), grp in parts_period.groupby(
            ["Model", "Grade", parts_period["Received Date"].dt.date]
        ):
            # Include Grade in the key — Model+Received Date alone can collide
            # when the same model has multiple grades received the same day
            # (e.g. A-grade and B-grade units of the same TV part checked in
            # on the same date). Without Grade in the key, two distinct part
            # groups would render with the same key in the frontend and the
            # user's single breakdown entry would get silently applied to
            # both groups server-side (validate_breakdowns keys off "key").
            key = f"part::{model}::{grade}::{recv_date}"
            part_groups.append({
                "key": key,
                "model": model,
                "grade": grade,
                "received_date": str(recv_date),
                "quantity": len(grp),
                "rows": grp[["Model", "Serial Number", "Received Date"]].to_dict("records"),
            })
    part_groups.sort(key=lambda g: (g["received_date"], g["model"]))
    log(f"Part groups needing box breakdown: {len(part_groups)}")

    return {
        "unit_groups":  unit_groups,
        "part_groups":  part_groups,
        "period_start": period_start,
        "period_end":   period_end,
        "unit_count":   len(units),
        "part_count":   len(parts_period),
    }


def _parse_breakdown(text: str):
    """'20,20,10' -> [20,20,10]. Raises ValueError on bad input."""
    parts = [p.strip() for p in str(text).split(",") if p.strip() != ""]
    if not parts:
        raise ValueError("empty breakdown")
    out = []
    for p in parts:
        n = int(float(p))
        if n <= 0:
            raise ValueError(f"box/pallet count must be positive, got {n}")
        out.append(n)
    return out


def validate_breakdowns(analysis: dict, unit_breakdowns: dict, box_breakdowns: dict):
    """
    unit_breakdowns: {group_key: "7,7,5,7,4"}  (comma list of units-per-pallet)
    box_breakdowns:  {group_key: "20,20,10"}   (comma list of parts-per-box)
    Every group in analysis must have an entry whose values sum to that group's quantity,
    and every box size must fall in a valid tier (1, or 2-20).
    Returns (errors: list[str], parsed_unit: dict[key->list[int]], parsed_box: dict[key->list[int]])
    """
    errors = []
    parsed_unit, parsed_box = {}, {}

    for g in analysis["unit_groups"]:
        raw = unit_breakdowns.get(g["key"], "")
        try:
            vals = _parse_breakdown(raw)
        except ValueError as e:
            errors.append(f"Unit batch {g['received_date']} ({g['quantity']} units): {e}")
            continue
        if sum(vals) != g["quantity"]:
            errors.append(
                f"Unit batch {g['received_date']}: pallet breakdown sums to {sum(vals)}, "
                f"expected {g['quantity']}")
            continue
        parsed_unit[g["key"]] = vals

    for g in analysis["part_groups"]:
        raw = box_breakdowns.get(g["key"], "")
        try:
            vals = _parse_breakdown(raw)
        except ValueError as e:
            errors.append(f"Part group {g['model']} / {g['received_date']} ({g['quantity']} pcs): {e}")
            continue
        if sum(vals) != g["quantity"]:
            errors.append(
                f"Part group {g['model']} / {g['received_date']}: box breakdown sums to {sum(vals)}, "
                f"expected {g['quantity']}")
            continue
        bad = [v for v in vals if v != 1 and not any(lo <= v <= hi for lo, hi, _, _ in BOX_TIERS)]
        if bad:
            errors.append(
                f"Part group {g['model']} / {g['received_date']}: box size(s) {bad} don't fit any "
                f"pricing tier (must be 1, or 2-20)")
            continue
        parsed_box[g["key"]] = vals

    return errors, parsed_unit, parsed_box


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — build_tcl_invoice(): price everything and write the Excel file
# ══════════════════════════════════════════════════════════════════════════════

def build_tcl_invoice(
    analysis: dict,
    unit_breakdowns: dict,   # {group_key: [ints]} already validated
    box_breakdowns: dict,    # {group_key: [ints]} already validated
    invoice_number="", invoice_date=None, due_date=None,
    po_number="Contract", terms="Net 30",
    bill_to=DEFAULT_BILL_TO, ship_to=None,
    period_label=None,
    output_path: Path = None,
    log=print,
):
    ship_to = ship_to or bill_to
    invoice_date = invoice_date or datetime.now()
    period_label = period_label or analysis["period_end"].strftime("%b %Y") + "***"

    # ── Build pallet line items (Line Items sheet rows + total pallet count) ──
    pallet_rows = []   # each: dict for one physical TV line
    total_pallets = 0
    for g in analysis["unit_groups"]:
        vals = unit_breakdowns.get(g["key"], [])
        rows = g["rows"]
        idx = 0
        for count in vals:
            total_pallets += 1
            batch_rows = rows[idx: idx + count]
            idx += count
            for j, r in enumerate(batch_rows):
                pallet_rows.append({
                    "model":     r["Model"],
                    "serial":    r["Serial Number"],
                    "qty_box":   1,
                    "recv_date": r["Received Date"],
                    "pallet":    total_pallets,
                    "charge":    0,   # filled in below on the last row of the pallet
                    "is_last":   (j == len(batch_rows) - 1),
                })
    p_rate = pallet_rate(total_pallets) if total_pallets else 0
    for r in pallet_rows:
        if r["is_last"]:
            r["charge"] = p_rate
    log(f"Pallets: {total_pallets} total -> ${p_rate:.2f}/pallet ({'<=10' if total_pallets<=PALLET_BREAK else '>=11'} tier)")

    # ── Build box line items (Line Items sheet rows + tier totals) ────────────
    box_rows = []      # each: one row per box
    tier_box_counts = {}   # label -> count of boxes
    tier_box_rate   = {}   # label -> rate
    for g in analysis["part_groups"]:
        vals = box_breakdowns.get(g["key"], [])
        rows = g["rows"]
        idx = 0
        for count in vals:
            rate, label = box_tier(count)
            batch_rows = rows[idx: idx + count]
            idx += count
            box_rows.append({
                "model":     g["model"],
                "serial":    "N/A",
                "qty_box":   count,
                "recv_date": batch_rows[0]["Received Date"] if batch_rows else None,
                "pallet":    "N/A",
                "charge":    rate,
                "tier_label": label,
            })
            tier_box_counts[label] = tier_box_counts.get(label, 0) + 1
            tier_box_rate[label]   = rate
    log(f"Part boxes: {len(box_rows)} across {len(tier_box_counts)} tier(s)")

    # ── Invoice-sheet summary lines (skip any zero-qty line, per spec) ────────
    summary_lines = []
    if total_pallets > 0:
        summary_lines.append({
            "label": "FGI TV Inventory (in pallet)",
            "qty":   total_pallets,
            "rate":  p_rate,
        })
    # Order tiers the same way the rate sheet lists them (smallest box first)
    tier_order = ["Serialized In Fee (1 part)"] + [t[3] for t in BOX_TIERS]
    for label in tier_order:
        qty = tier_box_counts.get(label, 0)
        if qty <= 0:
            continue
        display = "TV Part A-Grade Serialized Inventory (Single Part)" if label.startswith("Serialized") \
            else f"TV Part A-Grade Inventory ({label})"
        summary_lines.append({
            "label": display,
            "qty":   qty,
            "rate":  tier_box_rate[label],
        })

    subtotal = round(sum(l["qty"] * l["rate"] for l in summary_lines), 2)
    tax   = round(subtotal * TAX_RATE, 2)
    total = round(subtotal + tax, 2)
    log(f"Subtotal ${subtotal:,.2f}  Tax ${tax:,.2f}  Total ${total:,.2f}")

    # ── Write workbook ──────────────────────────────────────────────────────
    shutil.copy(TEMPLATE_FILE, output_path)
    wb = load_workbook(output_path)
    wb._external_links = []

    _build_invoice_sheet(wb, invoice_number, invoice_date, due_date, po_number, terms,
                          bill_to, ship_to, period_label, summary_lines, subtotal, tax, total)
    _build_line_items_sheet(wb, pallet_rows, box_rows)

    wb.save(output_path)
    log(f"Invoice saved -> {Path(output_path).name}")

    return {
        "total_pallets":  total_pallets,
        "pallet_rate":    p_rate,
        "box_count":      len(box_rows),
        "line_count":     len(summary_lines),
        "subtotal":       subtotal,
        "tax":            tax,
        "total":          total,
    }


# ── Sheet writers ───────────────────────────────────────────────────────────

def _f(bold=False, size=11):
    return Font(name="Carlito", bold=bold, size=size)


def _build_invoice_sheet(wb, invoice_number, invoice_date, due_date, po_number, terms,
                          bill_to, ship_to, period_label, summary_lines, subtotal, tax, total):
    ws = wb["Invoice"]

    # The template's merges below row 14 are sized for exactly 3 line items.
    # Unmerge ALL existing ranges and rebuild: fixed header merges always,
    # then dynamic ones sized to however many line items this invoice has.
    # (merged_cells.ranges.clear() alone leaves cells as read-only MergedCell
    # objects — must go through unmerge_cells() to restore normal Cells.)
    for rng in [str(r) for r in list(ws.merged_cells.ranges)]:
        ws.unmerge_cells(rng)

    # Wipe stale sample content from row 15 down (period label, line items,
    # wire info, totals, special instructions) — everything below this point
    # is rebuilt fresh below.
    clear_to = max(ws.max_row, 40)
    for r in range(15, clear_to + 1):
        for col in range(1, 10):
            c = ws.cell(row=r, column=col)
            c.value = None
            c.number_format = "General"

    def merge(rng):
        ws.merge_cells(rng)

    # ── Fixed header merges (rows 1-14 never move) ────────────────────────
    merge("G1:I1")
    merge("A5:D5"); merge("F5:I5")
    merge("A6:D9"); merge("F6:I9")
    merge("A14:F14")

    ws["H2"].value = "Invoice #"
    ws["H3"] = invoice_number
    ws["H3"].font = _f()
    ws["G3"].value = invoice_date
    ws["G3"].number_format = DATE_FMT
    ws["A12"].value = po_number
    ws["C12"].value = due_date
    ws["C12"].number_format = DATE_FMT
    ws["B12"].value = terms
    ws["A6"].value = bill_to
    ws["F6"].value = ship_to

    # Line items start at row 15 (row 15 = period label, rows 16.. = charges)
    first_row = 15
    ws.cell(row=first_row, column=1, value=period_label).font = _f()
    merge(f"A{first_row}:F{first_row}")

    row = first_row + 1
    for line in summary_lines:
        ws.cell(row=row, column=1, value=line["label"]).font = _f()
        merge(f"A{row}:F{row}")
        c_qty = ws.cell(row=row, column=7, value=line["qty"]); c_qty.font = _f()
        c_rate = ws.cell(row=row, column=8, value=line["rate"])
        c_rate.font = _f(); c_rate.number_format = ACCT_FMT
        c_amt = ws.cell(row=row, column=9, value=f"=H{row}*G{row}")
        c_amt.font = _f(); c_amt.number_format = ACCT_FMT
        row += 1
    last_item_row = row - 1
    if last_item_row < first_row + 1:
        # No billable lines at all (shouldn't normally happen) — keep a valid range
        last_item_row = first_row + 1
        ws.cell(row=last_item_row, column=9, value=0).number_format = ACCT_FMT

    # Shift the "Wire Information" note + totals block down to sit right after
    # the line items, mirroring the sample's fixed layout when there are more
    # or fewer lines than the 3-line example.
    wire_row = last_item_row + 2
    ws.cell(row=wire_row, column=1, value="USSI Global Wire Information").font = Font(name="Carlito", size=11)
    merge(f"A{wire_row}:F{wire_row+3}")

    totals_row = wire_row + 6
    ws.cell(row=totals_row, column=7, value="Subtotal").font = _f(bold=True)
    merge(f"G{totals_row}:H{totals_row}")
    c = ws.cell(row=totals_row, column=9, value=f"=SUM(I{first_row+1}:I{last_item_row})")
    c.font = _f(bold=True); c.number_format = r'\$#,##0.00'

    ws.cell(row=totals_row+1, column=7, value=f"Sales Tax ({TAX_RATE*100:.2f}%)").font = _f(bold=True)
    merge(f"G{totals_row+1}:H{totals_row+1}")
    c = ws.cell(row=totals_row+1, column=9, value=f"=I{totals_row}*{TAX_RATE}")
    c.font = _f(bold=True); c.number_format = r'\$#,##0.00'

    ws.cell(row=totals_row+2, column=7, value="Total").font = _f(bold=True)
    merge(f"G{totals_row+2}:H{totals_row+2}")
    c = ws.cell(row=totals_row+2, column=9, value=f"=I{totals_row}+I{totals_row+1}")
    c.font = _f(bold=True); c.number_format = r'\$#,##0.00'

    instr_row = totals_row + 4
    ws.cell(row=instr_row, column=1, value="Special instructions:").font = Font(name="Carlito", size=11)
    merge(f"A{instr_row}:I{instr_row+4}")


def _build_line_items_sheet(wb, pallet_rows, box_rows):
    ws = wb["Line Items"]
    ws.delete_rows(1, ws.max_row)
    headers = ["Line Item", "Model", "Serial", "Quantity in Box", "Receive Date", "Pallet Number", "Charge"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(name="Carlito", bold=True, size=11)

    row = 2
    for r in pallet_rows:
        vals = ["FGI TV Inventory (in pallet)", r["model"], r["serial"], r["qty_box"],
                r["recv_date"], r["pallet"], r["charge"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = Font(name="Carlito", size=11)
            if ci == 5 and isinstance(v, (pd.Timestamp, datetime)):
                c.number_format = DATE_FMT
            if ci == 7:
                c.number_format = ACCT_FMT
        row += 1

    for r in box_rows:
        label = "TV Part A-Grade Serialized Inventory (Single Part)" if r["tier_label"].startswith("Serialized") \
            else f"TV Part A-Grade Inventory ({r['tier_label']})"
        vals = [label, r["model"], r["serial"], r["qty_box"], r["recv_date"], r["pallet"], r["charge"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = Font(name="Carlito", size=11)
            if ci == 5 and isinstance(v, (pd.Timestamp, datetime)):
                c.number_format = DATE_FMT
            if ci == 7:
                c.number_format = ACCT_FMT
        row += 1

    for col, w in zip("ABCDEFG", [40.66, 19.11, 31.11, 18.33, 16.44, 17.11, 15.55]):
        ws.column_dimensions[col].width = w
