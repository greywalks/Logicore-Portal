"""
report_builder.py — Builds the TPV-PPDS Month End Report from the raw warehouse
system export (Raw_Philips_Data.xlsx: Inventory, Received, Repairs, Shipments,
RMA List tabs), producing the same Inventory / Shipping / Recieved / Repairs
tabs that philips_builder.analyze_philips() expects.

Validated against a real Month End Report (July 2026 period) by reverse-
engineering the transformation — see the design notes below for exactly what's
mechanical vs. what needed a judgment call.

── What's fully mechanical (validated to 97-100% row-for-row match) ──────────
  - Inventory:  the raw Inventory tab IS the current snapshot. Size = looked up
    from the Dimensions reference by Model (same lookup as the invoice builder).
    Rcv Date = most recent Received-log date for that Serial (best-effort;
    purely informational, not used in any billing calculation).
  - Shipping:   raw Shipments filtered to the billing period by Date. Exact
    match in testing.
  - Repairs:    raw Repairs filtered to the period, Status in
    {"Repaired","Harvested"} only (Pending items aren't billable — no price
    exists for them). RMA / Received Date looked up from the Received log by
    Serial (most recent record at-or-before the repair date).
  - Recieved:   raw Received filtered to the period. Grade/Warehouse looked up
    from the Inventory tab by Serial.

── What ISN'T mechanical (no clean rule found — flagged, not auto-excluded) ──
  Testing showed the source data does NOT cleanly separate billable from
  non-billable rows: e.g. of 27 same-period "re-received" serials, only 2 were
  actually excluded from the real report; of 13 "NonConforming"-labeled rows,
  9 were still billed. This looks like case-by-case human judgment rather than
  a formula. Rather than guess and silently misstate a bill, this module
  INCLUDES every in-period row by default and writes a separate "Flagged for
  Review" tab highlighting the ones worth a second look (NonConforming origin/
  RMA, missing tracking number) so they can be pulled out manually in Excel
  before the file is used to build an invoice.
"""

from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from philips_builder import load_dimensions, lookup_sqft, _tf as _pb_tf

DATE_FMT = "mm-dd-yy"


def _tf(bold=False, size=10):
    return Font(name="Calibri", bold=bold, size=size)


def _write_header_row(ws, headers, row=1, col_widths=None):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = _tf(bold=True, size=11)
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w


def _is_nonconforming(row):
    origin = str(row.get("Origin Company", "") or "")
    rma    = str(row.get("Arrival RMA", "") or "")
    return "nonconform" in origin.lower() or "nonconform" in rma.lower()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — analyze_raw_data(): filter/join the raw export into report shape
# ══════════════════════════════════════════════════════════════════════════════

def analyze_raw_data(raw_path, period_start, period_end, dims=None, log=print):
    period_start = pd.Timestamp(period_start)
    period_end   = pd.Timestamp(period_end)
    if dims is None:
        dims = load_dimensions()

    raw_inv  = pd.read_excel(raw_path, sheet_name="Inventory")
    raw_recv = pd.read_excel(raw_path, sheet_name="Received")
    raw_rep  = pd.read_excel(raw_path, sheet_name="Repairs")
    raw_ship = pd.read_excel(raw_path, sheet_name="Shipments")
    log(f"Loaded raw export — Inventory: {len(raw_inv)}, Received: {len(raw_recv)}, "
        f"Repairs: {len(raw_rep)}, Shipments: {len(raw_ship)}")

    raw_recv = raw_recv.copy()
    raw_recv["Date"] = pd.to_datetime(raw_recv["Date"], errors="coerce")
    raw_rep  = raw_rep.copy()
    raw_rep["Date"]  = pd.to_datetime(raw_rep["Date"], errors="coerce")
    raw_ship = raw_ship.copy()
    raw_ship["Date"] = pd.to_datetime(raw_ship["Date"], errors="coerce")

    # ── Inventory: current snapshot, dedup by Serial ────────────────────────
    inv = raw_inv.dropna(subset=["Serial"]).drop_duplicates(subset="Serial", keep="first").copy()
    inv["Size"] = inv["Model"].apply(lambda m: lookup_sqft(m, dims))
    missing_inv_dims = sorted(set(str(m) for m in inv[inv["Size"].isna()]["Model"].dropna()))
    recv_last = raw_recv.dropna(subset=["Serial"]).sort_values("Date").groupby("Serial")["Date"].last()
    inv["Rcv Date"] = inv["Serial"].map(recv_last)
    inventory_out = inv[["Type", "Grade", "RMA", "Model", "Size", "Serial", "Rcv Date"]].reset_index(drop=True)
    log(f"Inventory: {len(inventory_out)} unique units on hand "
        f"({len(missing_inv_dims)} model(s) missing a dimension)")

    # ── Shipping: period filter ──────────────────────────────────────────────
    ship_period = raw_ship[
        raw_ship["Date"].notna() & (raw_ship["Date"] >= period_start) & (raw_ship["Date"] <= period_end)
    ].copy()
    ship_period["Column11"] = "Track"
    shipping_out = ship_period[[
        "Date", "Stock Level (Primary)", "Arrival RMA", "Departure RMA",
        "Departure Tracking", "Ship to Name", "Carrier", "Column11", "Model", "Serial",
    ]].reset_index(drop=True)
    log(f"Shipping: {len(shipping_out)} shipment(s) in period")

    # ── Recieved: period filter + Grade/Warehouse join from Inventory ───────
    recv_period = raw_recv[
        raw_recv["Date"].notna() & (raw_recv["Date"] >= period_start) & (raw_recv["Date"] <= period_end)
    ].copy()
    inv_lkp = raw_inv.dropna(subset=["Serial"]).drop_duplicates(subset="Serial", keep="first").set_index("Serial")
    recv_period["Grade"]     = recv_period["Serial"].map(inv_lkp["Grade"])
    recv_period["Warehouse"] = recv_period["Serial"].map(inv_lkp["Type"])
    recv_period["_flag_nonconforming"] = recv_period.apply(_is_nonconforming, axis=1)
    recv_period["_flag_no_tracking"]   = recv_period["Tracking #"].isna()
    recv_period["_flagged"] = recv_period["_flag_nonconforming"] | recv_period["_flag_no_tracking"]

    received_out = recv_period.rename(columns={"Arrival RMA": "RMA", "Tracking #": "Tracking"})[
        ["Model", "Serial", "Grade", "Warehouse", "RMA", "Date", "Origin Company", "Tracking"]
    ].reset_index(drop=True)
    flagged_received = recv_period[recv_period["_flagged"]].rename(
        columns={"Arrival RMA": "RMA", "Tracking #": "Tracking"})[
        ["Model", "Serial", "RMA", "Date", "Origin Company", "Tracking",
         "_flag_nonconforming", "_flag_no_tracking"]
    ].reset_index(drop=True)
    log(f"Recieved: {len(received_out)} receipt(s) in period "
        f"({len(flagged_received)} flagged for review — included by default)")

    # ── Repairs: period filter, billable statuses only, RMA/date join ──────
    rep_period = raw_rep[
        raw_rep["Date"].notna() & (raw_rep["Date"] >= period_start) & (raw_rep["Date"] <= period_end)
    ].copy()
    billable_statuses = {"Repaired", "Harvested"}
    rep_billable = rep_period[rep_period["Repaired/ Harvested"].isin(billable_statuses)].copy()
    rep_pending  = rep_period[~rep_period["Repaired/ Harvested"].isin(billable_statuses)].copy()

    recv_sorted = raw_recv.dropna(subset=["Serial"]).sort_values("Date")
    def _rma_and_date(row):
        prior = recv_sorted[(recv_sorted["Serial"] == row["Serial"]) & (recv_sorted["Date"] <= row["Date"])]
        if len(prior):
            m = prior.iloc[-1]
            return pd.Series({"RMA": m["Arrival RMA"], "Received Date": m["Date"]})
        any_match = recv_sorted[recv_sorted["Serial"] == row["Serial"]]
        if len(any_match):
            m = any_match.iloc[0]
            return pd.Series({"RMA": m["Arrival RMA"], "Received Date": m["Date"]})
        return pd.Series({"RMA": None, "Received Date": pd.NaT})

    if len(rep_billable):
        joined = rep_billable.apply(_rma_and_date, axis=1)
        rep_billable = pd.concat([rep_billable.reset_index(drop=True), joined.reset_index(drop=True)], axis=1)
    else:
        rep_billable["RMA"] = []
        rep_billable["Received Date"] = []

    rep_billable["Repaired Y/N"] = rep_billable["Repaired/ Harvested"].map(
        lambda s: "Yes" if s == "Repaired" else "No")
    repairs_out = rep_billable.rename(columns={"Date": "Repair Date", "Repaired/ Harvested": "Status"})[
        ["Repair Date", "Received Date", "Model", "Serial", "RMA", "Status",
         "Diagnostics", "Parts Used", "Repaired Y/N"]
    ].reset_index(drop=True)
    pending_out = rep_pending.rename(columns={"Date": "Repair Date", "Repaired/ Harvested": "Status"})[
        ["Repair Date", "Model", "Serial", "Status", "Diagnostics", "Parts Used"]
    ].reset_index(drop=True)
    log(f"Repairs: {len(repairs_out)} billable (Repaired/Harvested) — "
        f"{len(pending_out)} pending/incomplete this period, not billed")

    return {
        "inventory_df": inventory_out,
        "shipping_df":  shipping_out,
        "received_df":  received_out,
        "repairs_df":   repairs_out,
        "flagged_received_df": flagged_received,
        "pending_repairs_df":  pending_out,
        "missing_inventory_dims": missing_inv_dims,
        "period_start": period_start, "period_end": period_end,
    }


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — build_month_end_report(): write the Excel file
# ══════════════════════════════════════════════════════════════════════════════

def build_month_end_report(analysis, output_path, log=print):
    wb = Workbook()
    wb.remove(wb.active)

    _write_inventory(wb, analysis["inventory_df"])
    _write_shipping(wb, analysis["shipping_df"])
    _write_received(wb, analysis["received_df"])
    _write_repairs(wb, analysis["repairs_df"])
    _write_flagged(wb, analysis["flagged_received_df"])
    _write_pending(wb, analysis["pending_repairs_df"])

    wb.save(output_path)
    log(f"Month End Report saved → {Path(output_path).name}")
    return output_path


def _write_inventory(wb, df):
    ws = wb.create_sheet("Inventory")
    headers = ["Type", "Grade", "RMA", "Model", "Size", "Serial", "Rcv Date"]
    _write_header_row(ws, headers, col_widths=[10, 8, 18, 20, 10, 20, 12])
    for i, (_, r) in enumerate(df.iterrows(), 2):
        vals = [r.get(h) for h in headers]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=v)
            c.font = _tf(size=10)
            if headers[ci-1] == "Rcv Date" and isinstance(v, (pd.Timestamp, datetime)):
                c.number_format = DATE_FMT


def _write_shipping(wb, df):
    ws = wb.create_sheet("Shipping")
    headers = ["Date", "Stock Level (Primary)", "Arrival RMA", "Departure RMA",
               "Departure Tracking", "Ship to Name", "Carrier", "Column11", "Model", "Serial"]
    _write_header_row(ws, headers, col_widths=[12, 14, 16, 16, 16, 22, 10, 10, 20, 20])
    for i, (_, r) in enumerate(df.iterrows(), 2):
        for ci, h in enumerate(headers, 1):
            v = r.get(h)
            c = ws.cell(row=i, column=ci, value=v)
            c.font = _tf(size=10)
            if h == "Date" and isinstance(v, (pd.Timestamp, datetime)):
                c.number_format = DATE_FMT


def _write_received(wb, df):
    ws = wb.create_sheet("Recieved")
    headers = ["Model", "Serial", "Grade", "Warehouse", "RMA", "Date", "Origin Company", "Tracking"]
    _write_header_row(ws, headers, col_widths=[20, 20, 8, 12, 16, 12, 22, 16])
    for i, (_, r) in enumerate(df.iterrows(), 2):
        for ci, h in enumerate(headers, 1):
            v = r.get(h)
            c = ws.cell(row=i, column=ci, value=v)
            c.font = _tf(size=10)
            if h == "Date" and isinstance(v, (pd.Timestamp, datetime)):
                c.number_format = DATE_FMT


def _write_repairs(wb, df):
    ws = wb.create_sheet("Repairs")
    headers = ["Repair Date", "Received Date", "Model", "Serial", "RMA", "Status",
               "Diagnostics", "Parts Used", "Repaired Y/N"]
    _write_header_row(ws, headers, col_widths=[12, 12, 20, 20, 16, 12, 24, 16, 10])
    for i, (_, r) in enumerate(df.iterrows(), 2):
        for ci, h in enumerate(headers, 1):
            v = r.get(h)
            c = ws.cell(row=i, column=ci, value=v)
            c.font = _tf(size=10)
            if h in ("Repair Date", "Received Date") and isinstance(v, (pd.Timestamp, datetime)):
                c.number_format = DATE_FMT


def _write_flagged(wb, df):
    """Rows included in Recieved but worth a manual look — not auto-excluded
    because testing showed no reliable rule separates billable from not."""
    ws = wb.create_sheet("Flagged for Review")
    headers = ["Model", "Serial", "RMA", "Date", "Origin Company", "Tracking", "Reason"]
    _write_header_row(ws, headers, col_widths=[20, 20, 16, 12, 22, 16, 30])
    if df is None or len(df) == 0:
        ws.cell(row=2, column=1, value="No flagged rows this period.").font = _tf(size=10)
        return
    for i, (_, r) in enumerate(df.iterrows(), 2):
        reasons = []
        if r.get("_flag_nonconforming"): reasons.append("NonConforming origin/RMA")
        if r.get("_flag_no_tracking"):   reasons.append("No tracking number")
        vals = [r.get("Model"), r.get("Serial"), r.get("RMA"), r.get("Date"),
                r.get("Origin Company"), r.get("Tracking"), " + ".join(reasons)]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=v)
            c.font = _tf(size=10)
            if ci == 4 and isinstance(v, (pd.Timestamp, datetime)):
                c.number_format = DATE_FMT


def _write_pending(wb, df):
    """Repairs still Pending this period — excluded from billing (no price),
    kept here for visibility so nothing silently disappears."""
    ws = wb.create_sheet("Pending Repairs (Not Billed)")
    headers = ["Repair Date", "Model", "Serial", "Status", "Diagnostics", "Parts Used"]
    _write_header_row(ws, headers, col_widths=[12, 20, 20, 14, 26, 16])
    if df is None or len(df) == 0:
        ws.cell(row=2, column=1, value="No pending repairs this period.").font = _tf(size=10)
        return
    for i, (_, r) in enumerate(df.iterrows(), 2):
        for ci, h in enumerate(headers, 1):
            v = r.get(h)
            c = ws.cell(row=i, column=ci, value=v)
            c.font = _tf(size=10)
            if h == "Repair Date" and isinstance(v, (pd.Timestamp, datetime)):
                c.number_format = DATE_FMT
