"""
philips_builder.py — Builds the TPV-Philips Warehouse & Repair Invoice
Mirrors the layout of 072026TPV_Philips_Warehouse_Invoice.xlsx (Breakdown / Shipping /
Receieved / Repairs tabs), following the same pattern as builder.py / storage_builder.py.

Input each month: the "Month End Report" workbook (Inventory, Shipping, Recieved, Repairs
tabs) pulled from the warehouse system — see analyze_philips().

Reference data (bundled, not uploaded monthly): philips_reference_default.json
  - "dimensions":   {MODEL: sq_ft}  — box footprint per TV model
  - "repair_cost":  [{"size": "50" or "20-24", "rb_price":, "harvest_price":, "box_build":}]

Billing rules (per USSI, confirmed Aug 2026):
  - Warehouse Cost, First 500 sq ft: flat $1,910, qty 1, always billed.
  - Demo / Service Additional Square Footage = SUM(Dimensions sq ft for units currently
    on hand in that category, from the Inventory tab) + SUM(Dimensions sq ft for units
    that shipped that month in that category, from the Shipping tab), in 50-sq-ft units
    ($3.50/unit). The flat 500 sq ft minimum is credited against the DEMO total only
    (it's the first line item, listed under the Demo section).
  - Parts Additional Square Footage: entered manually each month (no source data for it).
  - Inbound Handling = count of Recieved rows, $6 each.
  - Outbound Handling = count of Shipping rows, $6 each.
  - Repair = count/$ of Repairs rows with Status == "Repaired", priced by display size
    (parsed from the model number) via the Repair Cost table's R/B price.
  - Harvest = count/$ of Repairs rows with Status == "Harvested", priced via the Repair
    Cost table's Harvest/Scrap price.
  - Tax: 7%.

ASSUMPTION FLAGGED FOR REVIEW: the 500 sq ft credit is applied only to the Demo bucket,
matching where the line sits in the template. If it should instead come off the combined
Demo+Service total, that's a one-line change in _sqft_breakdown() below.
"""

import json
import re as _re
import shutil
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils import get_column_letter

REFERENCE_FILE   = Path(__file__).parent / "philips_reference_default.json"   # bundled seed data
DIMENSIONS_FILE  = Path(__file__).parent / "philips_dimensions.json"          # live, editable store
REPAIR_COST_FILE = Path(__file__).parent / "philips_repair_cost.json"         # live, editable store

ACCT_FMT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
INT_FMT  = '#,##0'
NUM_FMT  = '#,##0.00'
DATE_FMT = "mm-dd-yy"

TAX_RATE     = 0.07
HANDLING_FEE = 6
SQFT_UNIT    = 50       # billed in 50-sq-ft increments
SQFT_PRICE   = 3.5      # $ per 50-sq-ft increment
BASE_SQFT    = 500
BASE_PRICE   = 1910

MODEL_SIZE_RE = _re.compile(r'^\s*(\d+)')


# ══════════════════════════════════════════════════════════════════════════════
# Reference data (Dimensions + Repair Cost) — persisted, editable, uploadable
# ══════════════════════════════════════════════════════════════════════════════
# Dimensions and Repair Cost pricing live in their own JSON files next to this
# module (DIMENSIONS_FILE / REPAIR_COST_FILE), seeded once from the bundled
# philips_reference_default.json. After that they're independent of the
# original upload — edited via the admin panel, or replaced by re-uploading a
# Dimensions workbook. They are NOT re-uploaded with every invoice.

def _seed_reference():
    return json.loads(REFERENCE_FILE.read_text())


def load_dimensions() -> dict:
    """{MODEL: sq_ft}, upper-cased keys. Seeds from the bundled default on first use."""
    if DIMENSIONS_FILE.exists():
        data = json.loads(DIMENSIONS_FILE.read_text())
    else:
        data = _seed_reference()["dimensions"]
        save_dimensions(data)
    return {k.upper(): v for k, v in data.items()}


def save_dimensions(dims: dict):
    DIMENSIONS_FILE.write_text(json.dumps(dims, indent=2, sort_keys=True))


def add_dimension(model: str, sqft: float):
    """Append/overwrite one model's footprint and persist immediately."""
    dims = load_dimensions()
    dims[_normalize_model(model)] = float(sqft)
    save_dimensions(dims)
    return dims


def replace_dimensions_from_rows(rows):
    """rows: iterable of (model, sqft) pairs — used when a Dimensions workbook is
    re-uploaded. Replaces the entire store (this becomes the new source of truth)."""
    dims = {}
    for model, sqft in rows:
        if model is None or sqft is None:
            continue
        dims[_normalize_model(model)] = float(sqft)
    save_dimensions(dims)
    return dims


def load_repair_cost_raw() -> list:
    """Raw editable form: list of {"size","rb_price","harvest_price","box_build"}."""
    if REPAIR_COST_FILE.exists():
        return json.loads(REPAIR_COST_FILE.read_text())
    data = _seed_reference()["repair_cost"]
    save_repair_cost_raw(data)
    return data


def save_repair_cost_raw(tiers_raw: list):
    REPAIR_COST_FILE.write_text(json.dumps(tiers_raw, indent=2))


def _compile_tiers(tiers_raw):
    tiers = []
    for row in tiers_raw:
        s = str(row["size"]).strip()
        if "-" in s:
            lo, hi = s.split("-")
            lo, hi = int(lo), int(hi)
        else:
            lo = hi = int(s)
        tiers.append((lo, hi, row["rb_price"], row["harvest_price"], row.get("box_build")))
    tiers.sort(key=lambda t: t[0])
    return tiers


def load_reference(reference_path=None):
    """Back-compat helper — reference_path lets a one-off file override the live
    store (used only for testing); normal operation uses load_dimensions()/
    load_repair_cost_raw()."""
    if reference_path:
        data = json.loads(Path(reference_path).read_text())
        dims = {k.upper(): v for k, v in data["dimensions"].items()}
        tiers = _compile_tiers(data["repair_cost"])
        return dims, tiers
    return load_dimensions(), _compile_tiers(load_repair_cost_raw())


def _normalize_model(model):
    if model is None:
        return ""
    return str(model).strip().upper()


def lookup_sqft(model, dims):
    """Look up box footprint for a model, trying progressively looser matches."""
    m = _normalize_model(model)
    if not m:
        return None
    if m in dims:
        return dims[m]
    # Strip a trailing "-B", "/NN", "-NN" suffix and retry
    for pattern in (r'-B$', r'/\d+$', r'-\d+$'):
        stripped = _re.sub(pattern, '', m)
        if stripped != m and stripped in dims:
            return dims[stripped]
    # Try matching by base model with any suffix stripped after "/" or last "-"
    base = _re.split(r'[/]', m)[0]
    if base in dims:
        return base and dims.get(base)
    return None


def extract_display_size(model):
    m = _normalize_model(model)
    match = MODEL_SIZE_RE.match(m)
    return int(match.group(1)) if match else None


def _split_base_credit(demo_total, service_total, total_credit=BASE_SQFT):
    """Split the flat 500 sq ft minimum evenly (250/250) between Demo and Service.
    If one side's total is smaller than its half-share, that side is credited down
    to zero and the *entire remainder* of the credit shifts to the other side
    (capped at that side's own total, so credit never pushes billable sq ft negative)."""
    half = total_credit / 2
    if demo_total < half:
        demo_credit = demo_total
        service_credit = min(total_credit - demo_total, service_total)
    elif service_total < half:
        service_credit = service_total
        demo_credit = min(total_credit - service_total, demo_total)
    else:
        demo_credit = half
        service_credit = half
    return demo_credit, service_credit


def lookup_repair_price(model, status, tiers):
    """status: 'Repaired' -> R/B price, 'Harvested' -> Harvest/Scrap price.
    The Repair Cost table has gaps between listed sizes (e.g. nothing between 85" and
    95", or between 55" and 60"). For a size that falls in a gap, this snaps to the
    NEAREST listed tier rather than leaving it unpriced."""
    size = extract_display_size(model)
    if size is None or not tiers:
        return None
    best, best_dist = None, None
    for lo, hi, rb, harvest, _box in tiers:
        dist = 0 if lo <= size <= hi else min(abs(size - lo), abs(size - hi))
        if best_dist is None or dist < best_dist:
            best, best_dist = (rb, harvest), dist
    return best[0] if status == "Repaired" else best[1]


def parse_dimensions_workbook(path):
    """Read an uploaded Dimensions workbook (Model / Sq Footage columns, any sheet name,
    header row auto-detected) and return a list of (model, sqft) pairs."""
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    model_col = next((i for i, h in enumerate(header) if "model" in h), 0)
    sqft_col  = next((i for i, h in enumerate(header) if "sq" in h or "footage" in h or "footprint" in h), 1)
    out = []
    for row in rows[1:]:
        if row is None or len(row) <= max(model_col, sqft_col):
            continue
        model, sqft = row[model_col], row[sqft_col]
        if model is None or sqft is None:
            continue
        try:
            out.append((str(model).strip(), float(sqft)))
        except (TypeError, ValueError):
            continue
    return out


def write_dimensions_workbook(path, dims: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dimensions"
    ws["A1"] = "Model"; ws["B1"] = "Sq Footage"
    ws["A1"].font = _tf(bold=True); ws["B1"].font = _tf(bold=True)
    for i, model in enumerate(sorted(dims), 2):
        ws.cell(row=i, column=1, value=model).font = _tf(size=10)
        ws.cell(row=i, column=2, value=dims[model]).font = _tf(size=10)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    wb.save(path)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — analyze_philips(): load the Month End Report, compute everything
# ══════════════════════════════════════════════════════════════════════════════

def analyze_philips(report_path, parts_sqft_manual=0, dims=None, tiers=None, log=print):
    """dims/tiers: pass explicitly to reuse a specific snapshot (e.g. after the user
    supplies missing dimensions in the review step). Defaults to the live persisted store."""
    if dims is None or tiers is None:
        dims, tiers = load_dimensions(), _compile_tiers(load_repair_cost_raw())

    inv  = pd.read_excel(report_path, sheet_name="Inventory")
    ship = pd.read_excel(report_path, sheet_name="Shipping")
    recv = pd.read_excel(report_path, sheet_name="Recieved")
    rep  = pd.read_excel(report_path, sheet_name="Repairs")
    log(f"Loaded Month End Report — Inventory: {len(inv)}, Shipping: {len(ship)}, "
        f"Recieved: {len(recv)}, Repairs: {len(rep)}")

    # ── Square footage: derive from the model reference wherever needed ───────
    excluded_rows = []   # collected across all three checks -> written to an output tab

    inv = inv.dropna(subset=["Type"]).copy()
    if "Size" in inv.columns:
        inv["_sqft"] = pd.to_numeric(inv["Size"], errors="coerce")
    else:
        inv["_sqft"] = float("nan")
        log("Inventory has no Size column — deriving box square footage from Model using the Dimensions reference")
    blank_size = inv["_sqft"].isna()
    if blank_size.any():
        inv.loc[blank_size, "_sqft"] = inv.loc[blank_size, "Model"].apply(
        lambda m: lookup_sqft(m, dims) if lookup_sqft(m, dims) is not None else float("nan"))
    missing_inv_df = inv[inv["_sqft"].isna()]
    missing_inv = [str(m) for m in missing_inv_df["Model"].tolist()]
    inv_sqft = inv.groupby("Type")["_sqft"].sum().to_dict()

    ship = ship.dropna(subset=["Stock Level (Primary)"]).copy()
    ship["_sqft"] = ship["Model"].apply(lambda m: lookup_sqft(m, dims))
    missing_ship_df = ship[ship["_sqft"].isna()]
    missing_ship = [str(m) for m in missing_ship_df["Model"].tolist()]
    ship_sqft = ship.groupby("Stock Level (Primary)")["_sqft"].sum().to_dict()

    for _, r in missing_inv_df.iterrows():
        excluded_rows.append({"Source Tab": "Inventory", "Reason": "No Size / no Dimensions match",
                               "Model": r.get("Model"), "Serial": r.get("Serial"),
                               "Type": r.get("Type"), "Detail": ""})
    for _, r in missing_ship_df.iterrows():
        excluded_rows.append({"Source Tab": "Shipping", "Reason": "No Dimensions match",
                               "Model": r.get("Model"), "Serial": r.get("Serial"),
                               "Type": r.get("Stock Level (Primary)"), "Detail": ""})

    if missing_inv:
        log(f"WARNING: {len(missing_inv)} Inventory row(s) missing a Size — excluded: "
            f"{sorted(set(missing_inv))[:10]}")
    if missing_ship:
        log(f"WARNING: {len(missing_ship)} Shipping row(s) had no Dimensions match — "
            f"excluded: {sorted(set(missing_ship))[:10]}")

    demo_total    = inv_sqft.get("Demo", 0) + ship_sqft.get("Demo", 0)
    service_total = inv_sqft.get("Service", 0) + ship_sqft.get("Service", 0)
    demo_credit, service_credit = _split_base_credit(demo_total, service_total)
    demo_additional    = demo_total - demo_credit
    service_additional = service_total - service_credit

    log(f"Sq Ft — Demo: {demo_total:.2f} total, {demo_credit:.2f} of the {BASE_SQFT} "
        f"credit applied → {demo_additional:.2f} billable | Service: {service_total:.2f} "
        f"total, {service_credit:.2f} credit applied → {service_additional:.2f} billable")

    # ── Handling ──────────────────────────────────────────────────────────────
    inbound_count  = recv["Model"].notna().sum()
    outbound_count = ship["Model"].notna().sum()

    # ── Repairs ───────────────────────────────────────────────────────────────
    rep = rep.copy()
    rep["_price"] = rep.apply(lambda r: lookup_repair_price(r["Model"], r["Status"], tiers), axis=1)
    missing_rep_df = rep[rep["_price"].isna()]
    missing_rep = [str(m) for m in missing_rep_df["Model"].tolist()]
    for _, r in missing_rep_df.iterrows():
        excluded_rows.append({"Source Tab": "Repairs", "Reason": "No display size parsed from model",
                               "Model": r.get("Model"), "Serial": r.get("Serial"),
                               "Type": r.get("Status"), "Detail": ""})
    if missing_rep:
        log(f"WARNING: {len(missing_rep)} Repairs row(s) had no display size parsed — "
            f"priced $0: {sorted(set(missing_rep))[:10]}")
    rep["_price"] = rep["_price"].fillna(0)

    repaired  = rep[rep["Status"] == "Repaired"]
    harvested = rep[rep["Status"] == "Harvested"]
    repair_count,  repair_total  = len(repaired),  repaired["_price"].sum()
    harvest_count, harvest_total = len(harvested), harvested["_price"].sum()

    log(f"Repairs — Repaired: {repair_count} (${repair_total:,.2f}), "
        f"Harvested: {harvest_count} (${harvest_total:,.2f})")

    return {
        "inventory_df": inv, "shipping_df": ship, "received_df": recv, "repairs_df": rep,
        "demo_total_sqft": demo_total, "demo_additional_sqft": demo_additional,
        "service_total_sqft": service_total, "service_additional_sqft": service_additional,
        "parts_sqft_manual": parts_sqft_manual,
        "inbound_count": int(inbound_count), "outbound_count": int(outbound_count),
        "repair_count": int(repair_count), "repair_total": float(repair_total),
        "harvest_count": int(harvest_count), "harvest_total": float(harvest_total),
        "missing_inventory_models": sorted(set(missing_inv)),
        "missing_shipping_models":  sorted(set(missing_ship)),
        "missing_repair_models":    sorted(set(missing_rep)),
        "missing_dimension_models": sorted(set(missing_inv) | set(missing_ship)),
        "excluded_df": pd.DataFrame(excluded_rows, columns=["Source Tab", "Reason", "Model", "Serial", "Type", "Detail"]),
    }


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — build_philips_invoice(): write the Excel file
# ══════════════════════════════════════════════════════════════════════════════

def _tf(bold=False, size=10):
    return Font(name="Calibri", bold=bold, size=size)

def _hdr_fill():
    f = PatternFill(fill_type="solid")
    f.fgColor = Color(theme=0, tint=-0.25, type="theme")
    return f

def _total_fill():
    f = PatternFill(fill_type="solid")
    f.fgColor = Color(theme=4, tint=0.8, type="theme")
    return f

def _write_header_row(ws, headers, row=1, col_widths=None):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = _tf(bold=True, size=11)
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

def _line(ws, row, label, uom, price, qty_val, qty_formula=None, total_formula=None):
    ws.cell(row=row, column=1, value=label).font = _tf(size=10)
    if uom:
        ws.cell(row=row, column=2, value=uom).font = _tf(size=10)
    if price is not None:
        c = ws.cell(row=row, column=3, value=price)
        c.font = _tf(size=10); c.number_format = ACCT_FMT
    cd = ws.cell(row=row, column=4, value=qty_formula if qty_formula else qty_val)
    cd.font = _tf(size=10); cd.number_format = INT_FMT
    ce = ws.cell(row=row, column=5, value=total_formula if total_formula else f"=D{row}*C{row}")
    ce.font = _tf(size=10); ce.number_format = ACCT_FMT


def build_philips_invoice(analysis, invoice_title, output_path, log=print):
    # Quantity = raw summed square footage, rounded to the nearest 50-sq-ft increment
    # (the pricing unit) — not divided into a count of blocks.
    demo_add_qty    = round(analysis["demo_additional_sqft"] / SQFT_UNIT) * SQFT_UNIT
    service_add_qty = round(analysis["service_additional_sqft"] / SQFT_UNIT) * SQFT_UNIT
    parts_add_qty   = round(analysis["parts_sqft_manual"] / SQFT_UNIT) * SQFT_UNIT

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Breakdown")
    for col, w in zip("ABCDE", [44.55, 24, 14, 14, 16]):
        ws.column_dimensions[col].width = w

    ws["A1"] = "9145 Ellis Road"
    ws["A2"] = "Melbourne, FL 32904"
    ws["A3"] = "www.ussiglobal.com"
    for r in (1, 2, 3):
        ws.cell(row=r, column=1).font = _tf(size=8)

    ws["A5"] = invoice_title
    ws["A5"].font = _tf(bold=True, size=14)

    def section_hdr(row, label):
        for ci, h in enumerate(["", "Units of Measure", "Unit Price", "Quantity", "Total Amount"], 1):
            c = ws.cell(row=row, column=ci, value=h if ci > 1 else label)
            c.fill = _hdr_fill()
            c.font = _tf(bold=True, size=10)
            c.alignment = Alignment(horizontal="center" if ci > 1 else None)

    row = 7
    section_hdr(row, "Warehouse: Demo"); row += 1
    _line(ws, row, "Warehouse Cost, First 500 sq. ft.", "Each", BASE_PRICE, 1); row += 1
    _line(ws, row, "Additional Square Footage, in 50 sq ft. increments", "Each", SQFT_PRICE, demo_add_qty); row += 2

    section_hdr(row, "Warehouse: Service"); row += 1
    _line(ws, row, "Additional Square Footage, in 50 sq ft. increments", "Each", SQFT_PRICE, service_add_qty); row += 2

    section_hdr(row, "Warehouse: Handling"); row += 1
    inbound_row = row
    _line(ws, row, "Inbound Handling", "Each", HANDLING_FEE, analysis["inbound_count"],
          qty_formula="=COUNTA(Receieved!A2:A100000)"); row += 1
    outbound_row = row
    # Outbound count is defined (both here and in analyze_philips()) as rows
    # with a non-blank Model — so the live formula must COUNTA the Model
    # column (G on the Shipping sheet), not column A ("Date"). Column A can be
    # blank on rows where Model is still populated (and vice versa), so
    # counting on the wrong column would silently drift the recalculated
    # Excel total away from the quantity actually used to price the invoice.
    _line(ws, row, "Outbound Handling", "Each", HANDLING_FEE, analysis["outbound_count"],
          qty_formula="=COUNTA(Shipping!G2:G100000)"); row += 2

    section_hdr(row, "Warehouse: Parts"); row += 1
    _line(ws, row, "Additional Square Footage, in 50 sq ft. increments (manual entry)",
          "Each", SQFT_PRICE, parts_add_qty); row += 2

    section_hdr(row, "Warehouse: Repairs"); row += 1
    repair_row = row
    _line(ws, row, "Repair", "Unit Size/Each", None, analysis["repair_count"],
          qty_formula='=COUNTIFS(Repairs!F:F,"Repaired")',
          total_formula='=SUMIFS(Repairs!J:J,Repairs!F:F,"Repaired")'); row += 1
    harvest_row = row
    _line(ws, row, "Harvest", "Each", None, analysis["harvest_count"],
          qty_formula='=COUNTIFS(Repairs!F:F,"Harvested")',
          total_formula='=SUMIFS(Repairs!J:J,Repairs!F:F,"Harvested")'); row += 2

    last_line = row - 1
    ws.cell(row=row, column=4, value="Subtotal").font = _tf(bold=True)
    ws.cell(row=row, column=5, value=f"=SUM(E8:E{last_line})").number_format = ACCT_FMT
    ws.cell(row=row, column=5).font = _tf(bold=True)
    subtotal_row = row; row += 1
    ws.cell(row=row, column=4, value="Tax").font = _tf(bold=True)
    ws.cell(row=row, column=5, value=f"=E{subtotal_row}*7%").number_format = ACCT_FMT
    ws.cell(row=row, column=5).font = _tf(bold=True)
    tax_row = row; row += 1
    ws.cell(row=row, column=4, value="Total").font = _tf(bold=True)
    tc = ws.cell(row=row, column=5, value=f"=SUM(E{subtotal_row}:E{tax_row})")
    tc.number_format = ACCT_FMT; tc.font = _tf(bold=True); tc.fill = _total_fill()

    # ── Raw data tabs ─────────────────────────────────────────────────────────
    _write_received_tab(wb, analysis["received_df"])
    _write_shipping_tab(wb, analysis["shipping_df"])
    _write_repairs_tab(wb, analysis["repairs_df"])
    _write_excluded_tab(wb, analysis.get("excluded_df"))

    wb.save(output_path)
    log(f"Invoice saved → {Path(output_path).name}")

    subtotal = (
        BASE_PRICE
        + demo_add_qty * SQFT_PRICE
        + service_add_qty * SQFT_PRICE
        + analysis["inbound_count"] * HANDLING_FEE
        + analysis["outbound_count"] * HANDLING_FEE
        + parts_add_qty * SQFT_PRICE
        + analysis["repair_total"]
        + analysis["harvest_total"]
    )
    tax = round(subtotal * TAX_RATE, 2)
    total = round(subtotal + tax, 2)
    return {"subtotal": round(subtotal, 2), "tax": tax, "total": total}


def _write_received_tab(wb, recv):
    ws = wb.create_sheet("Receieved")
    headers = ["Model", "Serial", "Column1", "Warehouse", "Date", "RMA", "Price"]
    _write_header_row(ws, headers, col_widths=[20, 20, 10, 12, 12, 18, 10])
    for i, (_, r) in enumerate(recv.iterrows(), 2):
        vals = [r.get("Model"), r.get("Serial"), r.get("Grade"), r.get("Warehouse"),
                r.get("Date"), r.get("RMA"), HANDLING_FEE]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=v)
            c.font = _tf(size=10)
            if ci == 5 and isinstance(v, (pd.Timestamp, datetime)):
                c.number_format = DATE_FMT
            if ci == 7:
                c.number_format = ACCT_FMT


def _write_shipping_tab(wb, ship):
    ws = wb.create_sheet("Shipping")
    headers = ["Date", "Type", "Departure RMA", "Ship to Name", "Tracking", "Carrier", "Model", "Price"]
    _write_header_row(ws, headers, col_widths=[12, 10, 16, 22, 16, 10, 20, 10])
    for i, (_, r) in enumerate(ship.iterrows(), 2):
        vals = [r.get("Date"), r.get("Stock Level (Primary)"), r.get("Departure RMA"),
                r.get("Ship to Name"), r.get("Departure Tracking"), r.get("Carrier"),
                r.get("Model"), HANDLING_FEE]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=v)
            c.font = _tf(size=10)
            if ci == 1 and isinstance(v, (pd.Timestamp, datetime)):
                c.number_format = DATE_FMT
            if ci == 8:
                c.number_format = ACCT_FMT


def _write_excluded_tab(wb, excluded_df):
    """Every row dropped from a calculation (no Dimensions match, no Size, no
    parseable display size) lands here so nothing silently disappears from the invoice."""
    ws = wb.create_sheet("Excluded Items")
    headers = ["Source Tab", "Reason", "Model", "Serial", "Type", "Detail"]
    _write_header_row(ws, headers, col_widths=[12, 30, 20, 20, 12, 20])
    if excluded_df is None or len(excluded_df) == 0:
        ws.cell(row=2, column=1, value="No excluded items this period.").font = _tf(size=10)
        return
    for i, (_, r) in enumerate(excluded_df.iterrows(), 2):
        for ci, col in enumerate(headers, 1):
            ws.cell(row=i, column=ci, value=r.get(col, "")).font = _tf(size=10)


def _write_repairs_tab(wb, rep):
    ws = wb.create_sheet("Repairs")
    headers = ["Repair Date", "Receive Date", "Model", "Serial", "RMA",
               "Repaired/Harvested", "Repaired Y/N", "Diagnostics", "Parts Used", "Price"]
    _write_header_row(ws, headers, col_widths=[12, 12, 20, 20, 16, 14, 10, 24, 16, 10])
    for i, (_, r) in enumerate(rep.iterrows(), 2):
        yn = "Yes" if r.get("Status") == "Repaired" else "No"
        vals = [r.get("Repair Date"), r.get("Received Date"), r.get("Model"), r.get("Serial"),
                r.get("RMA"), r.get("Status"), yn, r.get("Diagnostics"), r.get("Parts Used"),
                r.get("_price", 0)]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=v)
            c.font = _tf(size=10)
            if ci in (1, 2) and isinstance(v, (pd.Timestamp, datetime)):
                c.number_format = DATE_FMT
            if ci == 10:
                c.number_format = ACCT_FMT
